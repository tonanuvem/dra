""" app-endoscopia-v7
Interface Streamlit para Sistema de Controle de Procedimentos de Endoscopia

Instalar:
    pip install streamlit crewai python-dotenv pypdf openpyxl

Executar:
    streamlit run app_endoscopia.py
"""

import os
import logging
import queue
import threading
import time
from datetime import datetime
from pathlib import Path

import io
import re

import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from docx import Document
from pypdf import PdfReader

# CrewAI
from crewai import Agent, Task, Crew, Process, LLM

# Supabase (auth + dados)
from supabase import create_client as _supa_create_client

# =============================================================================
# RBAC — autenticação e controle de acesso
# =============================================================================

# Perfis permitidos no backend (visualizador usa apenas o frontend Next.js)
_BACKEND_ROLES   = {"editor", "financeiro", "admin"}
_FINANCIAL_ROLES = {"financeiro", "admin"}

# Colunas de valor financeiro que editores não visualizam
_FINANCIAL_COLS = [
    "ValorEstimado_TUSS",
    "ValorLiberado_REPASSE",
    "ValorBase_REPASSE",
]

_ROLE_LABELS = {"editor": "Editor", "financeiro": "Financeiro", "admin": "Admin"}
_ROLE_COLORS = {"editor": "#b45309", "financeiro": "#1d4ed8", "admin": "#b91c1c"}


def _supa_auth():
    """Cliente Supabase com chave publishable — usado só para autenticação."""
    return _supa_create_client(
        os.getenv("SUPABASE_URL", ""),
        os.getenv("SUPABASE_PUBLISHABLE_KEY", ""),
    )


def _supa_service():
    """Cliente Supabase com secret key — leitura de profiles sem RLS."""
    return _supa_create_client(
        os.getenv("SUPABASE_URL", ""),
        os.getenv("SUPABASE_SECRET_KEY") or os.getenv("SUPABASE_PUBLISHABLE_KEY", ""),
    )


def can_view_financial() -> bool:
    """True se o usuário logado pode ver informações financeiras."""
    return st.session_state.get("auth_role") in _FINANCIAL_ROLES


def is_admin() -> bool:
    return st.session_state.get("auth_role") == "admin"


def _strip_financial(df: pd.DataFrame) -> pd.DataFrame:
    """Remove colunas financeiras do DataFrame quando o usuário não tem permissão."""
    if can_view_financial():
        return df
    return df.drop(columns=[c for c in _FINANCIAL_COLS if c in df.columns])


def _logout():
    for k in ["auth_user", "auth_role", "auth_nome", "auth_email"]:
        st.session_state.pop(k, None)
    st.rerun()


def _show_login():
    """Tela de login — visual alinhado ao frontend Next.js (slate-900 + card branco)."""

    # ── CSS global ────────────────────────────────────────────
    st.markdown("""
    <style>
    /* Fundo escuro slate-900 */
    .stApp { background-color: #0f172a !important; }
    [data-testid="stHeader"]    { background-color: #0f172a !important; box-shadow: none !important; }
    [data-testid="stToolbar"],
    [data-testid="stDecoration"],
    #MainMenu, footer           { display: none !important; }

    /* Container centralizado, largura do card */
    .main .block-container {
        max-width: 400px !important;
        padding: 5rem 0 2rem !important;
        margin: 0 auto !important;
    }

    /* Card branco */
    [data-testid="stForm"] {
        background: #ffffff !important;
        border-radius: 16px !important;
        padding: 2rem 2rem 1.5rem !important;
        box-shadow: 0 25px 50px rgba(0, 0, 0, 0.55) !important;
        border: none !important;
    }

    /* Labels dos campos */
    [data-testid="stForm"] .stTextInput label p {
        color: #374151 !important;
        font-size: 14px !important;
        font-weight: 500 !important;
    }

    /* Inputs */
    [data-testid="stForm"] .stTextInput input {
        border: 1px solid #d1d5db !important;
        border-radius: 8px !important;
        font-size: 14px !important;
        color: #111827 !important;
        background: #ffffff !important;
        padding: 10px 12px !important;
    }
    [data-testid="stForm"] .stTextInput input:focus {
        border-color: #3b82f6 !important;
        box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.15) !important;
        outline: none !important;
    }
    [data-testid="stForm"] .stTextInput input::placeholder {
        color: #9ca3af !important;
    }

    /* Botão Entrar */
    [data-testid="stForm"] .stFormSubmitButton > button {
        background-color: #2563eb !important;
        color: #ffffff !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        font-size: 14px !important;
        padding: 0.65rem 1rem !important;
        transition: background-color 0.15s ease !important;
        letter-spacing: 0.01em !important;
    }
    [data-testid="stForm"] .stFormSubmitButton > button:hover {
        background-color: #1d4ed8 !important;
    }
    [data-testid="stForm"] .stFormSubmitButton > button:active {
        background-color: #1e40af !important;
    }

    /* Mensagem de erro abaixo do card */
    [data-testid="stAlert"] {
        border-radius: 8px !important;
        font-size: 14px !important;
        margin-top: 12px !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── Logo (fora do card, sobre o fundo escuro) ─────────────
    st.markdown("""
    <div style="
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 12px;
        margin-bottom: 28px;
        font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
    ">
        <div style="
            width: 48px; height: 48px;
            background: #2563eb;
            border-radius: 12px;
            display: flex; align-items: center; justify-content: center;
            box-shadow: 0 8px 24px rgba(37, 99, 235, 0.45);
            flex-shrink: 0;
        ">
            <span style="font-size: 22px; line-height: 1;">🩺</span>
        </div>
        <div>
            <p style="color:#f1f5f9; font-weight:700; font-size:18px;
                      margin:0; line-height:1.25;">Endoscopia</p>
            <p style="color:#94a3b8; font-size:13px;
                      margin:0; line-height:1.3;">Auditoria de Faturamento</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Formulário (renderizado como card branco via CSS) ─────
    with st.form("login_form", clear_on_submit=False):
        # Título dentro do card
        st.markdown("""
        <div style="
            margin-bottom: 20px;
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
        ">
            <h2 style="font-size:20px; font-weight:700; color:#111827;
                       margin:0 0 4px; padding:0; line-height:1.3;">Entrar</h2>
            <p  style="font-size:14px; color:#6b7280;
                       margin:0; padding:0;">Acesse sua conta para continuar</p>
        </div>
        """, unsafe_allow_html=True)

        email    = st.text_input("E-mail ou CPF",  placeholder="seu@email.com ou 000.000.000-00", label_visibility="visible")
        password = st.text_input("Senha",          placeholder="••••••••", type="password", label_visibility="visible")
        submitted = st.form_submit_button("Entrar", use_container_width=True, type="primary")

        # Rodapé do card
        st.markdown("""
        <p style="
            font-size: 11px; color: #9ca3af; text-align: center;
            margin: 14px 0 0; padding: 0;
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
        ">Acesso restrito — somente usuários autorizados</p>
        """, unsafe_allow_html=True)

    # ── Lógica de autenticação ────────────────────────────────
    if not submitted:
        return

    if not email or not password:
        st.error("Preencha e-mail ou CPF e senha.")
        return

    supabase_url = os.getenv("SUPABASE_URL", "")
    if not supabase_url:
        st.error("SUPABASE_URL não configurado. Verifique o arquivo .env")
        return

    # ── Resolve CPF → e-mail se o input não contiver '@' e tiver 11 dígitos ──
    login_email = email.strip()
    _digitos = "".join(c for c in login_email if c.isdigit())
    if "@" not in login_email and len(_digitos) == 11:
        try:
            _cpf_resp = (
                _supa_service()
                .table("profiles")
                .select("email")
                .eq("cpf", _digitos)
                .maybe_single()
                .execute()
            )
            if not _cpf_resp.data or not _cpf_resp.data.get("email"):
                st.error("CPF não encontrado. Verifique o número ou use seu e-mail.")
                return
            login_email = _cpf_resp.data["email"]
        except Exception as _cpf_exc:
            st.error(f"Erro ao buscar CPF: {_cpf_exc}")
            return

    try:
        resp = _supa_auth().auth.sign_in_with_password(
            {"email": login_email, "password": password}
        )
        user = resp.user

        # Busca role e status no profiles (usa service_role para bypassar RLS)
        profile_resp = (
            _supa_service()
            .table("profiles")
            .select("role, nome, ativo")
            .eq("id", str(user.id))
            .single()
            .execute()
        )
        profile = profile_resp.data or {}
        role  = profile.get("role", "visualizador")
        ativo = profile.get("ativo", True)

        if not ativo:
            st.error("⛔ Conta inativa. Contate o administrador.")
            return

        if role not in _BACKEND_ROLES:
            st.error("⛔ Perfil 'Visualizador' não tem acesso a esta interface.")
            return

        st.session_state["auth_user"]  = str(user.id)
        st.session_state["auth_email"] = user.email or login_email
        st.session_state["auth_nome"]  = profile.get("nome") or user.email or email
        st.session_state["auth_role"]  = role
        st.rerun()

    except Exception as exc:
        msg = str(exc)
        if "Invalid login credentials" in msg or "invalid_credentials" in msg:
            st.error("E-mail/CPF ou senha incorretos.")
        else:
            st.error(f"Erro ao autenticar: {msg}")


# =============================================================================
# CONFIGURAÇÃO DE VARIÁVEIS DE AMBIENTE E LOG
# =============================================================================

# Carrega .env da raiz do projeto (compartilhado com frontend)
load_dotenv(Path(__file__).parent.parent / ".env")
# Fallback para .env local do backend
load_dotenv()

def _resolve_log_path() -> Path:
    candidates = [Path("logs"), Path("/tmp/endoscopia_logs")]
    for candidate in candidates:
        try:
            candidate.mkdir(parents=True, exist_ok=True)
            test_file = candidate / ".write_test"
            test_file.touch()
            test_file.unlink()
            return candidate
        except OSError:
            continue
    return Path("/tmp")

log_dir = _resolve_log_path()

_log_handlers: list[logging.Handler] = [logging.StreamHandler()]
try:
    _log_handlers.insert(0, logging.FileHandler(log_dir / "crew_logs.log", encoding="utf-8"))
except OSError:
    pass

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=_log_handlers,
    force=True,
)
logger = logging.getLogger(__name__)


# =============================================================================
# VALORES PADRÃO DOS AGENTES E TASKS
# =============================================================================

ANALISTA_DEFAULTS = {
    "role": "Analista de Controle dos procedimentos de Endoscopia",
    "goal": """
- Retornar APENAS o CSV puro, sem markdown, sem explicações.
- Primeira linha obrigatoriamente o cabeçalho:
- Padronizar Data para DD/MM/AAAA.
- Padronizar nomes de paciente em MAIÚSCULAS sem espaços extras.
- Identificar o tipo pelo nome da planilha: PRODUCAO ou REPASSE
- Coluna TipoArquivo = "PRODUCAO" ou "REPASSE".

Desmembramento de Procedimentos Adicionais: 
- Se houver preenchimento na coluna "PROCEDIMENTOS ADICIONAIS" (ex: TESTE DE UREASE), você deve gerar uma LINHA EXTRA no CSV para cada procedimento listado.
- Na linha extra, o nome do procedimento adicional deve entrar na coluna Procedimento. As demais informações (Data, Atendimento, Paciente, Convênio, etc.) devem ser replicadas da linha principal.
- Para essas linhas extras, inclua obrigatoriamente a tag PROCEDIMENTO_ADICIONAL na coluna Observacao. A linha do procedimento principal deve ter a Observacao vazia.
""",
    "backstory": """
Expert em análise de faturamento hospitalar e auditoria de convênios médicos.
Especializado na Terminologia Unificada da Saúde Suplementar (TUSS) e nas regras de
faturamento dos principais convênios do Brasil.
Conhece em profundidade os campos das planilhas de produtividade de equipes de enfermagem
e os arquivos de repasse hospitalares (formato Hospital São Camilo e similares).
Sabe que o número de atendimento pode divergir entre os dois arquivos para o mesmo
paciente/procedimento, portanto não usa esse campo como chave única.
Seu principal objetivo é estruturar os dados com máxima fidelidade para que o agente
correlacionador consiga fazer o batimento correto.
    """,
    "task_description_template": """
Retorne APENAS o CSV puro, sem markdown, sem explicações.
Cabeçalho obrigatório na primeira linha
Não omita nenhuma linha — cada procedimento registrado deve constar no CSV gerado.
=== CONTEÚDO DO ARQUIVO ===
{conteudo_arquivo}
    """,
    "task_expected_output": (
        "CSV puro com cabeçalho na primeira linha, todas as linhas do arquivo estruturadas e coluna TipoArquivo preenchida com PRODUCAO ou REPASSE."
    ),
}

CORRELACIONADOR_DEFAULTS = {
    "role": "Correlacionador dos registros de endoscopia",
    "goal": """
Receber os CSVs padronizados gerados pelo Analista — um do tipo PRODUCAO e outro do tipo REPASSE — e realizar o Batimento Automático linha a linha, considerando as seguintes regras:

- A primeira linha deve ser obrigatoriamente o cabeçalho, indicando de qual planilha veio aquela coluna (acrescentar sufixo _PRODUCAO ou _REPASSE)
- NÃO remover nenhuma linha da PRODUCAO; Ordenar por Data crescente; Padronizar Data para DD/MM/AAAA.
- Padronizar ValorLiberado: ponto decimal, sem símbolo de moeda. 

- A saída do CSV_CORRELACAO deve conter todas as linhas do CSV de PRODUÇÃO como a base da sua resposta, onde devem ser incluídas as novas colunas com o relacionamento do CSV de REPASSE, sendo que para cada linha da PRODUCAO, deve tentar encontrar correspondência no REPASSE usando a chave composta com 3 campos (data + paciente + procedimento). Devem ser incluídas em cada linha somente as colunas do CSV de REPASSE que ainda não exista no CSV de PRODUÇÃO, e também deve ser incluída uma nova coluna de StatusCorrelacao, considerando a seguinte regra:
   a. Se ValorLiberado coincidir (ou REPASSE não tiver valor): StatusCorrelacao = "CORRELACIONADO"
   b. Se ValorLiberado for menor no REPASSE: StatusCorrelacao = "CORRELACIONADO_COM_DIVERGENCIA_VALOR"
   c. Se ValorLiberado = 0 no REPASSE: StatusCorrelacao = "CORRELACIONADO" (glosa total — ver ValorLiberado_REPASSE)
   d. Se ValorLiberado > 0 mas menor que esperado: StatusCorrelacao = "CORRELACIONADO" (glosa parcial — ver ValorLiberado_REPASSE)
   e. Se NÃO encontrar correspondência no REPASSE: StatusCorrelacao = "NAO_FATURADO_NO_REPASSE"

- Para linhas do REPASSE sem correspondência na PRODUCAO, inserir no final do arquivo CSV_CORRELACAO as linhas como valores zerados do PRODUCAO e os valores existentes da PRODUÇÃO, e com nova coluna StatusCorrelacao = "REPASSE_NAO_IDENTIFICADO_NA_PRODUCAO"

- Para fazer esse relacionamento, deve-se usar a chave de correlação COMPOSTA pelos 3 campos a seguir:
  1. Data (DD/MM/AAAA) — tolerância de ± 1 dia para o mesmo episódio
  2. Paciente (nome normalizado em maiúsculas)
  3. Procedimento (correspondência semântica: "COLONOSCOPIA" bate com "Colonoscopia (Inclui A Retossigmoidoscopia)", "TESTE DE UREASE" bate com "Pesquisa de H. pylori - Teste da Urease", etc.)
    """,
    "backstory": """
Especialista em auditoria de faturamento médico, ETL e reconciliação de dados hospitalares.
Usa correspondência semântica para não perder correlações válidas.
Nunca inventa procedimentos — apenas padroniza, correlaciona e sinaliza divergências.
Seu trabalho alimenta diretamente a cobrança estruturada junto ao hospital e convênios.
    """,
    "task_description_template": """
Você recebeu os seguintes blocos CSV padronizados pelo Analista de Endoscopia, identificados como PRODUCAO e REPASSE. Execute o batimento automático conforme as regras do seu objetivo (goal), usando chave composta de Data + Paciente + Procedimento.
- Retorne APENAS o CSV correlacionado
- A primeira linha deve ser obrigatoriamente o cabeçalho
=== CONTEÚDO ===
{blocos}
    """,
    "task_expected_output": (
        "CSV puro com cabeçalho na primeira linha, todas as linhas da PRODUCAO como base, colunas complementares do REPASSE adicionadas e coluna StatusCorrelacao preenchida em cada linha."
    ),
}


def _init_agent_session_state():
    """
    Inicializa o session_state com os valores padrão dos agentes,
    caso ainda não existam (chamado uma vez no início do main).
    """
    if "analista_cfg" not in st.session_state:
        st.session_state["analista_cfg"] = dict(ANALISTA_DEFAULTS)
    if "correlacionador_cfg" not in st.session_state:
        st.session_state["correlacionador_cfg"] = dict(CORRELACIONADOR_DEFAULTS)


def _get_analista_cfg() -> dict:
    return st.session_state.get("analista_cfg", ANALISTA_DEFAULTS)


def _get_correlacionador_cfg() -> dict:
    return st.session_state.get("correlacionador_cfg", CORRELACIONADOR_DEFAULTS)


# =============================================================================
# HANDLER DE LOG EM TEMPO REAL PARA O STREAMLIT
# =============================================================================

class StreamlitLogHandler(logging.Handler):
    def __init__(self, log_queue: queue.Queue):
        super().__init__()
        self.log_queue = log_queue

    def emit(self, record: logging.LogRecord):
        try:
            self.log_queue.put_nowait(self.format(record))
        except queue.Full:
            pass


def render_log_line(line: str, container) -> None:
    line_lower = line.lower()
    if any(k in line_lower for k in ("error", "erro", "exception", "traceback")):
        container.error(f"🔴 {line}")
    elif any(k in line_lower for k in ("warning", "warn", "aviso")):
        container.warning(f"🟡 {line}")
    elif any(k in line_lower for k in ("task", "tarefa", "iniciando", "starting")):
        container.info(f"📋 {line}")
    elif any(k in line_lower for k in ("agent", "agente", "thinking", "pensando")):
        container.info(f"🤖 {line}")
    elif any(k in line_lower for k in ("action", "ação", "tool", "ferramenta")):
        container.info(f"⚙️ {line}")
    elif any(k in line_lower for k in ("final answer", "resposta final", "finished", "concluído", "complete")):
        container.success(f"✅ {line}")
    else:
        container.text(f"   {line}")


# =============================================================================
# FUNÇÕES DE LEITURA DE ARQUIVOS
# =============================================================================

def read_word_file(file) -> str:
    try:
        doc = Document(file)
        text = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(text)
    except Exception as e:
        st.error(f"Erro ao ler arquivo Word: {e}")
        return None


def read_pdf_file(file) -> str:
    try:
        pdf_reader = PdfReader(file)
        text = [
            page.extract_text()
            for page in pdf_reader.pages
            if page.extract_text() and page.extract_text().strip()
        ]
        return "\n\n".join(text)
    except Exception as e:
        st.error(f"Erro ao ler arquivo PDF: {e}")
        return None


def read_text_file(file) -> str:
    try:
        file.seek(0)
        raw = file.read()
        if isinstance(raw, bytes):
            return raw.decode("utf-8", errors="ignore")
        return str(raw)
    except Exception as e:
        st.error(f"Erro ao ler arquivo de texto: {e}")
        return None


def read_excel_file(file) -> str:
    try:
        file.seek(0)
        # Identifica a extensão do arquivo
        file_extension = file.name.split('.')[-1].lower()
        
        # Define o engine correto: xlrd para .xls e openpyxl para .xlsx
        engine = "openpyxl" if file_extension == "xlsx" else "xlrd"
        
        # Lê o arquivo Excel com o engine apropriado
        xls = pd.ExcelFile(file, engine=engine)
        
        blocos = []
        for sheet_name in xls.sheet_names:
            df = xls.parse(sheet_name, dtype=str)
            df.fillna("", inplace=True)
            df.dropna(axis=1, how="all", inplace=True)
            df.dropna(axis=0, how="all", inplace=True)
            
            if df.empty:
                continue
                
            # Remove quebras de linha internas em células
            df = df.apply(lambda col: col.map(
                lambda v: re.sub(r"[\r\n]+", " ", str(v)).strip() if isinstance(v, str) else v
            ))
            
            csv_str = df.to_csv(index=False, sep=",", encoding="utf-8")
            blocos.append(f"=== ABA: {sheet_name} ===\n{csv_str}")
            
        if not blocos:
            st.warning("O arquivo Excel não contém dados nas abas.")
            return None
            
        return "\n\n".join(blocos)
    except Exception as e:
        st.error(f"Erro ao ler arquivo Excel (.{file_extension}): {e}")
        return None


def extract_text_from_file(uploaded_file) -> str:
    if uploaded_file is None:
        return None
    extension = uploaded_file.name.rsplit(".", 1)[-1].lower()
    uploaded_file.seek(0)
    readers = {
        "docx": read_word_file,
        "pdf": read_pdf_file,
        "txt": read_text_file,
        "csv": read_text_file,
        "xlsx": read_excel_file,
        "xls": read_excel_file,
    }
    reader = readers.get(extension)
    if reader:
        return reader(uploaded_file)
    st.error(f"Formato de arquivo não suportado: .{extension}")
    return None


# =============================================================================
# FUNÇÕES DE TRANSFORMAÇÃO DE ARQUIVOS
# =============================================================================

# =============================================================================
# CONSTANTES
# =============================================================================

# Padrão da linha de título informativo (descartada)
_RE_LINHA_INFORMATIVA = re.compile(
    r"CONTROLE DE EXAMES E PROCEDIMENTOS NA ENDOSCOPIA",
    re.IGNORECASE,
)

# Colunas canônicas do formato PRODUCAO com QTD (2026) — 13 campos
_COLUNAS_COM_QTD = [
    "QTD", "Data", "Paciente", "NrAtendimento",
    "Convenio", "Origem", "Procedimento",
    "ProcedimentosAdicionais", "MedicoExecutor",
    "LocalSetor", "Sala", "Carater", "Observacao",
]

# Colunas canônicas do formato PRODUCAO sem QTD (legado 2025) — 13 campos
# O primeiro campo é Data (sem QTD), e há EXAME REALIZADO 1 + EXAME REALIZADO 2
_COLUNAS_SEM_QTD = [
    "Data", "Paciente", "NrAtendimento",
    "Convenio", "Origem", "Procedimento", "Procedimento2",
    "ProcedimentosAdicionais", "MedicoExecutor",
    "LocalSetor", "Sala", "Carater", "Observacao",
]

# Colunas canônicas do formato PRODUCAO 2025 NOVO (MAIO+ 2025) — 12 campos
# Sem QTD, sem EXAME REALIZADO 1/2, apenas EXAME REALIZADO único
_COLUNAS_2025_NOVO = [
    "Data", "Paciente", "NrAtendimento",
    "Convenio", "Origem", "Procedimento",
    "ProcedimentosAdicionais", "MedicoExecutor",
    "LocalSetor", "Sala", "Carater", "Observacao",
]

# =============================================================================
# MAPEAMENTOS DE COLUNAS → NOMES PADRONIZADOS
# =============================================================================

_MAP_PRODUCAO = {
    "qtd":                                   "QTD",
    "data":                                  "Data",
    "nome do paciente":                      "Paciente",
    "nº atendimento":                        "NrAtendimento",
    "n° atendimento":                        "NrAtendimento",
    "n atendimento":                         "NrAtendimento",
    "nr atendimento":                        "NrAtendimento",
    "convênio":                              "Convenio",
    "convenio":                              "Convenio",
    "origem":                                "Origem",
    "exame realizado":                       "Procedimento",      # com ou sem espaço trailing
    "exame realizado 1":                     "Procedimento",
    "exame realizado 2":                     "Procedimento2",
    "procedimento adicional":                "ProcedimentosAdicionais",
    "procedimentos adicionais":              "ProcedimentosAdicionais",
    "procedimento adocional":                "ProcedimentosAdicionais",  # typo real do arquivo
    "médico executor":                       "MedicoExecutor",
    "medico executor":                       "MedicoExecutor",
    "local / setor":                         "LocalSetor",
    "local/setor":                           "LocalSetor",
    "sala":                                  "Sala",
    "carater":                               "Carater",
    "caráter":                               "Carater",
    "observação + procedimentos adicionais": "Observacao",
    "observacao":                            "Observacao",
    "observação":                            "Observacao",
}

# Mapeamento para formato 2025 NOVO (MAIO+ 2025) - sem QTD, sem EXAME REALIZADO 1/2
_MAP_PRODUCAO_2025_NOVO = {
    "data":                                  "Data",
    "nome do paciente":                      "Paciente",
    "nº atendimento":                        "NrAtendimento",
    "n° atendimento":                        "NrAtendimento",
    "n atendimento":                         "NrAtendimento",
    "nr atendimento":                        "NrAtendimento",
    "convênio":                              "Convenio",
    "convenio":                              "Convenio",
    "origem":                                "Origem",
    "exame realizado":                       "Procedimento",
    "procedimentos adicionais":              "ProcedimentosAdicionais",
    "procedimento adicional":                "ProcedimentosAdicionais",
    "médico executor":                       "MedicoExecutor",
    "medico executor":                       "MedicoExecutor",
    "local / setor":                         "LocalSetor",
    "local/setor":                           "LocalSetor",
    "sala":                                  "Sala",
    "carater":                               "Carater",
    "caráter":                               "Carater",
    "observação + procedimentos adicionais": "Observacao",
    "observacao":                            "Observacao",
    "observação":                            "Observacao",
}

# Mapeamento para formato 2026 (com QTD no início)
_MAP_PRODUCAO_2026 = {
    "qtd":                                   "QTD",
    "data":                                  "Data",
    "nome do paciente":                      "Paciente",
    "nº atendimento":                        "NrAtendimento",
    "n° atendimento":                        "NrAtendimento",
    "n atendimento":                         "NrAtendimento",
    "nr atendimento":                        "NrAtendimento",
    "convênio":                              "Convenio",
    "convenio":                              "Convenio",
    "origem":                                "Origem",
    "exame realizado":                       "Procedimento",
    "procedimento":                          "Procedimento",
    "procedimentos adicionais":              "ProcedimentosAdicionais",
    "procedimento adicional":                "ProcedimentosAdicionais",
    "médico executor":                       "MedicoExecutor",
    "medico executor":                       "MedicoExecutor",
    "local / setor":                         "LocalSetor",
    "local/setor":                           "LocalSetor",
    "sala":                                  "Sala",
    "carater":                               "Carater",
    "caráter":                               "Carater",
    "observação + procedimentos adicionais": "Observacao",
    "observacao":                            "Observacao",
    "observação":                            "Observacao",
}

_MAP_REPASSE = {
    "ds estabelecimento": "Estabelecimento",
    "cnpj estabelecimento": "CNPJ",
    "ds terceiro": "Terceiro",
    "ds status": "Status",
    "nr repasse terceiro": "NrRepasse",
    "tipo": "TipoItem",
    "nr atendimento": "NrAtendimento",
    "tp atend": "TipoAtendimento",
    "nr interno conta": "NrInternoConta",
    "paciente": "Paciente",
    "convenio": "Convenio",
    "convênio": "Convenio",
    "ds categoria": "Categoria",
    "cód item tuss": "CodigoTUSS",
    "cod item tuss": "CodigoTUSS",
    "ds procedimento": "Procedimento",
    "via unid med": "Via",
    "nm medico executor": "MedicoExecutor",
    "médico executor": "MedicoExecutor",
    "porcentagem": "Porcentagem",
    "ds funcao": "Funcao",
    "ds especialidade": "Especialidade",
    "qt procedimento": "QtProcedimento",
    "dt procedimento": "Data",
    "vl liberado": "ValorLiberado",
    "valor liberado": "ValorLiberado"
}

# Conjunto de todas as colunas canônicas do REPASSE (derivado automaticamente do mapa)
# Usado para garantir que todas as colunas estejam presentes após o processamento
_COLUNAS_CANONICAS_REPASSE = set(_MAP_REPASSE.values())

# =============================================================================
# FUNÇÕES AUXILIARES PRIVADAS
# =============================================================================

def _normalizar_coluna(col: str) -> str:
    return re.sub(r"\s+", " ", col.strip()).lower()


def _renomear_colunas(df: pd.DataFrame, mapa: dict) -> pd.DataFrame:
    novo = {
        col: mapa[_normalizar_coluna(col)]
        for col in df.columns
        if _normalizar_coluna(col) in mapa
    }
    return df.rename(columns=novo)


def _detectar_formato_producao(df: pd.DataFrame, nome_aba: str = "") -> str:
    """
    Detecta o formato da planilha baseado nas colunas e nome da aba.
    Retorna: '2025_LEGADO' (JAN-ABR 2025), '2025_NOVO' (MAI+ 2025), '2026'
    """
    colunas_norm = [_normalizar_coluna(col) for col in df.columns]
    aba_upper = nome_aba.upper()
    
    # Formato 2025 LEGADO: tem "EXAME REALIZADO 1" e "EXAME REALIZADO 2"
    if "exame realizado 1" in colunas_norm or "exame realizado 2" in colunas_norm:
        return "2025_LEGADO"
    
    # Formato 2026: tem coluna QTD no início
    if "qtd" in colunas_norm:
        return "2026"
    
    # Formato 2025 NOVO (a partir de MAIO): sem QTD, sem EXAME REALIZADO 1/2
    # Tem apenas "EXAME REALIZADO" e "PROCEDIMENTOS ADICIONAIS"
    if "exame realizado" in colunas_norm and "procedimentos adicionais" in colunas_norm:
        return "2025_NOVO"
    
    # Fallback: tenta detectar pelo nome da aba
    if "2025" in aba_upper:
        # Se é JAN-ABR 2025, provavelmente é legado
        meses_legado = ["JANEIRO", "FEVEREIRO", "MARÇO", "MARCO", "ABRIL"]
        if any(mes in aba_upper for mes in meses_legado):
            return "2025_LEGADO"
        return "2025_NOVO"
    
    return "2026"


def _padronizar_data(valor: str) -> str:
    """
    Converte qualquer formato de data reconhecível para DD/MM/AAAA.

    Formatos suportados (em ordem de tentativa):
      1. Serial numérico do Excel  — ex: 45344  → 22/02/2024
         Origem: arquivos de REPASSE exportados diretamente do Excel sem
         formatação de célula, onde a data fica como inteiro (dias desde
         30/12/1899, conforme convenção do Excel/Lotus).
         Faixa válida: 40000–50000 (aprox. 2009–2036) para evitar
         confundir com NrAtendimento ou outros números de 5 dígitos.
      2. ISO 8601  — ex: 2024-04-24 ou 2024-04-24 00:00:00
      3. Demais formatos reconhecidos pelo pandas (dayfirst=True)
    """
    if pd.isna(valor):
        return ""

    s = str(valor).strip()

    try:
        # 1. Serial numérico do Excel (5 dígitos na faixa 40000–50000)
        if re.match(r'^\d{5}$', s) and 40000 <= int(s) <= 50000:
            from datetime import date, timedelta
            dt = date(1899, 12, 30) + timedelta(days=int(s))
            return dt.strftime("%d/%m/%Y")

        # 2. ISO 8601 (YYYY-MM-DD ou YYYY-MM-DD HH:MM:SS)
        if re.match(r'^\d{4}-\d{2}-\d{2}', s):
            dt = pd.to_datetime(s, format='ISO8601')
            return dt.strftime("%d/%m/%Y")

        # 3. Outros formatos reconhecidos pelo pandas
        dt = pd.to_datetime(s, dayfirst=True)
        return dt.strftime("%d/%m/%Y")

    except Exception:
        return s


def _corrigir_data_malformada(data: str, aba_origem: str) -> str:
    """
    Corrige datas malformadas usando o mês da AbaOrigemDados como referência.
    Exemplos:
    - 18/032025 + ABA: MARÇO 2025 → 18/03/2025
    - 170/2025 + ABA: MAIO 2025 → 17/05/2025
    - 2706/2025 + ABA: JUNHO 2025 → 27/06/2025
    """
    if not data or not aba_origem:
        return data
    
    # Extrai mês e ano da AbaOrigemDados
    meses = {
        "JANEIRO": "01", "FEVEREIRO": "02", "MARÇO": "03", "MARCO": "03",
        "ABRIL": "04", "MAIO": "05", "JUNHO": "06",
        "JULHO": "07", "AGOSTO": "08", "SETEMBRO": "09",
        "OUTUBRO": "10", "NOVEMBRO": "11", "DEZEMBRO": "12"
    }
    
    mes_ref = None
    ano_ref = None
    aba_upper = aba_origem.upper()
    
    for nome_mes, num_mes in meses.items():
        if nome_mes in aba_upper:
            mes_ref = num_mes
            # Extrai ano (4 dígitos)
            match_ano = re.search(r'20\d{2}', aba_upper)
            if match_ano:
                ano_ref = match_ano.group()
            break
    
    if not mes_ref or not ano_ref:
        return data
    
    # Padrões de datas malformadas
    data_limpa = data.replace("/", "").replace("-", "").strip()
    
    # Padrão: 18032025 (sem separadores)
    if len(data_limpa) == 8 and data_limpa.isdigit():
        dia = data_limpa[:2]
        mes = data_limpa[2:4]
        ano = data_limpa[4:]
        if mes == mes_ref:  # Valida se o mês bate
            return f"{dia}/{mes}/{ano}"
    
    # Padrão: 170/2025 (dia com 3 dígitos)
    if re.match(r'^\d{3}/\d{4}$', data):
        dia = data[:2]  # Pega os 2 primeiros dígitos
        ano = data[-4:]
        return f"{dia}/{mes_ref}/{ano}"
    
    # Padrão: 2706/2025 (dia+mes sem separador)
    if re.match(r'^\d{4}/\d{4}$', data):
        dia = data[:2]
        mes = data[2:4]
        ano = data[-4:]
        if mes == mes_ref:  # Valida se o mês bate
            return f"{dia}/{mes}/{ano}"
    
    # Padrão: 17/042025 (ano sem separador)
    if re.match(r'^\d{2}/\d{6}$', data):
        dia = data[:2]
        mes = data[3:5]
        ano = data[5:]
        if mes == mes_ref:  # Valida se o mês bate
            return f"{dia}/{mes}/{ano}"
    
    return data


def _padronizar_valor(valor: str) -> str:
    """Remove R$, espaços e normaliza separadores decimais."""
    s = str(valor).strip() if valor else ""
    if not s:
        return ""
    s = re.sub(r"[R$\s]", "", s)
    if re.search(r"\d\.\d{3},", s):      # 1.234,56 → 1234.56
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")           # 664,61 → 664.61
    return re.sub(r"[,.]$", "", s)


def _identificar_tipo_arquivo(nome_aba: str, nome_arquivo: str = "") -> str:
    """
    Retorna 'REPASSE' ou 'PRODUCAO'.
    Prioridade: nome_aba → nome_arquivo → fallback PRODUCAO.
    """
    if "REPASSE" in nome_aba.upper():
        return "REPASSE"
    if "REPASSE" in nome_arquivo.upper():
        return "REPASSE"
    return "PRODUCAO"


def _detectar_tipo_por_cabecalho(linhas: list[str]) -> str:
    """
    Detecta o tipo (PRODUCAO ou REPASSE) inspecionando as primeiras linhas.
    Usa colunas exclusivas do REPASSE como indicadores.
    """
    _COLS_REPASSE = {
        "ds estabelecimento", "ds terceiro", "nr repasse terceiro",
        "ds procedimento", "nm medico executor", "dt procedimento",
        "vl liberado", "cód item tuss", "ds categoria",
    }
    for linha in linhas[:3]:
        cols = {c.strip().lower() for c in linha.split(",")}
        if cols & _COLS_REPASSE:
            return "REPASSE"
    return "PRODUCAO"


def _linha_e_valida(linha: str) -> bool:
    """
    Descarta:
    - linhas em branco ou só vírgulas/espaços
    - linhas com apenas 1 valor não vazio (ex: "429,,,,,,,,,,,")
    - linha de título informativo ("CONTROLE DE EXAMES...")
    - linha com colunas todas 'Unnamed: N'
    """
    s = linha.strip()
    if not s or re.match(r"^[,\s]*$", s):
        return False
    nao_vazios = [p.strip() for p in s.split(",") if p.strip()]
    if len(nao_vazios) <= 1:
        return False
    if _RE_LINHA_INFORMATIVA.search(s):
        return False
    if all(re.match(r"^unnamed\s*:\s*\d+$", v, re.I) for v in nao_vazios):
        return False
    
    return True


def _detectar_cabecalho_producao(linhas: list[str]) -> tuple[str, int]:
    """
    Inspeciona as primeiras linhas válidas e retorna:
      ("COM_QTD", idx)   — cabeçalho explícito com a coluna QTD
      ("SEM_QTD", idx)   — cabeçalho explícito sem QTD (formato 2025)
      ("SEM_HEADER", 0)  — sem cabeçalho; dados começam na linha 0

    idx = índice em `linhas` onde está o cabeçalho (ou 0 se não houver).
    """
    for idx, linha in enumerate(linhas[:5]):
        upper = linha.upper()
        campos = [c.strip() for c in linha.split(",")]

        # Cabeçalho 2026 com QTD
        if "QTD" in campos[0].upper() and "DATA" in upper:
            return "COM_QTD", idx

        # Cabeçalho 2025 sem QTD — começa com DATA
        if campos[0].strip().upper() in ("DATA",) and "NOME DO PACIENTE" in upper:
            return "SEM_QTD", idx

    return "SEM_HEADER", 0


def _atribuir_colunas(df: pd.DataFrame, formato: str, nome_aba: str = "") -> pd.DataFrame:
    """
    Atribui cabeçalho canônico a DataFrames sem cabeçalho explícito.
    """
    n = len(df.columns)
    
    # Detecta qual base usar baseado no número de colunas e nome da aba
    if n == 12:
        # 12 colunas = formato 2025 NOVO (MAIO+ 2025)
        base = _COLUNAS_2025_NOVO
    elif n == 13:
        # 13 colunas: pode ser COM_QTD (2026) ou SEM_QTD (2025 LEGADO)
        # Verifica se primeira coluna parece QTD
        amostra = df.iloc[:, 0].dropna().head(5).tolist()
        primeiro_parece_qtd = all(
            re.match(r"^\d*$", str(v).strip()) for v in amostra if str(v).strip()
        )
        base = _COLUNAS_COM_QTD if primeiro_parece_qtd else _COLUNAS_SEM_QTD
    else:
        # Fallback: tenta detectar pelo nome da aba
        if "2025" in nome_aba.upper():
            meses_legado = ["JANEIRO", "FEVEREIRO", "MARÇO", "MARCO", "ABRIL"]
            if any(mes in nome_aba.upper() for mes in meses_legado):
                base = _COLUNAS_SEM_QTD
            else:
                base = _COLUNAS_2025_NOVO
        else:
            base = _COLUNAS_COM_QTD

    df.columns = (base + [f"Extra_{k}" for k in range(n - len(base))])[:n] if n <= len(base) else \
                 base + [f"Extra_{k}" for k in range(n - len(base))]
    return df


# =============================================================================
# PROCESSADORES POR TIPO
# =============================================================================

def _processar_aba_producao(df: pd.DataFrame, nome_aba: str) -> pd.DataFrame:
    """
    Normaliza um DataFrame de PRODUCAO:
    - Detecta formato (2025_LEGADO, 2025_NOVO, 2026)
    - Renomeia colunas pelo mapa apropriado
    - Garante QTD (0 se ausente), Data, Paciente, Procedimento
    - Gera linhas extras para 'Procedimento2' (EXAME REALIZADO 2 das abas 2025 LEGADO)
    - Adiciona TipoArquivo = 'PRODUCAO' e AbaOrigemDados
    - Descarta linhas sem Paciente válido
    """
    # Detecta formato baseado nas colunas e nome da aba
    formato = _detectar_formato_producao(df, nome_aba)
    
    # Seleciona mapa apropriado
    if formato == "2025_LEGADO":
        mapa = _MAP_PRODUCAO
    elif formato == "2025_NOVO":
        mapa = _MAP_PRODUCAO_2025_NOVO
    else:  # 2026
        mapa = _MAP_PRODUCAO_2026
    
    # CORREÇÃO: Remove primeira coluna se for Unnamed e formato não é 2026
    # Isso acontece quando pandas lê cabeçalho sem QTD mas dados têm coluna extra
    if formato != "2026" and len(df.columns) > 0:
        primeira_col = str(df.columns[0]).lower()
        if "unnamed" in primeira_col or primeira_col.strip() in ("", "nan"):
            df = df.iloc[:, 1:]  # Remove primeira coluna
    
    df = _renomear_colunas(df, mapa)
    
    # CORREÇÃO: Limpa valores "Unnamed: N" dos dados (problema do Excel)
    for col in df.columns:
        df[col] = df[col].apply(
            lambda v: "" if str(v).strip().upper().startswith("UNNAMED:") else v
        )

    for col in ("QTD", "Data", "Paciente", "Procedimento", "Procedimento2",
                "Convenio", "Origem", "NrAtendimento", "MedicoExecutor",
                "LocalSetor", "Sala", "Carater", "Observacao",
                "ProcedimentosAdicionais"):
        if col not in df.columns:
            df[col] = ""

    # Normaliza QTD
    df["QTD"] = df["QTD"].apply(
        lambda v: "0" if str(v).strip() in ("", "nan", "NaN") else str(v).strip()
    )

    # Primeiro aplica padronização básica
    df["Data"] = df["Data"].apply(_padronizar_data)
    
    # Depois corrige datas malformadas usando AbaOrigemDados
    df["Data"] = df.apply(
        lambda row: _corrigir_data_malformada(row["Data"], nome_aba),
        axis=1
    )
    
    # Tenta padronizar novamente após correção
    df["Data"] = df["Data"].apply(_padronizar_data)
    
    df["Paciente"]     = df["Paciente"].apply(
        lambda v: re.sub(r"\s+", " ", str(v).strip()).upper()
    )
    df["Procedimento"] = df["Procedimento"].apply(
        lambda v: str(v).strip().upper() if str(v).strip() not in ("", "-", "nan") else ""
    )

    linhas_saida = []
    for _, row in df.iterrows():
        paciente = str(row.get("Paciente", "")).strip()
        # Descarta linhas de formatação sem paciente real
        if not paciente or paciente in ("NAN", ""):
            continue

        linha_principal = row.copy()
        linha_principal["TipoArquivo"]    = "PRODUCAO"
        linha_principal["AbaOrigemDados"] = f"ABA: {nome_aba}"
        linhas_saida.append(linha_principal)

        # Trata Procedimento2 (EXAME REALIZADO 2 das abas 2025 LEGADO) como linha extra
        proc2 = str(row.get("Procedimento2", "")).strip().upper()
        if proc2 and proc2 not in ("", "-", "NAN", "NONE"):
            extra = row.copy()
            extra["Procedimento"]    = proc2
            extra["Procedimento2"]   = ""
            extra["Observacao"]      = "PROCEDIMENTO_ADICIONAL"
            extra["TipoArquivo"]     = "PRODUCAO"
            extra["AbaOrigemDados"]  = f"ABA: {nome_aba}"
            extra["QTD"]             = "0"
            linhas_saida.append(extra)

    df_saida = pd.DataFrame(linhas_saida)
    df_saida.drop(columns=["Procedimento2"], inplace=True, errors="ignore")
    # Remove colunas lixo: Extra_N e duplicatas pandas (sufixo .1, .2, ...)
    lixo = [c for c in df_saida.columns
            if re.match(r"^Extra_\d+$", str(c)) or re.search(r"\.\d+$", str(c))]
    df_saida.drop(columns=lixo, inplace=True, errors="ignore")
    return df_saida


def _processar_aba_repasse(df: pd.DataFrame, nome_aba: str) -> pd.DataFrame:
    """
    Normaliza um DataFrame de REPASSE:
    - Renomeia colunas pelo mapa
    - Limpa valores "Unnamed: N" gerados pelo pandas ao ler Excel
    - Garante que TODAS as colunas canonicas do _MAP_REPASSE existam (vazias se ausentes)
    - Padroniza Data, Paciente e ValorLiberado
    - Adiciona TipoArquivo = 'REPASSE' e AbaOrigemDados
    - Remove colunas lixo: Extra_N e duplicatas pandas (sufixo .1, .2, ...)
    """
    df = _renomear_colunas(df, _MAP_REPASSE)

    # Limpa valores "Unnamed: N" nos dados (problema do Excel com colunas sem cabecalho)
    for col in df.columns:
        df[col] = df[col].apply(
            lambda v: "" if str(v).strip().upper().startswith("UNNAMED:") else v
        )

    # Garante TODAS as colunas canonicas do _MAP_REPASSE, nao apenas um subconjunto.
    # Isso evita que colunas ausentes no arquivo fisico (ex: NrAtendimento) sejam
    # omitidas do CSV de saida. Derivado automaticamente do mapa - sem listas manuais.
    for col in _COLUNAS_CANONICAS_REPASSE:
        if col not in df.columns:
            df[col] = ""

    df["Data"]          = df["Data"].apply(_padronizar_data)
    df["Paciente"]      = df["Paciente"].apply(
        lambda v: re.sub(r"\s+", " ", str(v).strip()).upper()
    )
    df["ValorLiberado"] = df["ValorLiberado"].apply(_padronizar_valor)
    df["TipoArquivo"]   = "REPASSE"
    df["AbaOrigemDados"] = f"ABA: {nome_aba}"

    # Remove colunas lixo: Extra_N e duplicatas pandas (sufixo .1, .2, ...)
    # Alinhado com o mesmo tratamento ja feito em _processar_aba_producao
    lixo = [c for c in df.columns
            if re.match(r"^Extra_\d+$", str(c)) or re.search(r"\.\d+$", str(c))]
    df.drop(columns=lixo, inplace=True, errors="ignore")
    return df


# =============================================================================
# PROCESSAMENTO DE UM BLOCO TEXTO (UMA ABA)
# =============================================================================

def _processar_bloco_texto(
    dados_aba: str,
    nome_aba: str,
    nome_arquivo: str,
) -> str:
    """Filtra, detecta formato e processa o texto de uma única aba."""

    # FIX defensivo: colapsa \n internos em campos entre aspas (origem: CSV/TXT)
    dados_aba = re.sub(
        r'"([^"]*)"',
        lambda m: '"' + m.group(1).replace('\n', ' ').replace('\r', '') + '"',
        dados_aba,
    )
    # ── 1. Filtrar linhas inválidas ──────────────────────────────────────────
    linhas_validas = [l for l in dados_aba.splitlines() if _linha_e_valida(l)]
    if not linhas_validas:
        return ""

    tipo = _identificar_tipo_arquivo(nome_aba, nome_arquivo)
    if tipo == "PRODUCAO":
        tipo = _detectar_tipo_por_cabecalho(linhas_validas)

    # ── 2. REPASSE — cabeçalho sempre explícito ──────────────────────────────
    if tipo == "REPASSE":
        try:
            df = pd.read_csv(io.StringIO("\n".join(linhas_validas)), dtype=str)
            df.fillna("", inplace=True)
            df_proc = _processar_aba_repasse(df, nome_aba)
            return df_proc.to_csv(index=False, encoding="utf-8")
        except Exception as exc:
            logger.warning(f"Erro REPASSE '{nome_aba}': {exc}")
            return ""

    # ── 3. PRODUCAO — detecta formato ────────────────────────────────────────
    formato, idx_hdr = _detectar_cabecalho_producao(linhas_validas)

    if formato in ("COM_QTD", "SEM_QTD"):
        # Cabeçalho explícito na linha idx_hdr; descarta linhas anteriores
        linhas_csv = linhas_validas[idx_hdr:]
        try:
            df = pd.read_csv(io.StringIO("\n".join(linhas_csv)), dtype=str)
            df.fillna("", inplace=True)
        except Exception as exc:
            logger.warning(f"Erro PRODUCAO com header '{nome_aba}': {exc}")
            return ""
    else:
        # SEM_HEADER: nenhum cabeçalho encontrado → aplica cabeçalho canônico
        try:
            df = pd.read_csv(
                io.StringIO("\n".join(linhas_validas)),
                header=None, dtype=str,
            )
            df.fillna("", inplace=True)
            df = _atribuir_colunas(df, formato, nome_aba)
        except Exception as exc:
            logger.warning(f"Erro PRODUCAO sem header '{nome_aba}': {exc}")
            return ""

    df_proc = _processar_aba_producao(df, nome_aba)
    return df_proc.to_csv(index=False, encoding="utf-8")


def _consolidar_blocos(blocos_csv: list[str]) -> str:
    """Une múltiplos CSVs em um único, com cabeçalho único na primeira linha."""
    dfs = []
    for bloco in blocos_csv:
        bloco = bloco.strip()
        if not bloco:
            continue
        try:
            df = pd.read_csv(io.StringIO(bloco), dtype=str)
            df.fillna("", inplace=True)
            dfs.append(df)
        except Exception:
            pass
    if not dfs:
        return ""
    df_final = pd.concat(dfs, ignore_index=True, sort=False)
    df_final.fillna("", inplace=True)
    return df_final.to_csv(index=False, encoding="utf-8")


# =============================================================================
# FUNÇÃO PÚBLICA PRINCIPAL
# =============================================================================

def transformar_csv_arquivo(
    conteudo_texto: str,
    nome_arquivo: str = "",
    desmembrar_procedimentos_adicionais: bool = False,
) -> str:
    """
    Transforma o conteúdo extraído de um arquivo de endoscopia (PRODUCAO ou REPASSE)
    em um CSV padronizado puro, sem linhas vazias ou texto explicativo.

    Regras aplicadas:
    - Remove linhas completamente vazias, com só vírgulas (ex: "429,,,,,,,,,,,,")
      ou com apenas um valor não vazio
    - Remove linhas informativas de título ("CONTROLE DE EXAMES E PROCEDIMENTOS
      NA ENDOSCOPIA / MÊS ANO,Unnamed: 1,Unnamed: 2,...")
    - Extrai o nome da aba do marcador "=== ABA: <nome> ===" e insere na coluna
      AbaOrigemDados (ex: "ABA: JANEIRO 2026")
    - Primeira linha do CSV = cabeçalho obrigatório
    - Data padronizada para DD/MM/AAAA
    - Paciente em MAIÚSCULAS sem espaços extras
    - TipoArquivo = 'PRODUCAO' ou 'REPASSE' (identificado pelo nome da aba/arquivo)
    - Coluna QTD incluída; preenchida com "0" quando ausente na aba de origem
    - Abas com formato legado 2025 (EXAME REALIZADO 1 / EXAME REALIZADO 2) têm
      "EXAME REALIZADO 2" tratado como linha extra de procedimento
    - Todas as colunas canônicas do _MAP_REPASSE são sempre incluídas no CSV de REPASSE,
      mesmo que o arquivo físico não as contenha (ficam vazias)

    Args:
        conteudo_texto:   Texto extraído por extract_text_from_file.
                          Para Excel: contém blocos "=== ABA: <nome> ===" separando abas.
        nome_arquivo:     Nome original do arquivo (fallback para identificar tipo).
        desmembrar_procedimentos_adicionais:
                          Se True E TipoArquivo for PRODUCAO, chama
                          desmembrar_procedimentos_adicionais_csv_arquivo antes de retornar.

    Returns:
        CSV puro (str) com cabeçalho na primeira linha, ou "" em caso de falha total.
    """
    blocos_csv: list[str] = []

    # Divide por marcadores de aba (formato gerado por read_excel_file)
    partes = re.split(r"(===\s*ABA:\s*.+?===)", conteudo_texto)

    if len(partes) > 1:
        it = iter(partes[1:])
        for marcador, dados_aba in zip(it, it):
            nome_aba = re.sub(r"===\s*ABA:\s*|===", "", marcador).strip()
            dados_aba = dados_aba.strip()
            if not dados_aba:
                continue
            try:
                bloco = _processar_bloco_texto(dados_aba, nome_aba, nome_arquivo)
                if bloco:
                    blocos_csv.append(bloco)
            except Exception as exc:
                logger.warning(f"Falha ao processar aba '{nome_aba}': {exc}")
    else:
        # Arquivo CSV/TXT simples sem marcador de aba
        nome_aba = re.sub(r"\.[^.]+$", "", nome_arquivo)
        try:
            bloco = _processar_bloco_texto(conteudo_texto.strip(), nome_aba, nome_arquivo)
            if bloco:
                blocos_csv.append(bloco)
        except Exception as exc:
            logger.error(f"Falha ao processar '{nome_arquivo}': {exc}")
            return ""

    if not blocos_csv:
        return ""

    resultado = _consolidar_blocos(blocos_csv)

    # Desmembramento opcional (só PRODUCAO)
    if desmembrar_procedimentos_adicionais and _identificar_tipo_arquivo(nome_arquivo) == "PRODUCAO":
        resultado = desmembrar_procedimentos_adicionais_csv_arquivo(resultado)

    return resultado


# =============================================================================
# FUNÇÃO PÚBLICA: DESMEMBRAMENTO DE PROCEDIMENTOS ADICIONAIS
# =============================================================================

def desmembrar_procedimentos_adicionais_csv_arquivo(csv_texto: str) -> str:
    """
    Recebe um CSV de PRODUCAO já padronizado e desmembra a coluna
    'ProcedimentosAdicionais' em linhas extras.

    Regras:
    - Para cada valor em 'ProcedimentosAdicionais', gera uma LINHA EXTRA com:
        * Procedimento  = nome do procedimento adicional (MAIÚSCULAS)
        * Observacao    = "PROCEDIMENTO_ADICIONAL"
        * QTD           = "0"
        * demais campos = replicados da linha principal
    - A linha principal mantém sua Observacao original (vazia ou com valor já existente).
    - Não altera linhas de REPASSE (TipoArquivo != 'PRODUCAO').
    - Remove a coluna 'ProcedimentosAdicionais' do CSV final.
    - Delimitadores suportados: ponto-e-vírgula, '+', quebra de linha.
      Vírgulas são delimitadoras apenas quando não fazem parte de número decimal.

    Args:
        csv_texto: CSV puro (string) com cabeçalho na primeira linha.

    Returns:
        CSV puro (string) com as linhas extras já inseridas e sem
        a coluna 'ProcedimentosAdicionais'.
    """
    if not csv_texto or not csv_texto.strip():
        return csv_texto

    try:
        df = pd.read_csv(io.StringIO(csv_texto.strip()), dtype=str)
        df.fillna("", inplace=True)
    except Exception as exc:
        logger.warning(f"desmembrar_procedimentos_adicionais_csv_arquivo: erro ao ler CSV: {exc}")
        return csv_texto

    if "ProcedimentosAdicionais" not in df.columns:
        return csv_texto

    linhas_saida = []
    for _, row in df.iterrows():
        linhas_saida.append(row.copy())

        # Só desmembra linhas de PRODUCAO
        if str(row.get("TipoArquivo", "")).strip().upper() != "PRODUCAO":
            continue

        proc_adicionais = str(row.get("ProcedimentosAdicionais", "")).strip()
        if proc_adicionais.upper() in ("", "NAN", "NONE", "NAN", "-"):
            continue

        # Delimitadores: ; | + | quebra de linha
        # Vírgula apenas quando não separa dígitos (evita quebrar "1,234")
        separados = re.split(r"[;+\n\r]+|(?<!\d),(?!\d)", proc_adicionais)

        for proc in separados:
            proc = proc.strip().upper()
            if not proc or proc in ("-", ""):
                continue
            extra = row.copy()
            extra["Procedimento"]            = proc
            extra["ProcedimentosAdicionais"] = ""
            extra["Observacao"]              = "PROCEDIMENTO_ADICIONAL"
            extra["QTD"]                     = "0"
            linhas_saida.append(extra)

    df_saida = pd.DataFrame(linhas_saida)
    df_saida.drop(columns=["ProcedimentosAdicionais"], inplace=True, errors="ignore")
    df_saida.fillna("", inplace=True)
    return df_saida.to_csv(index=False, encoding="utf-8")

###############################################################################


# =============================================================================
# FUNÇÕES DE CORRELAÇÃO LOCAL
# =============================================================================

from datetime import datetime, timedelta
from difflib import SequenceMatcher
from typing import Dict, List, Tuple, Optional

# Dicionário de sinônimos e variações de procedimentos
SINONIMOS_PROCEDIMENTOS = {
    # Colonoscopia e variações
    "COLONO": ["COLONOSCOPIA", "COLONOSCOPIA INCLUI", "RETOSSIGMOIDOSCOPIA"],
    "COLONOSCOPIA": ["COLONO", "COLONOSCOPIA INCLUI", "RETOSSIGMOIDOSCOPIA"],
    "COLONOCOPIA": ["COLONO", "COLONOSCOPIA"],  # Typo comum
    "RETOSSIGMOIDOSCOPIA": ["COLONOSCOPIA", "COLONO"],
    
    # Endoscopia Digestiva Alta
    "ENDOSCOPIA": ["EDA", "ENDOSCOPIA DIGESTIVA ALTA", "ESOFAGOGASTRODUODENOSCOPIA", "ESOFAGO GASTRO DUODENOSCOPIA"],
    "EDA": ["ENDOSCOPIA", "ENDOSCOPIA DIGESTIVA ALTA", "ESOFAGOGASTRODUODENOSCOPIA"],
    "ENDOCOSPIA": ["ENDOSCOPIA"],  # Typo
    "ENDDOSCOPIA": ["ENDOSCOPIA"],  # Typo
    "ESOFAGOGASTRODUODENOSCOPIA": ["ENDOSCOPIA", "EDA"],
    "ESOFAGO GASTRO DUODENOSCOPIA": ["ENDOSCOPIA", "EDA"],
    
    # Ecoendoscopia
    "ECOENDOSCOPIA": ["ECOEDA", "ULTRASSOM ENDOSCOPICO"],
    "ECOEDA": ["ECOENDOSCOPIA", "ULTRASSOM ENDOSCOPICO"],
    
    # Teste de Urease / H. Pylori
    "TESTE": ["PESQUISA", "EXAME"],
    "UREASE": ["H PYLORI", "HELICOBACTER", "HELICOBACTER PYLORI"],
    "TESTE DE UREASE": ["PESQUISA H PYLORI", "TESTE UREASE", "UREASE", "H PYLORI", "HELICOBACTER"],
    "TESTE DA UREASE": ["TESTE DE UREASE", "PESQUISA H PYLORI", "HELICOBACTER"],
    "HP": ["H PYLORI", "HELICOBACTER"],
    "HPYLORI": ["H PYLORI", "HELICOBACTER"],
    "HELICOBACTER": ["H PYLORI", "UREASE", "HP"],
    "HELICOBACTER PYLORI": ["H PYLORI", "UREASE", "HP"],
    
    # Anátomo Patológico / Biópsia / Citologia
    "ANATOMO": ["BIOPSIA", "CITOLOGIA", "AP"],
    "ANATOMIA": ["BIOPSIA", "CITOLOGIA", "AP"],
    "ANATOMIA PATOLOGICA": ["BIOPSIA", "CITOLOGIA", "AP", "ANÁTOMO PATOLÓGICO"],
    "ANATOMO PATOLOGICA": ["BIOPSIA", "CITOLOGIA", "AP", "ANÁTOMO PATOLÓGICO"],
    "ANATOMO PATOLOGICO": ["AP", "BIOPSIA", "CITOLOGIA", "ANÁTOMO PATOLÓGICO", "ANATOMOPATOLOGICO"],
    "ANÁTOMO PATOLÓGICO": ["AP", "BIOPSIA", "CITOLOGIA", "ANATOMO PATOLOGICO", "ANATOMOPATOLOGICO"],
    "ANÁTOMO PATLÓGICO": ["ANÁTOMO PATOLÓGICO", "BIOPSIA", "AP"],  # Typo
    "ANATOMOPATOLOGICO": ["AP", "BIOPSIA", "CITOLOGIA", "ANATOMO PATOLOGICO"],
    "ANÁTOMOPATOLOGICO": ["AP", "BIOPSIA", "CITOLOGIA", "ANÁTOMO PATOLÓGICO"],
    "AP": ["ANATOMO PATOLOGICO", "BIOPSIA", "CITOLOGIA", "ANÁTOMO PATOLÓGICO", "ANATOMOPATOLOGICO"],
    "BIOPSIA": ["ANATOMO PATOLOGICO", "AP", "CITOLOGIA", "BIOPSIAS"],
    "BIOPSIAS": ["BIOPSIA", "ANATOMO PATOLOGICO", "AP"],
    "BIOPSIA SERIADO": ["BIOPSIA", "ANÁTOMO PATOLÓGICO", "AP"],
    "CITOLOGIA": ["BIOPSIA", "ANATOMO PATOLOGICO", "AP"],
    
    # Polipectomia
    "POLIPECTOMIA": ["POLIPO", "RESSECCAO POLIPO", "RETIRADA POLIPO"],
    "POLIPO": ["POLIPECTOMIA", "RESSECCAO POLIPO"],
    "PÓLIPO": ["POLIPECTOMIA", "POLIPO"],
    "POLIOECTOMIA": ["POLIPECTOMIA"],  # Typo
    
    # Hemostasia
    "HEMOSTASIA": ["HEMOSTASE", "CONTROLE SANGRAMENTO", "HEMOSTASE ENDOSCOPICA"],
    "HEMOSTASE": ["HEMOSTASIA", "CONTROLE SANGRAMENTO"],
    
    # Mucosectomia
    "MUCOSECTOMIA": ["RESSECCAO MUCOSA", "RESSECCAO ENDOSCOPICA"],
    
    # Tatuagem
    "TATUAGEM": ["MARCACAO", "TATUAGEM ENDOSCOPICA"],
    
    # Biópsia Hepática
    "BIOPSIA HEPATICA": ["BIOPSIA", "HEPATICA"],
    "HEPATICA": ["BIOPSIA HEPATICA", "FIGADO"],
    
    # CPRE / Colangiopancreatografia / Papilotomia
    "CPRE": ["COLANGIOPANCREATOGRAFIA", "COLANGIOPANCREATOGRAFIA RETROGRADA", "PAPILOTOMIA"],
    "COLANGIOPANCREATOGRAFIA": ["CPRE", "PAPILOTOMIA", "COLANGIOPANCREATOGRAFIA RETROGRADA"],
    "COLANGIOPANCREATOGRAFIA RETROGRADA": ["CPRE", "COLANGIOPANCREATOGRAFIA", "PAPILOTOMIA"],
    "PAPILOTOMIA": ["CPRE", "COLANGIOPANCREATOGRAFIA"],
    "PAPILOTOMIA ENDOSCOPICA": ["CPRE", "PAPILOTOMIA"],
    
    # Dilatação
    "DILATACAO": ["DILATAÇÃO", "DILATACAO PNEUMATICA"],
    "DILATAÇÃO": ["DILATACAO", "DILATACAO PNEUMATICA"],
    
    # Gastrostomia
    "GASTROSTOMIA": ["GTT", "GASTROSTOMIA ENDOSCOPICA"],
    "GASTROSTOMIA ENDOSCOPICA": ["GASTROSTOMIA", "GTT"],
    "GTT": ["GASTROSTOMIA"],
    
    # Passagem de Sonda
    "SONDA": ["PASSAGEM SONDA", "SONDA NASO ENTERAL", "SONDA NASOENTERAL"],
    "PASSAGEM DE SONDA": ["SONDA", "SONDA NASO ENTERAL", "PASSAGEM SONDAS"],
    "PASSAGEM DE SONDAS": ["PASSAGEM DE SONDA", "SONDA"],
    "SONDA NASO ENTERAL": ["SONDA", "PASSAGEM SONDA"],
    
    # Anuscopia / Retoscopia
    "ANUSCOPIA": ["RETOSCOPIA", "EXAME ANORRETAL"],
    "RETOSCOPIA": ["ANUSCOPIA"],
    
    # Prótese
    "PROTESE": ["COLOCACAO PROTESE", "IMPLANTE PROTESE"],
    "PRÓTESE": ["PROTESE", "COLOCACAO PROTESE"],
    "COLOCACAO DE PROTESE": ["PROTESE", "IMPLANTE PROTESE"],
    "COLOCACAO PROTESE": ["PROTESE", "COLOCACAO DE PROTESE"],
    
    # Visita Hospitalar
    "VISITA": ["VISITA HOSPITALAR", "VISITA PACIENTE INTERNADO", "HOSPITALAR"],
    "VISITA HOSPITALAR": ["VISITA", "VISITA PACIENTE INTERNADO", "HOSPITALAR"],
    "HOSPITALAR": ["VISITA", "VISITA HOSPITALAR"],
    
    # Cromoscopia
    "CROMOSCOPIA": ["CROMOSCOPIA ENDOSCOPICA", "COLORACAO"],
    
    # Esôfago
    "ESOFAGO": ["ESOFAGICO", "ESOFAGICA"],
    "ESOFAGICO": ["ESOFAGO"],
    
    # Cálculo / Coledociano
    "CALCULO": ["CALCULOSE", "PEDRA"],
    "COLEDOCIANO": ["COLECOCO", "BILIAR"],
    
    # Drenagem
    "DRENAGEM": ["DRENAGEM BILIAR", "DESCOMPRESSAO"],
    "DESCOMPRESSAO": ["DRENAGEM"],
    "DESCOMPRESSAO COLONICA": ["DESCOMPRESSAO", "COLONICA"],
    "COLONICA": ["COLON", "COLONOSCOPIA"],
    
    # Ostomia
    "OSTOMIA": ["GASTROSTOMIA", "COLOSTOMIA"],
    "TROCA DE GTT": ["GTT", "GASTROSTOMIA"],
}

# Pares anatomicamente incompatíveis: EDA (trato alto) ≠ Colonoscopia (trato baixo).
# Usados para sinalizar matches via SequenceMatcher acidental (ex: ENDOSCOPIA↔COLONOSCOPIA
# compartilham "OSCOPIA" como substring, gerando sim=0.73 acima do limiar 0.65).
_PARES_ANATOMICAMENTE_DIVERGENTES: list[frozenset] = [
    frozenset({"ENDOSCOPIA", "COLONOSCOPIA"}),
    frozenset({"ENDOSCOPIA", "COLONO"}),
    frozenset({"ENDOSCOPIA", "RETOSSIGMOIDOSCOPIA"}),
    frozenset({"EDA", "COLONOSCOPIA"}),
    frozenset({"EDA", "COLONO"}),
    frozenset({"EDA", "RETOSSIGMOIDOSCOPIA"}),
]

# Palavras-chave de procedimentos companion (Urease/Helicobacter standalone).
_KEYWORDS_COMPANION_PROC = {"UREASE", "HELICOBACTER"}
# Palavras-chave de procedimentos principais — excluem uma linha do perfil companion.
_KEYWORDS_PRINCIPAL_PROC = {
    "ENDOSCOPIA", "COLONOSCOPIA", "COLONO", "ECOENDOSCOPIA",
    "GASTROSTOMIA", "MUCOSECTOMIA", "CPRE", "COLANGIOPANCREATOGRAFIA",
    "PAPILOTOMIA", "BRONCOSCOPIA", "RETOSSIGMOIDOSCOPIA",
}


def _sao_anatomicamente_divergentes(proc1: str, proc2: str) -> bool:
    """
    Retorna True quando os dois procedimentos pertencem a grupos anatômicos
    reconhecidamente distintos (ex: ENDOSCOPIA/EDA vs COLONOSCOPIA/trato baixo).
    Não bloqueia o match — apenas sinaliza para revisão humana via status.
    """
    p1 = _normalizar_procedimento(proc1)
    p2 = _normalizar_procedimento(proc2)
    tokens1 = set(p1.split())
    tokens2 = set(p2.split())
    for par in _PARES_ANATOMICAMENTE_DIVERGENTES:
        t1, t2 = tuple(par)
        if (t1 in tokens1 and t2 in tokens2) or (t2 in tokens1 and t1 in tokens2):
            return True
    return False


def _e_procedimento_companion(proc: str) -> bool:
    """
    Retorna True para procedimentos reconhecidamente acompanhantes/adicionais
    (ex: Teste de Urease standalone) que na PRODUCAO costumam estar em
    ProcedimentosAdicionais. Exclui procedimentos compostos que já contêm o
    exame principal no nome (ex: 'EDA Com Biópsia E Teste De Urease').
    """
    p_norm = _normalizar_procedimento(proc)
    tokens = set(p_norm.split())
    return bool(tokens & _KEYWORDS_COMPANION_PROC) and not bool(tokens & _KEYWORDS_PRINCIPAL_PROC)


def _normalizar_procedimento(proc: str) -> str:
    """Normaliza nome de procedimento para comparação semântica."""
    if not proc or str(proc).strip() in ("", "nan", "NaN"):
        return ""
    s = str(proc).upper().strip()
    # Remove acentos
    s = s.replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U')
    s = s.replace('Â', 'A').replace('Ê', 'E').replace('Ô', 'O')
    s = s.replace('Ã', 'A').replace('Õ', 'O')
    s = s.replace('Ç', 'C')
    s = re.sub(r'[^\w\s]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    # Remove plural (S no final de palavras)
    s = re.sub(r'\bBIOPSIAS\b', 'BIOPSIA', s)
    s = re.sub(r'\bPOLIPOS\b', 'POLIPO', s)
    s = re.sub(r'\bSONDAS\b', 'SONDA', s)
    # Remove separadores comuns em procedimentos compostos
    s = re.sub(r'\s*[/+]\s*', ' ', s)  # Remove / e +
    return s

def _expandir_abreviacao(proc: str) -> str:
    """Expande abreviações comuns de procedimentos."""
    ABREVIACOES = {
        "COLONO": "COLONOSCOPIA",
        "EDA": "ENDOSCOPIA DIGESTIVA ALTA",
        "RETO": "RETOSSIGMOIDOSCOPIA",
        "RETOSSIGMOIDO": "RETOSSIGMOIDOSCOPIA",
        "POLIPECTOMIA": "POLIPECTOMIA",
        "LIGADURA": "LIGADURA ELASTICA",
        "DILATACAO": "DILATACAO ESOFAGICA",
        "ESCLEROSE": "ESCLEROTERAPIA",
        "MUCOSECTOMIA": "MUCOSECTOMIA",
        "HEMOSTASIA": "HEMOSTASIA ENDOSCOPICA",
    }
    proc_norm = _normalizar_procedimento(proc)
    
    # Verifica match exato
    if proc_norm in ABREVIACOES:
        return ABREVIACOES[proc_norm]
    
    # Verifica se começa com abreviação
    for abrev, expandido in ABREVIACOES.items():
        if proc_norm.startswith(abrev + " ") or proc_norm == abrev:
            return proc_norm.replace(abrev, expandido, 1)
    
    return proc_norm

def _match_palavra_contida(proc1: str, proc2: str) -> bool:
    """Verifica se um procedimento está contido no outro ou compartilha palavra-chave."""
    p1 = _normalizar_procedimento(proc1).strip()
    p2 = _normalizar_procedimento(proc2).strip()
    
    if not p1 or not p2:
        return False
    
    # Verifica contenção direta
    if p1 in p2 or p2 in p1:
        return True
    
    # Verifica se primeira palavra significativa é igual (mínimo 4 caracteres)
    palavras1 = [p for p in p1.split() if len(p) >= 4]
    palavras2 = [p for p in p2.split() if len(p) >= 4]
    
    if palavras1 and palavras2:
        # Verifica se primeira palavra significativa é igual
        if palavras1[0] == palavras2[0]:
            return True
        # Verifica se alguma palavra significativa está contida
        for palavra in palavras1:
            if len(palavra) >= 6 and any(palavra in p2_word or p2_word in palavra for p2_word in palavras2):
                return True
    
    return False

def _verificar_sinonimo(proc1: str, proc2: str) -> bool:
    """Verifica se dois procedimentos são sinônimos conhecidos."""
    p1_norm = _normalizar_procedimento(proc1)
    p2_norm = _normalizar_procedimento(proc2)
    
    # Verifica match direto no dicionário
    for chave, sinonimos in SINONIMOS_PROCEDIMENTOS.items():
        if chave in p1_norm:
            for sin in sinonimos:
                if sin in p2_norm:
                    return True
        if chave in p2_norm:
            for sin in sinonimos:
                if sin in p1_norm:
                    return True
    
    # Para procedimentos compostos (ex: "ANATOMO + POLIPECTOMIA")
    # Verifica se todas as palavras-chave de um estão no outro
    palavras1 = set(p1_norm.split())
    palavras2 = set(p2_norm.split())
    
    # Se tem pelo menos 2 palavras em comum e são palavras-chave
    palavras_chave = {'BIOPSIA', 'POLIPECTOMIA', 'POLIPO', 'ANATOMO', 'TESTE', 'UREASE', 
                      'MUCOSECTOMIA', 'HEMOSTASIA', 'ENDOSCOPIA', 'COLONOSCOPIA', 'COLONO'}
    comuns = palavras1 & palavras2 & palavras_chave
    if len(comuns) >= 2:
        return True
    
    return False

def _similaridade_procedimento(proc1: str, proc2: str, cache: Dict = None) -> float:
    """Calcula similaridade entre dois procedimentos (0.0 a 1.0) com cache e sinônimos."""
    if cache is not None:
        key = (proc1, proc2)
        if key in cache:
            return cache[key]
    
    # Expande abreviações antes de comparar
    p1_expandido = _expandir_abreviacao(proc1)
    p2_expandido = _expandir_abreviacao(proc2)
    
    if not p1_expandido or not p2_expandido:
        return 0.0
    
    # Match exato após expansão
    if p1_expandido == p2_expandido:
        if cache is not None:
            cache[(proc1, proc2)] = 1.0
        return 1.0
    
    # Verifica match por palavra contida
    if _match_palavra_contida(p1_expandido, p2_expandido):
        if cache is not None:
            cache[(proc1, proc2)] = 0.95
        return 0.95
    
    # Verifica sinônimos conhecidos
    if _verificar_sinonimo(proc1, proc2):
        if cache is not None:
            cache[(proc1, proc2)] = 1.0
        return 1.0
    
    # Verifica se um contém o outro (palavras-chave)
    palavras1 = set(p1_expandido.split())
    palavras2 = set(p2_expandido.split())
    intersecao = palavras1 & palavras2
    
    if intersecao:
        # Se tem palavras em comum, calcula similaridade
        ratio = len(intersecao) / max(len(palavras1), len(palavras2))
        if ratio >= 0.4:
            score = SequenceMatcher(None, p1_expandido, p2_expandido).ratio()
            if cache is not None:
                cache[(proc1, proc2)] = score
            return score
    
    # Fallback para SequenceMatcher completo
    score = SequenceMatcher(None, p1_expandido, p2_expandido).ratio()
    if cache is not None:
        cache[(proc1, proc2)] = score
    return score

def _datas_compativeis(data1: str, data2: str, tolerancia_dias: int = 1) -> bool:
    """Verifica se duas datas são compatíveis (tolerância ±N dias)."""
    if data1 == data2:
        return True
    try:
        d1 = datetime.strptime(data1, "%d/%m/%Y")
        d2 = datetime.strptime(data2, "%d/%m/%Y")
        return abs((d1 - d2).days) <= tolerancia_dias
    except:
        return False

def _extrair_valor_numerico(valor: str) -> float:
    """Extrai valor numérico de string."""
    if not valor or str(valor).strip() in ("", "nan", "NaN"):
        return 0.0
    try:
        s = str(valor).replace("R$", "").replace(",", ".").strip()
        return float(s)
    except:
        return 0.0

def _normalizar_nome_paciente(nome: str) -> str:
    """Normaliza nome de paciente removendo acentos e caracteres especiais."""
    if not nome or str(nome).strip() in ("", "nan", "NaN"):
        return ""
    s = str(nome).upper().strip()
    # Remove acentos
    s = s.replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U')
    s = s.replace('Â', 'A').replace('Ê', 'E').replace('Ô', 'O')
    s = s.replace('Ã', 'A').replace('Õ', 'O')
    s = s.replace('Ç', 'C')
    s = s.replace('À', 'A').replace('È', 'E').replace('Ì', 'I').replace('Ò', 'O').replace('Ù', 'U')
    s = s.replace('Ü', 'U').replace('Ö', 'O').replace('Ä', 'A')
    # Remove caracteres especiais, mantém apenas letras e espaços
    s = re.sub(r'[^A-Z\s]', '', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def _criar_indice_repasse(df_rep: pd.DataFrame) -> Dict[Tuple[str, str], List[int]]:
    """Cria índice para busca rápida no REPASSE por (Data, Paciente)."""
    indice = {}
    for idx, row in df_rep.iterrows():
        data = row.get("Data", "")
        paciente = _normalizar_nome_paciente(row.get("Paciente", ""))
        key = (data, paciente)
        if key not in indice:
            indice[key] = []
        indice[key].append(idx)
    return indice

def _criar_indice_repasse_atendimento(df_rep: pd.DataFrame) -> Dict[Tuple[str, str], List[int]]:
    """Cria índice para busca rápida no REPASSE por (Data, NrAtendimento)."""
    indice = {}
    for idx, row in df_rep.iterrows():
        data = row.get("Data", "")
        nr_atend = str(row.get("NrAtendimento", "")).strip()
        if nr_atend and nr_atend not in ("", "nan", "NaN"):
            key = (data, nr_atend)
            if key not in indice:
                indice[key] = []
            indice[key].append(idx)
    return indice

def _determinar_status_correlacao(valor_liberado: float, tem_match: bool) -> str:
    """Determina o status de correlação baseado no valor liberado."""
    if not tem_match:
        return "NAO_FATURADO_NO_REPASSE"
    
    if valor_liberado == 0:
        return "CORRELACIONADO"
    elif valor_liberado > 0:
        return "CORRELACIONADO"
    else:
        return "CORRELACIONADO"

def _extrair_tokens_nome(nome: str) -> list[str]:
    """
    Retorna os tokens significativos do nome normalizado, limitado a 4.

    Remove partículas (DE, DA, DO, DOS, DAS, E) e tokens curtos (≤ 2 chars).
    Usado pelo FB1 para gerar combinações de busca por partes do nome.

    Exemplos:
        "MARIA DAS GRACAS SILVA"  → ["MARIA", "GRACAS", "SILVA"]
        "JOAO DE SOUZA PEREIRA"   → ["JOAO", "SOUZA", "PEREIRA"]
        "ANA"                     → ["ANA"]
    """
    PARTICULAS = {"DE", "DA", "DO", "DOS", "DAS", "E"}
    partes = _normalizar_nome_paciente(nome).split()
    validos = [p for p in partes if p not in PARTICULAS and len(p) > 2]
    return validos[:4]  # limita a 4 tokens


def _tokens_fuzzy_em_comum(
    tokens_prod: list[str],
    nome_rep_norm: str,
    limiar_token: float = 0.82,
) -> list[tuple[str, str, float]]:
    """
    Para cada token da PRODUCAO, encontra o token mais similar no REPASSE
    usando SequenceMatcher. Retorna apenas os pares que atingem o limiar.

    Cada token do REPASSE só pode ser usado uma vez (match exclusivo),
    evitando que um único token do REPASSE satisfaça múltiplos tokens da PRODUCAO.

    Retorna lista de (token_prod, token_rep, score) ordenada por score desc.

    Exemplo:
        tokens_prod = ["MICHELE", "KAROLINNE", "FERNANDES"]
        nome_rep    = "MICHELLE KAROLINNE FERNANDES GUERREIRO"
        → [("MICHELE","MICHELLE",0.93), ("KAROLINNE","KAROLINNE",1.0), ("FERNANDES","FERNANDES",1.0)]
    """
    tokens_rep = [t for t in nome_rep_norm.split() if len(t) > 2]
    usados: set[str] = set()
    resultado: list[tuple[str, str, float]] = []

    for tp in tokens_prod:
        melhor_rep = None
        melhor_s = 0.0
        for tr in tokens_rep:
            if tr in usados:
                continue
            s = SequenceMatcher(None, tp, tr).ratio()
            if s >= limiar_token and s > melhor_s:
                melhor_s = s
                melhor_rep = tr
        if melhor_rep is not None:
            resultado.append((tp, melhor_rep, round(melhor_s, 3)))
            usados.add(melhor_rep)

    return sorted(resultado, key=lambda x: x[2], reverse=True)


def _buscar_fallback1_combinacoes_nome(
    data_prod: str,
    nome_prod: str,
    df_rep: pd.DataFrame,
    indice_repasse: Dict[Tuple[str, str], List[int]],
    proc_prod: str,
    cache_similaridade: Dict,
    limiar_similaridade: float = 0.65,
    limiar_token: float = 0.82,
    min_tokens_similares: int = 3,
) -> Tuple[Optional[int], float]:
    """
    Fallback 1 — Similaridade fuzzy por token + Data ±1 dia + Procedimento.

    Estratégia:
      Para cada candidato no REPASSE (dentro de ±1 dia), compara token a token
      o nome da PRODUCAO com o nome do REPASSE usando SequenceMatcher por token.
      Aceita o match se AMBAS as condições forem satisfeitas:
        1. >= min_tokens_similares tokens com similaridade >= limiar_token
        2. O primeiro token da PRODUCAO tem um similar no REPASSE
           (evita matches onde só sobrenomes coincidem)

    Por que fuzzy e não match exato:
      Typos reais como MICHELE/MICHELLE, VASCONCELOS/VASCONCELLOS,
      APPARECIDA/APARECIDA passam no limiar 0.82 e seriam perdidos
      com match exato de string.

    Por que primeiro token obrigatório:
      Impede que pares de sobrenomes genéricos (SILVA+SANTOS, PEREIRA+SILVA)
      matcheiem pacientes completamente diferentes.

    Por que min 3 tokens:
      Com apenas 2 tokens, combinações como MARIA+SILVA ou FABIANO+SILVA
      ainda geram falsos positivos frequentes em bases hospitalares.

    Parâmetros:
        limiar_token:        similaridade mínima por par de tokens (padrão 0.82)
        min_tokens_similares: quantidade mínima de tokens similares (padrão 3)

    Retorna (idx, score_procedimento) do melhor candidato, ou (None, 0.0).
    """
    tokens_prod = _extrair_tokens_nome(nome_prod)
    if len(tokens_prod) < min_tokens_similares:
        # Não há tokens suficientes para atingir o mínimo — não tenta
        return None, 0.0

    primeiro_token = tokens_prod[0]

    melhor_idx: Optional[int] = None
    melhor_score: float = 0.0

    datas_candidatas = [data_prod]
    try:
        data_dt = datetime.strptime(data_prod, "%d/%m/%Y")
        datas_candidatas += [
            (data_dt + timedelta(days=d)).strftime("%d/%m/%Y")
            for d in [-1, 1]
        ]
    except ValueError:
        pass

    for data_cand in datas_candidatas:
        for (data_rep, pac_rep), idxs in indice_repasse.items():
            if data_rep != data_cand:
                continue

            # Avalia tokens fuzzy contra o nome normalizado do candidato
            matches = _tokens_fuzzy_em_comum(tokens_prod, pac_rep, limiar_token)

            # Condição 1: mínimo de tokens similares
            if len(matches) < min_tokens_similares:
                continue

            # Condição 2: primeiro token da PRODUCAO deve ter similar no REPASSE
            primeiro_tem_match = any(tp == primeiro_token for tp, tr, s in matches)
            if not primeiro_tem_match:
                continue

            # Candidato passou nas duas condições — avalia procedimento
            for i in idxs:
                if df_rep.at[i, "_matched"]:
                    continue
                proc_rep = df_rep.iloc[i].get("Procedimento", "")
                score = _similaridade_procedimento(proc_prod, proc_rep, cache_similaridade)
                if score >= limiar_similaridade and score > melhor_score:
                    melhor_score = score
                    melhor_idx = i

    return melhor_idx, melhor_score


def _buscar_fallback2_paciente_proc_data_ampla(
    data_prod: str,
    paciente_norm: str,
    proc_prod: str,
    df_rep: pd.DataFrame,
    indice_repasse: Dict[Tuple[str, str], List[int]],
    cache_similaridade: Dict,
    limiar_similaridade: float = 0.65,
    tolerancia_dias: int = 7,
) -> Tuple[Optional[int], float]:
    """
    Fallback 2 — Nome exato + Procedimento ≥ limiar + Data ±7 dias.

    Usa o mesmo índice por (data, paciente), mas varre a janela de 2 a 7 dias
    (os ±1 dia já foram cobertos pela chave principal). O nome deve ser idêntico
    após normalização; o procedimento deve atingir o mesmo limiar de similaridade.

    Cenário típico resolvido:
        Procedimento realizado em 10/03, lançado no REPASSE como 14/03
        → ±1 dia não bate, mas ±7 dias captura.

    Retorna (idx, score) do melhor candidato, ou (None, 0.0).
    """
    melhor_idx: Optional[int] = None
    melhor_score: float = 0.0

    try:
        data_dt = datetime.strptime(data_prod, "%d/%m/%Y")
    except ValueError:
        return None, 0.0

    for delta in range(-tolerancia_dias, tolerancia_dias + 1):
        if abs(delta) <= 1:
            continue  # ±1 dia já coberto pela chave principal
        data_candidata = (data_dt + timedelta(days=delta)).strftime("%d/%m/%Y")
        key = (data_candidata, paciente_norm)
        if key not in indice_repasse:
            continue
        for i in indice_repasse[key]:
            if df_rep.at[i, "_matched"]:
                continue
            proc_rep = df_rep.iloc[i].get("Procedimento", "")
            score = _similaridade_procedimento(proc_prod, proc_rep, cache_similaridade)
            if score >= limiar_similaridade and score > melhor_score:
                melhor_score = score
                melhor_idx = i

    return melhor_idx, melhor_score


# =============================================================================
# TABELA TUSS — geração, carregamento e verificação pós-correlação
# =============================================================================

_TUSS_DIR = Path(__file__).parent
_TUSS_LOOKUP_PATH = _TUSS_DIR / "tuss_lookup_table.csv"
_TUSS_VALORES_PATH = _TUSS_DIR / "tuss_valores.csv"
_TUSS_XLSX_PATH    = _TUSS_DIR / "TUSS (ATUALIZADO).xlsx"

# Mapeamento PRODUCAO→TUSS embutido. Atualizar quando o XLSX mudar (re-executar _gerar_tabela_tuss).
_TABELA_TUSS_EMBUTIDA: dict[str, dict] = {
    'ANUSCOPIA_': {"TipoCobranca": 'sem_mapeamento_tuss', "CodigosTUSS": '', "codigo_base_proc_principal": '', "Descricao_REPASSE": 'Sem correspondência exata TUSS'},
    'COLONO_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201082', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia (Inclui A Retossigmoidoscopia)'},
    'COLONO_ANATOMO+POLIPECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202666, 40202542', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Biópsia E Polipectomia'},
    'COLONO_ANATOMO PATLOGICO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202666', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Biópsia E/Ou Citologia'},
    'COLONO_ANATOMO PATOLOGICO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202666', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Biópsia E/Ou Citologia'},
    'COLONO_ANATOMO PATOLOGICO+MUCOSECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202666, 40202712', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Biópsia E Mucosectomia'},
    'COLONO_ANATOMO PATOLOGICO+POLIPECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202666, 40202542', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Biópsia E Polipectomia'},
    'COLONO_ANATOMO PATOLOGICO+TATUAGEM': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202666, 40202135', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Biópsia E Tatuagem'},
    'COLONO_ANATOMO PATOLOGICO+MUCOSECTOMIA+POLIPECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202666, 40202712, 40202542', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Biópsia, Mucosectomia E Polipectomia'},
    'COLONO_ANATOMO PATOLOGICO+POLIPO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202666, 40202542', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Biópsia E Polipectomia'},
    'COLONO_ANATOMO+MUCOSECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202666, 40202712', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Biópsia E Mucosectomia'},
    'COLONO_ANATOMO+POLIOECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202666, 40202542', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Biópsia E Polipectomia'},
    'COLONO_ANATOMOPATOLOGICO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202666', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Biópsia E/Ou Citologia'},
    'COLONO_ANATOMOPATOLOGICO+POLIPO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202666, 40202542', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Biópsia E Polipectomia'},
    'COLONO_BRADES': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201082', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia (Inclui A Retossigmoidoscopia)'},
    'COLONO_MUCOSECTOMIA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202712', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Mucosectomia'},
    'COLONO_POLEPECTOMIA+MUCOSECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202542, 40202712', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Polipectomia E Mucosectomia'},
    'COLONO_POLIECTOMIA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202542', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Polipectomia De Cólon (Independente Do Número De Pólipos)'},
    'COLONO_POLIPECTOMIA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202542', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Polipectomia De Cólon (Independente Do Número De Pólipos)'},
    'COLONO_POLIPECTOMIA+MUCOSECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202542, 40202712', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Polipectomia E Mucosectomia'},
    'COLONO_POLIPOECTOMIA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202542', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Polipectomia De Cólon (Independente Do Número De Pólipos)'},
    'COLONO_RETIRADA DE CORPO ESRANHO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202569', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Retirada de corpo estranho do cólon'},
    'COLONO_TESTE DE UREASE': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40201082, 40202615', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia + Teste De Urease Para Pesquisa De Helicobacter Pylori'},
    'COLONO_TESTE DE UREASE+ANATOMO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40201082, 40202615', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia + Teste De Urease Para Pesquisa De Helicobacter Pylori'},
    'COLONO_TESTE UREASE': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40201082, 40202615', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia + Teste De Urease Para Pesquisa De Helicobacter Pylori'},
    'COLONOCOPIA_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201082', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia (Inclui A Retossigmoidoscopia)'},
    'COLONOCOPIA_POLIPECTOMIA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202542', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Polipectomia De Cólon (Independente Do Número De Pólipos)'},
    'COLONOSCOPIA_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201082', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia (Inclui A Retossigmoidoscopia)'},
    'COLONOSCOPIA_-': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201082', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia (Inclui A Retossigmoidoscopia)'},
    'COLONOSCOPIA_ANATOMO PATOLOGICO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202666', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Biópsia E/Ou Citologia'},
    'COLONOSCOPIA_ANATOMO PATOLOGICO+POLIPECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202666, 40202542', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Biópsia E Polipectomia'},
    'COLONOSCOPIA_BIOPSIA SERIADO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202666', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Biópsia E/Ou Citologia'},
    'COLONOSCOPIA_DESCOMPRESSAO COLONICA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202143', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Descompressão colônica por colonoscopia'},
    'COLONOSCOPIA_HEMOSTASIA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202313', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Hemostasias de cólon'},
    'COLONOSCOPIA_MUCOSECTOMIA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202712', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia Com Mucosectomia'},
    'COLONOSCOPIA_POLIPECTOMIA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202542', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Polipectomia De Cólon (Independente Do Número De Pólipos)'},
    'COLONOSCOPIA_TESTE DE UREASE': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40201082, 40202615', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia + Teste De Urease Para Pesquisa De Helicobacter Pylori'},
    'COLONOSCOPIA_TESTE DE UREASE - NEGATIVO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40201082, 40202615', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia + Teste De Urease Para Pesquisa De Helicobacter Pylori'},
    'COLONOSCOPIA_TESTE UREASE - NEGATIVO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40201082, 40202615', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia + Teste De Urease Para Pesquisa De Helicobacter Pylori'},
    'COLONOSCOPIA_TESTE UREASE+ANATOMO PATOLOGICO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40201082, 40202615', "codigo_base_proc_principal": '40201082', "Descricao_REPASSE": 'Colonoscopia + Teste De Urease Para Pesquisa De Helicobacter Pylori'},
    'CPRE_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201074', "codigo_base_proc_principal": '40201074', "Descricao_REPASSE": 'Colangiopancreatografia Retrógrada Endoscópica'},
    'CPRE_ANATOMO PATOLOGICO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40201074, 40202038', "codigo_base_proc_principal": '40201074', "Descricao_REPASSE": 'Colangiopancreatografia Retrógrada Endoscópica Com Biópsia'},
    'CPRE_COLOCACAO DE PROTESE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40813320', "codigo_base_proc_principal": '40201074', "Descricao_REPASSE": 'Colocação De Stent Biliar'},
    'CPRE_PROTESE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40813320', "codigo_base_proc_principal": '40201074', "Descricao_REPASSE": 'Colocação De Stent Biliar'},
    'CPRE_RETIRADA DE PROTESE BILIAR': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201074', "codigo_base_proc_principal": '40201074', "Descricao_REPASSE": 'Colangiopancreatografia Retrógrada Endoscópica'},
    'ECOEDA ALTA C+PUNCAO_ANATOMO PATOLOGICO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202240, 40202038', "codigo_base_proc_principal": '', "Descricao_REPASSE": 'Ecoendoscopia Alta Com Punção E Biópsia'},
    'ECOEDA ALTA_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201104', "codigo_base_proc_principal": '40201104', "Descricao_REPASSE": 'Ecoendoscopia alta sem punção'},
    'ECOENDOCOPIA ALTA_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201104', "codigo_base_proc_principal": '40201104', "Descricao_REPASSE": 'Ecoendoscopia alta sem punção'},
    'ECOENDOSCOPIA BAIXA_': {"TipoCobranca": 'sem_mapeamento_tuss', "CodigosTUSS": '', "codigo_base_proc_principal": '', "Descricao_REPASSE": 'Não há código TUSS específico para baixa na tabela (apenas alta)'},
    'ECOENDOSCOPIA ALTA_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201104', "codigo_base_proc_principal": '40201104', "Descricao_REPASSE": 'Ecoendoscopia alta sem punção'},
    'ECOENDOSCOPIA ALTA_ANATOMO PATOLOGICO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40201104, 40202038', "codigo_base_proc_principal": '40201104', "Descricao_REPASSE": 'Ecoendoscopia Alta Com Biópsia E/Ou Citologia'},
    'ECOENDOSCOPIA ALTA_ANATOMOPATOLOGICO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40201104, 40202038', "codigo_base_proc_principal": '40201104', "Descricao_REPASSE": 'Ecoendoscopia Alta Com Biópsia E/Ou Citologia'},
    'ECOENDOSCOPIA_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201104', "codigo_base_proc_principal": '40201104', "Descricao_REPASSE": 'Ecoendoscopia alta sem punção'},
    'ECOENDOSCOPIA_ANATOMO PATOLOGICO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40201104, 40202038', "codigo_base_proc_principal": '40201104', "Descricao_REPASSE": 'Ecoendoscopia Alta Com Biópsia E/Ou Citologia'},
    'ENCOSCOPIA_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201120', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta'},
    'ENCOSCOPIA_TESTE UREASE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDDOSCOPIA_ANATOMO PATOLOGICO+TESTE UREASE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOCOSPIA_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201120', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta'},
    'ENDOSCOPIA_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201120', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta'},
    'ENDOSCOPIA_ENDOSCOPIA': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201120', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta'},
    'ENDOSCOPIA_UREASE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_ANATOMIA PATOLOGICA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202038', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E/Ou Citologia'},
    'ENDOSCOPIA_ANATOMO PATOLOGICA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202038', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E/Ou Citologia'},
    'ENDOSCOPIA_ANATOMO PATOLOGICO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202038', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E/Ou Citologia'},
    'ENDOSCOPIA_ANATOMO PATOLOGICO+POLIPECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202038, 40202550', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Alta Com Biópsia E Polipectomia'},
    'ENDOSCOPIA_ANATOMO PATOLOGICO+HP+CICATRIZ GASTRICA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_ANATOMO PATOLOGICO+TESTE DE UREASE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_ANATOMO PATOLOGICO TESTE UREASE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_ANATOMO PATOLOGICO+HPYLORI+BUBO GASTRICO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_ANATOMO PATOLOGICO+TESTE UREASE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_ANATOMO PATOLOGICO+POLIPO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202038, 40202550', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Alta Com Biópsia E Polipectomia'},
    'ENDOSCOPIA_ANATOMO+MUCOSECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202038, 40202470', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Alta Com Biópsia E Mucosectomia'},
    'ENDOSCOPIA_ANATOMO+POLIPECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202038, 40202550', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Alta Com Biópsia E Polipectomia'},
    'ENDOSCOPIA_ANATOMOPATOLOGICO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202038', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E/Ou Citologia'},
    'ENDOSCOPIA_ANATOMOPATOLOGICO+TESTE DE UREASE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_BIOPSIA HEPATICA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202038', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E/Ou Citologia'},
    'ENDOSCOPIA_COM DILATACAO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202186', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Dilatação De Esôfago Com Balão Pneumático'},
    'ENDOSCOPIA_CPRE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40201074', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Colangiopancreatografia Retrógrada Endoscópica'},
    'ENDOSCOPIA_DILATACAO DE OSTOMIA+TROCA DE GTT': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202283', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Gastrostomia endoscópica'},
    'ENDOSCOPIA_DILATACAO ESOFAGICA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202186', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Dilatação De Esôfago Com Balão Pneumático'},
    'ENDOSCOPIA_DILATACAO PNEUMATICA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202186', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Dilatação De Esôfago Com Balão Pneumático'},
    'ENDOSCOPIA_GASTROSTOMIA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202283', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Gastrostomia Endoscópica'},
    'ENDOSCOPIA_GASTROSTOMIA (1ª PASSGEM)': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202283', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Gastrostomia Endoscópica'},
    'ENDOSCOPIA_GTT': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202283', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Gastrostomia Endoscópica'},
    'ENDOSCOPIA_GTT 1ª PASSAGEM': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202283', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Gastrostomia Endoscópica'},
    'ENDOSCOPIA_HEMOCLIP': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202291', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Hemostasia mecânica do esôfago, estômago ou duodeno'},
    'ENDOSCOPIA_HEMOSTASIA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202291', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Hemostasia mecânica do esôfago, estômago ou duodeno'},
    'ENDOSCOPIA_HEMOSTASIA LEVE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202291', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Hemostasia mecânica do esôfago, estômago ou duodeno'},
    'ENDOSCOPIA_HPYLORI': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_HPYLORI+POLIPOS GASTRICOS+LIGADURA ELASTICA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202615, 40202550, 40202453', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Alta Com Urease, Polipectomia E Ligadura'},
    'ENDOSCOPIA_LIGADURA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202453', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Ligadura Elástica Do Esôfago, Estômago Ou Duodeno'},
    'ENDOSCOPIA_LIGADURA DE VARIZES ESOFAGO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202453', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Ligadura Elástica Do Esôfago, Estômago Ou Duodeno'},
    'ENDOSCOPIA_LIGADURA ELASTICA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202453', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Ligadura Elástica Do Esôfago, Estômago Ou Duodeno'},
    'ENDOSCOPIA_MUCOSECTOMIA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202470', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Mucosectomia Do Esôfago, Estômago Ou Duodeno'},
    'ENDOSCOPIA_PASSAGEM SNE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202534', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Passagem de sonda naso-enteral'},
    'ENDOSCOPIA_PASSAGEM DE SNE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202534', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Passagem de sonda naso-enteral'},
    'ENDOSCOPIA_PASSAGEM DE SONDA NASO ENTERAL': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202534', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Passagem de sonda naso-enteral'},
    'ENDOSCOPIA_PASSAGEM DE SONDA NASOENTERAL': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202534', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Passagem de sonda naso-enteral'},
    'ENDOSCOPIA_PASSAGEN SNE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202534', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Passagem de sonda naso-enteral'},
    'ENDOSCOPIA_PLASMA DE ARGONIO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40201376', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Aplicação de plasma de argônio por endoscopia digestiva alta'},
    'ENDOSCOPIA_POLIPECTOMIA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202550', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Polipectomia Do Esôfago, Estômago Ou Duodeno (Independente Do Número De Pólipos)'},
    'ENDOSCOPIA_REMOCAO DE BALAO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202577', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Retirada de corpo estranho do esôfago, estômago ou duodeno'},
    'ENDOSCOPIA_RETIRADA DE CORPO ESTRANHO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202577', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Retirada de corpo estranho do esôfago, estômago ou duodeno'},
    'ENDOSCOPIA_RETIRADA DE GTT': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202577', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Retirada de corpo estranho do esôfago, estômago ou duodeno'},
    'ENDOSCOPIA_TESTE UREASE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_TESTE DE UEASE+POLIPECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202615, 40202550', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Alta Com Urease E Polipectomia'},
    'ENDOSCOPIA_TESTE DE UREASE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_TESTE DE UREASE - NEGATIVO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_TESTE DE UREASE+ANATOMO PATOLOGICO+POLIPECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202615, 40202550', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Alta Com Urease E Polipectomia'},
    'ENDOSCOPIA_TESTE DE UREASE - POSITIVO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_TESTE DE UREASE+ANATOMO+POLIOECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202615, 40202550', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Alta Com Urease E Polipectomia'},
    'ENDOSCOPIA_TESTE DE UREASE+ANATOMO PATOLOGICO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_TESTE DE UREASE+POLIPECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202615, 40202550', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Alta Com Urease E Polipectomia'},
    'ENDOSCOPIA_TESTE DE UREASE+ANATOMO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_TESTE DE UREASE+ANATOMO+POLIPECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202615, 40202550', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Alta Com Urease E Polipectomia'},
    'ENDOSCOPIA_TESTE DE UREASE+POLIPECTOMIA+ANATOMO PATOLOGICO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202615, 40202550', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Alta Com Urease E Polipectomia'},
    'ENDOSCOPIA_TESTE DE UREASEA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_TESTE DE URESE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_TESTE E UREASE': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_TESTE UREASE - NEGATIVO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_TESTE UREASE - POSITIVO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_TESTE UREASE+ANATOMO PATOLOGICO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202615', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)'},
    'ENDOSCOPIA_TESTE UREASE+ANATOMO+POLIPECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202615, 40202550', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Endoscopia Alta Com Urease E Polipectomia'},
    'ENDOSCOPIA_TROCA DE GTT (BOTTON)': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202283', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Gastrostomia endoscópica'},
    'ENDOSCOPIA_TROCA DE GTT': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202283', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Gastrostomia endoscópica'},
    'ENDOSCOPIA_TROCA GTT': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202283', "codigo_base_proc_principal": '40201120', "Descricao_REPASSE": 'Gastrostomia endoscópica'},
    'EXAME REALIZADO_PROCEDIMENTOS ADICIONAIS': {"TipoCobranca": 'sem_mapeamento_tuss', "CodigosTUSS": '', "codigo_base_proc_principal": '', "Descricao_REPASSE": 'Erro de leitura de cabeçalho do PDF'},
    'GASTROSTOMIA 1ª PASSAGEM_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40202283', "codigo_base_proc_principal": '40202283', "Descricao_REPASSE": 'Gastrostomia Endoscópica'},
    'GASTROSTOMIA ENDOSCOPICA_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40202283', "codigo_base_proc_principal": '40202283', "Descricao_REPASSE": 'Gastrostomia endoscópica'},
    'GASTROSTOMIA ENDOSCOPICA_ANATOMO PATOLOGICO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202283, 40202038', "codigo_base_proc_principal": '40202283', "Descricao_REPASSE": 'Gastrostomia Endoscópica Com Biópsia E/Ou Citologia'},
    'GASTROSTOMIA ENDOSCOPICA_GASTROSTOMIA+ANATOMO PATOLOGICO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202283, 40202038', "codigo_base_proc_principal": '40202283', "Descricao_REPASSE": 'Gastrostomia Endoscópica Com Biópsia E/Ou Citologia'},
    'GASTROSTOMIA_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40202283', "codigo_base_proc_principal": '40202283', "Descricao_REPASSE": 'Gastrostomia Endoscópica'},
    'GTT_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40202283', "codigo_base_proc_principal": '40202283', "Descricao_REPASSE": 'Gastrostomia Endoscópica'},
    'HEMOSTASIA_POLIPECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202291, 40202550', "codigo_base_proc_principal": '', "Descricao_REPASSE": 'Endoscopia Alta Com Hemostasia E Polipectomia'},
    'LIGADURA ELASTICA_HEMOSTASIA LEVE': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202453, 40202291', "codigo_base_proc_principal": '', "Descricao_REPASSE": 'Endoscopia Alta Com Ligadura E Hemostasia'},
    'MUCOSECTOMIA_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40202470', "codigo_base_proc_principal": '40202470', "Descricao_REPASSE": 'Mucosectomia Do Esôfago, Estômago Ou Duodeno'},
    'MUCOSECTOMIA_POLIPECTOMIA': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202470, 40202550', "codigo_base_proc_principal": '40202470', "Descricao_REPASSE": 'Endoscopia Alta Com Mucosectomia E Polipectomia'},
    'PASSAGEM DE SNE_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40202534', "codigo_base_proc_principal": '40202534', "Descricao_REPASSE": 'Passagem de sonda naso-enteral'},
    'PASSAGEM DE SONDA POR ENDOSCOPIA_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40202534', "codigo_base_proc_principal": '40202534', "Descricao_REPASSE": 'Passagem de Sondas por Endoscopia'},
    'PASSAGEM SNE_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40202534', "codigo_base_proc_principal": '40202534', "Descricao_REPASSE": 'Passagem de sonda naso-enteral'},
    'RETIRADA DE PROTESE TRANSPAPILAR_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201074', "codigo_base_proc_principal": '40201074', "Descricao_REPASSE": 'Colangiopancreatografia Retrógrada Endoscópica'},
    'RETO_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201171', "codigo_base_proc_principal": '40201171', "Descricao_REPASSE": 'Retossigmoidoscopia Flexível'},
    'RETO_ANATOMO PATOLOGICO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202690', "codigo_base_proc_principal": '40201171', "Descricao_REPASSE": 'Retossigmoidoscopia flexível com biópsia e/ou citologia'},
    'RETOSIGMOIDECTOMIA_ANATOMO PATOLOGICO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202690', "codigo_base_proc_principal": '', "Descricao_REPASSE": 'Retossigmoidoscopia flexível com biópsia e/ou citologia'},
    'RETOSSIGMOIDECTOMIA FLEXIVEL_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201171', "codigo_base_proc_principal": '40201171', "Descricao_REPASSE": 'Retossigmoidoscopia Flexível'},
    'RETOSSIGMOIDECTOMIA FLEXIVEL_ANATOMO PATOLOGICO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202690', "codigo_base_proc_principal": '40201171', "Descricao_REPASSE": 'Retossigmoidoscopia flexível com biópsia e/ou citologia'},
    'RETOSSIGMOIDOSCOPIA FLEXIVEL_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201171', "codigo_base_proc_principal": '40201171', "Descricao_REPASSE": 'Retossigmoidoscopia Flexível'},
    'RETOSSIGMOIDOSCOPIA_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40201171', "codigo_base_proc_principal": '40201171', "Descricao_REPASSE": 'Retossigmoidoscopia Flexível'},
    'RETOSSIGMOIDOSCOPIA_ANATOMO PATOLOGICO': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202690', "codigo_base_proc_principal": '40201171', "Descricao_REPASSE": 'Retossigmoidoscopia flexível com biópsia e/ou citologia'},
    'RETOSSIGMOIDOSCOPIA_POLIPECTOMIA': {"TipoCobranca": 'unico_cod_tuss_inclui_proc_adicional_e_principal', "CodigosTUSS": '40202682', "codigo_base_proc_principal": '40201171', "Descricao_REPASSE": 'Retossigmoidoscopia Flexível Com Polipectomia'},
    'TROCA DE GTT_': {"TipoCobranca": 'unico_cod_tuss_somente_proc_principal', "CodigosTUSS": '40202283', "codigo_base_proc_principal": '40202283', "Descricao_REPASSE": 'Gastrostomia endoscópica'},
    'TROCA DE GTT_ANATOMO PATOLOGICO': {"TipoCobranca": 'multiplos_cod_tuss_proced_adicional', "CodigosTUSS": '40202283, 40202038', "codigo_base_proc_principal": '40202283', "Descricao_REPASSE": 'Gastrostomia Endoscópica Com Biópsia E/Ou Citologia'},
}

# Logo do Hospital São Camilo embutida em base64 (extraída do template original)
_LOGO_SAOCAMILO_B64 = (
    "/9j/4AAQSkZJRgABAQEAeAB4AAD/2wBDAAgGBgcGBQgHBwcJCQgKDBQNDAsLDBkSEw8UHRofHh0a"
    "HBwgJC4nICIsIxwcKDcpLDAxNDQ0Hyc5PTgyPC4zNDL/2wBDAQkJCQwLDBgNDRgyIRwhMjIyMjIy"
    "MjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjIyMjL/wAARCACOAYsDASIA"
    "AhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQA"
    "AAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3"
    "ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWm"
    "p6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEA"
    "AwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSEx"
    "BhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElK"
    "U1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3"
    "uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD3+iii"
    "gAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKA"
    "CiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooqrqG"
    "o2mlWT3d9OkECDlmP6D1PtQ3bccYuTUYq7ZaorgLP4lDU7zUksrDFvaWUtyjythnKdMgdAan8IfE"
    "vTPErpaXKix1Bvuxu2UkP+y3r7Hn61mqsHsztq5biqSbnDbfyudxRRRWhwhRRRQAUUUUAFFFFABR"
    "Wbo+oy6it6ZURfIu5IF291XGCffmkTUZW8SzaaUTyktVmDfxZLEY+nFAGnRWIL/UNTvLmLTDBDbW"
    "0hie4mQuXkHUKoI4HQknrU2nalcPfzaZqEcaXkSCVXizsljJxuGeQQeCKANWisG3v9U1dp5tPe1g"
    "tIpWiQzRs7SlTgngjaM5Hc8VZ1fUbrS9FW58uFrotHGRk7AzMFz645oA1aKwr291XRbc3t3JbXVo"
    "hHnCOMxuik43Dkg4z04q1LqUieIrTT1VDDNbyTFucgqVAx7c0AadFc5e32uWmp2NoJNPYXjuqt5T"
    "/LtXdz83NbSSy21g01/JFujUvI0akLgc8Ak9qALNFZHh/V5tVtpjdQrBcxON0an+BgGQ/ipH4g1J"
    "LqUsfiW200InlS2zzFudwKsAB9OaANOiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKK"
    "KACiiigAooooAKKKKACiioL68g0+xnvLl9kMKF3PoBQNJydluU9e16x8O6a97eyYUcJGPvSN6AV4"
    "P4j8TX/ia/NxdvtiUnyYFPyxj+p96TxL4iuvEurPeXBKxDKwQ54jX/H1NY9eZXruo7LY/QMoyiGD"
    "iqlTWo/w8l+rOp8Ex+ZJro/6hE/8hXAocBSCQRggjtXpfw4g86bXuOumOn5//qrzNOEH0qbe4vma"
    "OV8XVXbl/I9f+H3xLZ3i0bX5sscJb3jnqeyuf5N+det18jkZFe1fCrxpLqcR0HUZC91Am63lY8yR"
    "jqp9x/L6V1UK1/dkfO5vliinXorTqv1X6np1FFFdZ84FFFFABRRRQBymhrqhfVjZy2axf2jNxNGz"
    "NnjuGFS6cLseN7r7Y8DyfYEwYVKjG8+pPNamj6dLpy3oldG8+7knXbnhWxgH34pE06VfEs2pF08p"
    "7VYQvO7IYnP05oAp+Dv+QBz9/wC0T7/97zGzRe/8jxpW3732Sff/ALuVx+tSDT9R0y8uZdMFvNbX"
    "Mhle3mYoUkPUqwB4PXBHWptO025XUJtT1GSN7uRBEiRZ2Qxg52gnkknkmgCo2m6lo8s82jvFPbSO"
    "ZWsZvlwx5bY46Z9CMZqrr2ox6p4OhvrZTiSeEhH4IYSgFT6cjFaUzeIW8yKKLTl3EhJzI52jsdmO"
    "T+OKgn8OlfDUGk2sozFJG5kk/iIcOxOO55oAW/tNU1u3NlcQQ2dnIR57CXzHdQc7V4AGcdaS5AHj"
    "jTgOgspsf99JW/WZNp0sniO11EOgiht5ImU53EsVII9uKAKetf8AIyeHv+us3/os0eKrkCzt9PxI"
    "xvZgjrEhZvKHzPgDnoMfjVzUNOlu9W0u7R0CWjyM4OcncuBilXT5X8QvqMzoY0gEMCDOVycuT9cK"
    "PwoAyV1GGLxZazRw3MUN7EbaTzYGjG9fmTkjrjcKsXH/ACPtj/14S/8Aoa1e13TZNU0xoYHWO5R1"
    "lgkborqcgn+X41WvtP1FtdtdTtBasY7ZoXjldl5Yg5BAPpQBt0VXs2vGjb7bHAj5+UQuWGPfIFWK"
    "ACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAoornvFHjDT/DFv++PnXbj"
    "Mdsh+Y+59B70pSUVdmtGjUrTVOmrtnQ1S1fT01bSLuwkOFuImjz6Ejg/nXjE/wATPEst6biO5hij"
    "zxAsQKAenPJ+ua7Hw78U7K+ZLfWY1spjwJlOYmPv3X+XvWCxFOfunr1cjx2GSqxV7a6atf15XPJb"
    "yzn0+9ms7lCk8LlHX3FQV698SPCg1W0GvaaokuIk/fLHz5sf94Y6kfy+leQ1wVabpysfZZdjo4yg"
    "qi36rs/62PS/hHa+c2suejRpF+e6vI5UMU8sZ6o7KfwNe5fCK1MegXtyR/rrnaPoqj/GvH/E9p9h"
    "8V6tbYwEu5MfQnI/nW0o2pRZ5NKspZjXj6fgrGVXqPwc8PyTalca9KpEMCmCA/3nP3j+A4/GuE8N"
    "eHrvxPrUWnWgIB+aWXHESd2P9PU19AXGo6F4F0O3tJJVhhhTbFCvzSSepx3JPJNVQgr88tkYZtiZ"
    "cv1akrzl0Xb/AIJ0VFeMaz8VNXvJGXTI47GDsSA8h+pPA/AVe8M/FOeOVLbXwJImOBdIuGX/AHlH"
    "Ue4rdYqm3Y8mXD+NjS9pZX7X1/y/E9ZoqOCeK5gSeCRZInG5HQ5DD1BqSug8Vpp2YUUUUCCioLy9"
    "tdPtmuLyeOCFeryNgVhDx54cL4+34X++YnC/nigDpKRmVFLMQAO5NRWt3b3tutxazRzQv910bINN"
    "vrG21K0ktLyISwSY3IScHBz2oAk+0Qf89o/++hTkljkJ2OrY9DmuUuvDngyyu7e0ubO2iuLk4hjZ"
    "2y59ua29M0DS9GkkfT7NIGkADlSTkDp1NAGlRRVLUdX0/SIRLf3cVuh6bzyfoOpoAu0VzsXjnw7J"
    "IEN/5eTgNLEyKfxIxXQRyJLGskbq6MMqynII9jQA6iiq95fWunW5nvLiOCIdWkbAoAsUVzQ8feGj"
    "Ls/tEf73lPj88VvWd7bX9utxaTxzwt0eNsigCeiisu98R6Rpt4LS8v4obggEI2c89KANSiiigAoq"
    "jqWs6do6xtqF3HbiUkIXzyR1qeyvbbUbRLq0mWaB87XXoe1AE9FFZeoeJNH0q5+z31/FBNtDbGzn"
    "B70AalFNR1kRXQgqwBBHcVm3viLSNOvRZ3d9FFcHBEbZzz07UAalFRzzx20Ek8zhIo1LOx6ADkmq"
    "ema5pmsGQafeR3Hl437M8Z6fyoA0KKKCQASTgCgAorn7zxt4esZTFLqUbOpwREpfH4gYq5pniPSN"
    "YbZY30UsmM+X91vyPNAGpRRRQAUUVR1q9uNP0a7u7W1e6uI4yY4UBJdug4FJuyuVCLnJRW7Of8a+"
    "NYfDVt9nt9supyrlIz0jH95v6DvXh93d3F9dSXV1M808p3O7nJJrVu9E8TajeS3dzpOpS3Ezbnc2"
    "7cn8ulNXwl4ibpol/wDjCRXmVZzqPbQ/QstwuFwNO3OnJ7u6/qxjUV0EfgfxPIeNGuB/vFV/mauw"
    "/DTxRNjNlFEPV51/pms1Tm+jO6WPwsd6kfvRB4X8baj4akWIE3NgT81s56e6nsf0q74j8P2uqQHx"
    "B4YVp7OVv9ItEX95buf9kdj/AJ4q/a/CLVpCDdX9pCO4QM5/pXdeFPA9n4Vklniup7i5lTY7N8q4"
    "znhR/XNdEKVSS5ZrT8jwsZmWCoT9vhpe/wBUtpevT0e/qW/B2kPonhWxs5V2zbN8o9GY5I/DOPwr"
    "zL4j+DtTvfHMUumWbzLqKr8yj5UdRg7j2GAD+de1VHPCtxbyQvnZIpRtpwcEY4PauuVJOCj2PmaG"
    "YVKWIliOsr3+Z5D/AG5p3w+0p9H0Ix3mrP8A8fl6RlFf0HrjsOg781wd3d3N/dPc3c7zzuctI5yT"
    "Xp2ofB9CzNpuqMi54juE3Y/4EP8ACsK4+FXiKEnyjZzj/ZlK/wAxXFUp1Xo1ofXYDGZbTTlGouZ7"
    "t6N/8DyWhxFFdNL8PvFMPXS2f/clQ/1qq/g3xInXRLz8Ez/KsPZz7M9aONw0tqkfvRoeDPGtz4Zu"
    "RBOXm0yRvni6mM/3l/qO9e5Wt1Be2sVzbSrLBKoZHU5BFfO7eFvEK9dE1D/wHau5+G82v6TqJ0u9"
    "0y+TT58srSQsFhfGc5PQH+eK6sPUlF8klofPZ5gcPWg8TRkuZb6rX/g/meq0UUV3nxpwaxr4n+Il"
    "3BejzLLSkHlwN91n45I79/yFdubaBovKMMZjxjYUGMfSuK1qz1Hw34nk8R6dbNd2lwu27gT7w9x+"
    "QNbuj+L9G1orHb3YS4P/ACwm+V8+nPX8KANe1tLexgEFrCkMQJIRFwAScnipqKyvEmqDRvD95e5w"
    "6RkR+7ngfqaAOG1uCbxDda/rEBJ/svZHaEf3kO5yP1r0HR9QTVdItL6PpNGGPse4/PNZ3hPSRp/h"
    "W2tZlzJMhknB7s/Jz+BxWV4GkbT7nVfDsp+aynLw57xt/kfnQB1d/eR6fYXF5McRwxl2/AVyHhDS"
    "/wC2mfxNrCCe5uHP2ZHGViQHAwP8/rV74hztB4Muwpx5jIh+hYf4VsaBAtv4e06JRgLbR/8AoINA"
    "Fq5srW8gaC5t45YmGCjqCMVxmjs/hTxg2gGRm029Uy2gc58tu6j8iPyruq4Xx+fs2reHb1eHS625"
    "9sqaAO2uJ47W3lnlbbHGpdj6ADJryWziu/iJ4pkkupHjsYRu2qf9WmeFH+0e5+tdz4+uGt/B17tO"
    "DJtj/AsM/pWT8LbdU0O8uMfNJcbSfZVH+Jpgb8fg7w9HAIRpVuVxjLLlj+PWr+l6VZ6NZi0sYvKh"
    "DFtuc8k+pq7RSARmVEZmICqMknsK8Q1GKfxG2u6+N3lQSJtH+yTgfkADXpXjzVf7L8LXGxsS3P7h"
    "Px6n8s1W8IaDGngX7LOmGv0Z5MjswwP0xTA1PCWqf2v4as7ljmVV8uT/AHl4P+P41t15n8Nb17HV"
    "dQ0O4OGyXUH+8pww/LB/CvTKQHm3xJBvNc0bTlODJn8CzBR/Kr3wzvGWyv8ASZTiS0nJAPoeD+oP"
    "51Q19vtnxY0yDqIfK4+mXqWL/iQfFh0+7BqS5Hpluf8A0IH86YHoteGeJJJNY1DVdYBJgS5W3Q+2"
    "Dj9F/WvXfE+o/wBleG767Bw6xFU/3jwP1NefyaQbX4SCVl/eSzrct64J2j9P50ID0Tw9cfavDmnT"
    "ZyWt0z9QMV5p48/5H2H/AHYf513HgCfz/BtkO8e6M/gxrh/Hn/I+w/7sP86APTPEH/Iual/16yf+"
    "gmuF+FH39U+kf/s1d14g/wCRc1L/AK9ZP/QTXC/Cj7+qfSP/ANmoA9Lrzr4j69cCeHQbJ2DSgNPt"
    "OC2ThU/Hqfwr0WvIpT9u+LYEnIW9A59EHH8qQHbaF4H0nS7CNLi0huroqDLJKobn0APAFc9448J2"
    "+mWq63o6G1kgcGVYjgAE8MPQg4r0is3xDAtx4c1KJhkG2f8ARSaAKfhDXTr+gRXEuPtMZ8qbHdh3"
    "/EYNb1eafCidt2pQZ+UiOT8eRXpdABWdrunzarod5Y291JazTRlY5o2IKN1ByOeorRopNXVioycZ"
    "KS3R8w32peJdJv5rG81PUYbmFtrobl+vqOeQfWo08U+IY/u65qA/7eG/xr3Lx14Ft/Fdn50O2HVI"
    "VxFKejj+43t79q+f76xutMvZbO9geC4iOHjccj/Ee9cFSEoPfQ+0wGJoYyHwrmW6t/Whrx+OPFMX"
    "3devT/vPu/mKvwfE7xdBj/iaCQeksCH+lcjRWanJdTslhMPLeC+5Hotn8ZdfhIF1Z2NyO+AyH9CR"
    "+leh+DPH9t4vlmt1sZ7a4hQO+fnTGccMO/sa8l8I/D/U/FDrcODaaYD81y45ceiDv9elaviLXbPT"
    "7Y+HfC+bfToW/fXCN89y47lupH8/pW0as4Lmk9Dy62XYXE1PYUI+91a2j6+fl+R7rUc8y29vLO4Y"
    "rGhchFySAM8DuaxPBmrPrXhSxu5W3ThfLlPqynBP44z+NeYfEnxbqVt46jj068khGmqu0K3ys5GW"
    "3DoeCBzXTKqlBS7ng0MvqVcRLD7ON7/Iv6l8a33sul6QNoJAe6k5P/AV6fnXPXHxa8VTk+XNaW4P"
    "aOAH/wBCJrYk8P6b8SdLk1jRfKsddj/4/LQnEcjevtnsfz9a83v9Pu9LvJLO+t5Le4Q/MjjB+o9R"
    "7iuWc6i1vofRYTDYGV4KCUlunq/x6eaN6b4heLJyd2tTr/uKq/yFU38YeJZPv69qB+kxH8qxaKy5"
    "5dz0VhqMdoL7kabeJNdb72tagf8At5f/ABr0T4V6frmq6l/bV9qN82nwbljWSdiJnIx0J5A/nXOe"
    "BPAFz4puFu7tXh0lG+Z+jTH+6vt6mvf7W1gsrWK2toligiUKiIMBQO1dFCnJvmex4mbY2lTi6FJL"
    "me/l/wAElooorsPlwrA17wjpeuQuXgWG7xlLiIbWB98dfxq+mu6VJdzWg1C3FxC22SNnCkH8etVt"
    "Y8UaVo1m8013FJIB8kMbhmc9hgfzoAy/AerXd7Y3en37mS50+XyjITksvOM+vQ/pVLxvJcarrOl6"
    "BZRpNJu+1Sxu21SF6AnsOv6Vb8B6Zc2Wm3mpX6mKe/lMxVuCq8kZ9OpP5VB4NB1jXdY8ROCUkk+z"
    "2+eyL/kUwNH7Z4wHTSdM/wDAlv8ACucvZ9V0jxpp2tapa29tHcn7LKYJS4Yepz07flXpFc9420s6"
    "r4Wu40GZYR50eOuV5/lmkBX+IcDT+DLzaM+WUf8AAMP8a1/D863Ph3TpVOQ1sn6KBVDQLqLxP4Ni"
    "E53edCYJv94DB/xrF8Lav/wjcj+GtckEDwuTazvwkqE9jQB3dcL4+H2nV/DlinLvdbse2VrrLnWd"
    "Ms7dp7i/t0jAzkyDn6etcpoiTeKfFjeI5Inj061UxWQcYLnu36n9PSgDU8e2zXPg692jJj2yY9gw"
    "z+mayPhZcq+jXltn54p92PZgP8DXcTwx3EEkEqho5FKsp7gjBryOP7d8OvE7M8TS2MuVDDpLHnjB"
    "/vD0/wAaYHsFFc3H498NyW4lOoBDjJR423D2xir9h4hsNS0WXVbeRvs0W/fuGCNvXj/PWkB5/wDE"
    "O9k1PxLa6VbxvOLZRmKPks7ckD/gIH51vR+MNZijWNPB98qIAqjJ4A/4DWP8PoJNY8TahrtwMlSS"
    "pP8Aff8AwXj8a9PpgeK3up3Vh41h1ubTZ9PLyCRoZc5YdHI4HXmvaEdZEV0IKsAQR3FcV8TdM+1a"
    "DFfKuXtJPm/3G4P64rQ8Bar/AGl4WgV2zLanyHyew6H8sUAc1p5+2/GG5c8iEuPptTb/ADq98TLV"
    "4YtN1mEYltZtpI9+R+o/Ws7wCftnjbV7w8/LIQf95/8A61dz4n07+1PDd9agZdoiyf7w5H8qAOS8"
    "dah/bFpoemWrZOouspA9DgD9Sfyro/FNki+CL61jGEitsKPQLg/0rg/AEUur+JLaaf5o9NtsJnty"
    "Qo/U/lXqOqw/aNIvYcZ3wOv5qaAOS+F1x5nh24hzzFcn8iAa5rx5/wAj7D/uw/zrS+E8/wA2p256"
    "kRyY/MH+lZvjz/kfYf8Adh/nQB6Z4g/5FzUv+vWT/wBBNcL8KPv6p9I//Zq7rxB/yLmpf9esn/oJ"
    "rzX4d65puitfnULpYBKE2bgTnGc9B70Aet15DcD7B8Wg0nyq16rZ9nH/ANeu/i8a+HZpUij1ONnd"
    "gqjY3JJwB0rm/iN4euJni1yxRmkhULMEGWAByrj6d6APQ6zPEc623hvUpWOALZx+Yx/WsDQ/iHpN"
    "5Yx/2jcC1u1UCQMp2sfUEfyrC8Y+L4tdhTRdEWS485x5jqpG/B4VR169T7UgJfhRbsDqVwR8oEcY"
    "+vJ/wr0qsPwloR0DQYrWTBuHJkmI/vHt+AwK3KACiiigArnfFXg3TPFloEul8q6QfurmMfOnsfUe"
    "xroqKTSasy6dSdKSnB2aPAZ/hJ4ojvzBFHbTQ54uBMFXHuDyPyNdx4Z+Emm6WyXOsSDULleRHjEK"
    "n6dW/H8q9GqlrGoppOj3eoONy28TPj1IHA/OslRhHU9Oea4vEJUk7X000b/ryOE+JXiz+z7b+wNO"
    "cJNIg+0MnHlx9lHoT/L615HU13dzX15Nd3Ll55nLux7k1DXm1ajqSuz7nL8FDB0FTjv1fdnsPwiu"
    "jJoN9bE/6m53D6Mo/wAK8f8AE139u8Vatc5z5l3Jj6BsD+Vek/CO58l9ZQ9BEkv5bq8jlcyzSSHq"
    "7lj+JreUr0oo8mlRUcxry/w/irml4d1+78NazDqVmcshxJHniVO6n/PBr6Bn0/QPH3h+3uZoFngm"
    "TdFIOJIj3APUEHqK+aq9P+DmvSQarcaFIxMFwhmhH911+9+Y/lVUJ2fK9mc+b4Vyh9Yp6Sj+X/AI"
    "dc+DurWkrPo9xHewE8JIwjkH9D+laHhT4Qyecl34kZQinK2cTZ3f77Dt7D869forpWHgnc8OWc4u"
    "VPkv8+oyGGO3hSGGNY4kAVUUYCj0Ap9FFbHlbhRRRQBlah4b0bVZPMvdOglkPVyuGP4jmo7LwnoO"
    "nTCa20yBZR0ZgWI+mc4rZooAbJGk0TxSKGRwVZT3BqKzsrXT7ZbazgSCFSSEQYAzU9FABSEBlIIy"
    "CMEGlooAq2Om2WmRNFZW0cEbNuKxjAJ9aL7TbLU4fJvrWK4j7CRc4+npVqigDBh8FeHIJRImkwbg"
    "cjdlgPwJxW6qqihUUKoGAAMAClooAKhubW3vYGguYI5om6pIoYfrU1FAHPjwR4bEm/8AsqHPXGWx"
    "+Wa1v7MsfsDWItYhaMMGFVAUj6CrVFAFWx02y0yForG2jt42bcVjXAJ9atUUUARXFvDd28lvcRrJ"
    "DINrowyCKr2GkafpiyLY2kVuJMbxGMbsetXaKAKNho2naW8j2NnDA0gAcxrgt9avUUUAU7HSbDTD"
    "KbK0igMpy5jXG76/nVsgMCCMg8GlooAoWGiaZpkry2NlDbu42s0a4JHWku9C0q+uxdXVhBNOMYkd"
    "cnjpWhRQAyaGO4heGVA8cilWU9CD1FZH/CI+H/8AoEWv/fFbVFAGPH4V0GKRJE0q2V0YMpCdCOhr"
    "YoooAxbvwjoF9KZZ9LgMhOSygrn8sVa07QtL0nJsLGGBjwWVfm/M81oUUAFFFFABRRRQAUUUUAFQ"
    "XlpDf2c1pcIHhmQo6nuDU9FA02ndHzr4n8N3XhnVWtZgXgbLQTY4kX/EdxWLX0lreiWWv6a9lfR7"
    "kblWH3kbswPrXhPibwtf+GL7yrlS9u5/c3Cj5XHp7H2rzK9Bwd1sff5Rm8MXFU6jtUX4+a/VGz8O"
    "Z/Jm17nppkj/AJf/AK680T7o+ld94KkMcmun/qET/wAhXBIpbaqgljgAAZJPpU39xfM1cbYuq+/L"
    "+Qp4r234W+CpNJgOuajGUvLhNsETDmKM9z7n9BVT4ffDQ2rRazr0I88Ya3tGGdnoz+/t2r1auuhR"
    "t70j53Ns0U06FF6dX+iCiiiuo+cCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAC"
    "iiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACq99YWupWklpe"
    "QJPBIMMjjIqxRRuOMnF3Tszg7T4aw6beai9lfsILuzktljlTcYy465zyBVnwj8N9K8MMt1IxvdQA"
    "4nkXAT/cXt9etdnRWapQWyOyrmOKqpqc99/OwUUUVocQUUUUAFFFFABRRRQAUUUUAFFFFABRRRQA"
    "UUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAf/Z"
)


def _criar_workbook_formulario_template():
    """Cria o workbook do formulário de cobrança programaticamente.

    Reproduz fielmente o layout do template original (mesmas merges, estilos,
    larguras, bordas e logo do Hospital São Camilo) sem depender de arquivo externo.
    Retorna um openpyxl.Workbook pronto para receber preenchimento a partir da linha 11.
    """
    import base64
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Plan1"

    # ── Larguras de coluna (idênticas ao template original) ────────────────
    col_widths = {
        "A": 12.5, "B": 14.33, "C": 49.5, "D": 23.5, "E": 26.33,
        "F": 21.5,  "G": 33.33, "H": 12.0, "I": 41.33, "J": 18.66,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── Alturas de linhas (cabeçalho) ──────────────────────────────────────
    for r in (1, 3, 4, 6):
        ws.row_dimensions[r].height = 12
    ws.row_dimensions[7].height = 15.75

    # ── Estilos reutilizáveis ──────────────────────────────────────────────
    f8b      = Font(name="Arial", size=8, bold=True)
    f8       = Font(name="Arial", size=8)
    center   = Alignment(horizontal="center", vertical="center")
    left_al  = Alignment(horizontal="left",   vertical="center")
    thin     = Side(style="thin")
    b_all    = Border(left=thin, right=thin, top=thin, bottom=thin)
    b_top_bot = Border(top=thin, bottom=thin)
    b_left   = Border(left=thin, top=thin, bottom=thin)
    b_right  = Border(right=thin, top=thin, bottom=thin)
    fill_azul = PatternFill("solid", fgColor="DDEBF7")  # aproxima o fill de tema das colunas I/J

    # ── Linha 2–3: Título do hospital (merge C2:J3) ────────────────────────
    ws.merge_cells("C2:J3")
    ws["C2"].value     = "HOSPITAL  SÃO CAMILO "
    ws["C2"].font      = f8b
    ws["C2"].alignment = center

    # ── Linha 5: Título do formulário (merge C5:J5) ───────────────────────
    ws.merge_cells("C5:J5")
    ws["C5"].value     = "FORMULÁRIO PARA SOLICITAÇÃO DE REVISÃO DE PROCEDIMENTOS NÃO REPASSADOS "
    ws["C5"].font      = f8b
    ws["C5"].alignment = center

    # ── Linha 7: campos de cabeçalho ──────────────────────────────────────
    ws.merge_cells("A7:C7")
    ws["A7"].value     = "EMPRESA"       # sobrescrito pelo código chamador
    ws["A7"].font      = f8b
    ws["A7"].alignment = left_al

    ws.merge_cells("D7:F7")
    ws["D7"].value     = "MÉDICO RESPONSÁVEL"   # sobrescrito
    ws["D7"].font      = f8b
    ws["D7"].alignment = left_al

    ws["G7"].value     = "DATA: "               # sobrescrito
    ws["G7"].font      = f8b
    ws["G7"].alignment = left_al

    ws.merge_cells("I7:J7")
    ws["I7"].value     = "PRAZO DE RETORNO AO PRESTADOR: 30 dias"
    ws["I7"].font      = f8b
    ws["I7"].alignment = left_al

    # ── Linha 9: seções (merge A9:H9 / I9:J9) ────────────────────────────
    ws.merge_cells("A9:H9")
    ws["A9"].value     = "PREENCHIMENTO PRESTADOR"
    ws["A9"].font      = f8b
    ws["A9"].alignment = center
    # bordas externas do bloco A9:H9
    for c in range(1, 9):
        cell = ws.cell(9, c)
        cell.font      = f8b
        cell.alignment = center
        left_s  = thin if c == 1 else Side()
        right_s = thin if c == 8 else Side()
        cell.border = Border(top=thin, bottom=thin, left=left_s, right=right_s)

    ws.merge_cells("I9:J9")
    ws["I9"].value     = "PREENCHIMENTO CONTAS MÉDICAS"
    ws["I9"].font      = f8b
    ws["I9"].alignment = center
    ws["I9"].fill      = fill_azul
    for c in (9, 10):
        cell = ws.cell(9, c)
        cell.fill   = fill_azul
        cell.font   = f8b
        left_s  = thin if c == 9 else Side()
        cell.border = Border(top=thin, bottom=thin, left=left_s, right=thin)

    # ── Linha 10: cabeçalhos das colunas de dados ─────────────────────────
    headers = [
        "DATA ", "Nº ATEND.", "PACIENTE", "CONVÊNIO", "PRESTADOR",
        "CÓDIGO", "PROCEDIMENTO", "Função", "OBSERVAÇÃO CONTAS MÉDICAS", "VALOR",
    ]
    for c, hdr in enumerate(headers, 1):
        cell = ws.cell(10, c)
        cell.value     = hdr
        cell.font      = f8b
        cell.alignment = center
        cell.border    = b_all
        if c in (9, 10):
            cell.fill  = fill_azul

    # ── Logo do Hospital São Camilo ───────────────────────────────────────
    try:
        logo_bytes = base64.b64decode(_LOGO_SAOCAMILO_B64)
        img        = XLImage(BytesIO(logo_bytes))
        img.width  = 93   # convertido de EMU: 885825 ÷ 9525
        img.height = 49   # convertido de EMU: 462915 ÷ 9525
        ws.add_image(img, "A1")
    except Exception:
        pass  # falha silenciosa: logo não é crítica para o formulário

    return wb


def _normalizar_chave_tuss(s: str) -> str:
    """Normaliza chave CONCATENAR: remove acentos, colapsa separadores e espaços."""
    import unicodedata
    s = str(s).strip().upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"[+/;,]+", "+", s)
    s = re.sub(r"\s*\+\s*", "+", s)
    return re.sub(r"\s+", " ", s).strip()


def _gerar_tabela_tuss(
    caminho_xlsx: Path = _TUSS_XLSX_PATH,
    caminho_saida: Path = _TUSS_LOOKUP_PATH,
) -> pd.DataFrame:
    """
    Lê TUSS (ATUALIZADO).xlsx e gera tuss_lookup_table.csv com classificação de tipo
    de cobrança. Retorna o DataFrame gerado. Chame sempre que o XLSX for atualizado.
    """
    import unicodedata

    df = pd.read_excel(caminho_xlsx, sheet_name=0)
    df.columns = ["Proc_PRODUCAO", "ProcAdic_PRODUCAO", "CONCATENAR", "CodigoTUSS", "Descricao_REPASSE"]
    # Remover linha de cabeçalho duplicado que entra como dado
    df = df[~df["Proc_PRODUCAO"].astype(str).str.strip().isin(["EXAME REALIZADO 1", "EXAME REALIZADO 2"])]

    # Mapa código base por procedimento principal (linha sem adicional)
    codigos_base: dict[str, str] = {}
    sem_adic = df[df["ProcAdic_PRODUCAO"].isna() | (df["ProcAdic_PRODUCAO"].astype(str).str.strip() == "")]
    for _, r in sem_adic.iterrows():
        proc_norm = _normalizar_chave_tuss(str(r["Proc_PRODUCAO"]))
        raw = str(r["CodigoTUSS"]).strip()
        sem_tuss = raw.lower() in ("nan", "sem correspondência exata tuss", "sem correspondência")
        if not sem_tuss:
            codigos_base[proc_norm] = raw.split(",")[0].strip()

    # Descrições corrigidas por conjunto de códigos (base de conhecimento curada)
    _CORRECOES_DESC: dict[frozenset, str] = {
        frozenset(["40201082"]): "Colonoscopia (Inclui A Retossigmoidoscopia)",
        frozenset(["40202038"]): "Endoscopia Digestiva Alta Com Biópsia E/Ou Citologia",
        frozenset(["40202186"]): "Dilatação De Esôfago Com Balão Pneumático",
        frozenset(["40202453"]): "Ligadura Elástica Do Esôfago, Estômago Ou Duodeno",
        frozenset(["40202470"]): "Mucosectomia Do Esôfago, Estômago Ou Duodeno",
        frozenset(["40202542"]): "Polipectomia De Cólon (Independente Do Número De Pólipos)",
        frozenset(["40202550"]): "Polipectomia Do Esôfago, Estômago Ou Duodeno (Independente Do Número De Pólipos)",
        frozenset(["40202615"]): "Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)",
        frozenset(["40202666"]): "Colonoscopia Com Biópsia E/Ou Citologia",
        frozenset(["40202712"]): "Colonoscopia Com Mucosectomia",
        frozenset(["40813320"]): "Colocação De Stent Biliar",
        frozenset(["40201082", "40202615"]): "Colonoscopia + Teste De Urease Para Pesquisa De Helicobacter Pylori",
        frozenset(["40201074", "40202038"]): "Colangiopancreatografia Retrógrada Endoscópica Com Biópsia",
        frozenset(["40201104", "40202038"]): "Ecoendoscopia Alta Com Biópsia E/Ou Citologia",
        frozenset(["40202283", "40202038"]): "Gastrostomia Endoscópica Com Biópsia E/Ou Citologia",
        frozenset(["40202666", "40202542"]): "Colonoscopia Com Biópsia E Polipectomia",
        frozenset(["40202666", "40202712"]): "Colonoscopia Com Biópsia E Mucosectomia",
        frozenset(["40202666", "40202712", "40202542"]): "Colonoscopia Com Biópsia, Mucosectomia E Polipectomia",
        frozenset(["40202666", "40202135"]): "Colonoscopia Com Biópsia E Tatuagem",
        frozenset(["40202542", "40202712"]): "Colonoscopia Com Polipectomia E Mucosectomia",
        frozenset(["40202038", "40202550"]): "Endoscopia Alta Com Biópsia E Polipectomia",
        frozenset(["40202038", "40202470"]): "Endoscopia Alta Com Biópsia E Mucosectomia",
        frozenset(["40202615", "40202550"]): "Endoscopia Alta Com Urease E Polipectomia",
        frozenset(["40202615", "40202550", "40202453"]): "Endoscopia Alta Com Urease, Polipectomia E Ligadura",
        frozenset(["40202240", "40202038"]): "Ecoendoscopia Alta Com Punção E Biópsia",
        frozenset(["40202453", "40202291"]): "Endoscopia Alta Com Ligadura E Hemostasia",
        frozenset(["40202291", "40202550"]): "Endoscopia Alta Com Hemostasia E Polipectomia",
        frozenset(["40202470", "40202550"]): "Endoscopia Alta Com Mucosectomia E Polipectomia",
    }

    registros = []
    for _, r in df.iterrows():
        proc_raw  = str(r["Proc_PRODUCAO"]).strip()
        adic_raw  = str(r["ProcAdic_PRODUCAO"]).strip() if pd.notna(r["ProcAdic_PRODUCAO"]) else ""
        concat_raw = str(r["CONCATENAR"]).strip()
        codigos_raw = str(r["CodigoTUSS"]).strip()

        chave_norm = _normalizar_chave_tuss(concat_raw)
        proc_norm  = _normalizar_chave_tuss(proc_raw)
        sem_tuss   = codigos_raw.lower() in ("nan", "sem correspondência exata tuss", "sem correspondência")

        if sem_tuss:
            codigos_lista, tipo = [], "sem_mapeamento_tuss"
        else:
            codigos_lista = [c.strip() for c in codigos_raw.split(",") if c.strip()]
            base = codigos_base.get(proc_norm)
            adic_vazio = not adic_raw or adic_raw in ("nan",)
            if adic_vazio or (len(codigos_lista) == 1 and base and codigos_lista[0] == base):
                tipo = "unico_cod_tuss_somente_proc_principal"
            elif len(codigos_lista) > 1:
                tipo = "multiplos_cod_tuss_proced_adicional"
            else:
                tipo = "unico_cod_tuss_inclui_proc_adicional_e_principal"

        chave_frozen = frozenset(codigos_lista)
        desc = _CORRECOES_DESC.get(chave_frozen, str(r["Descricao_REPASSE"]).strip())

        registros.append({
            "chave_norm":              chave_norm,
            "Proc_PRODUCAO_raw":       proc_raw,
            "ProcAdic_PRODUCAO_raw":   adic_raw,
            "CONCATENAR_raw":          concat_raw,
            "CodigosTUSS":             ", ".join(codigos_lista),
            "QtdCodigos":              len(codigos_lista),
            "TipoCobranca":            tipo,
            "codigo_base_proc_principal": codigos_base.get(proc_norm, ""),
            "Descricao_REPASSE":       desc,
        })

    df_out = pd.DataFrame(registros)
    df_out.to_csv(caminho_saida, index=False, encoding="utf-8-sig")
    logger.info(f"tuss_lookup_table.csv gerado: {len(df_out)} linhas, {df_out['chave_norm'].nunique()} chaves únicas")
    return df_out


def _normalizar_convenio(s: str) -> str:
    """Normaliza nome de convênio: maiúsculas, sem acentos, espaços colapsados."""
    import unicodedata as _ud
    s = str(s).upper().strip()
    s = _ud.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s)


def _gerar_valores_tuss(
    df_correlacao: pd.DataFrame,
    caminho_saida: Path = _TUSS_VALORES_PATH,
    ano: int | None = None,
) -> pd.DataFrame:
    """
    Calcula estatísticas de valor por (Convênio, Código TUSS) a partir do
    DataFrame de correlação e atualiza (ou cria) tuss_valores.csv.

    Estrutura gerada:
        Ano | Convenio | CodigoTUSS | Descricao | Qtd | Media | UltimoValor | DataUltimo | Confianca

    Uma linha por (Convênio normalizado, Código) + linha Convenio="GERAL" como
    fallback agregado por código.  Chamada automaticamente pela aba Gerar Cobrança
    quando o arquivo não existe.
    """
    import unicodedata as _ud

    # ── Descrições oficiais ────────────────────────────────────────────────
    _DESC_OFICIAL: dict[str, str] = {
        "40201074": "Colangiopancreatografia Retrógrada Endoscópica",
        "40201082": "Colonoscopia (Inclui A Retossigmoidoscopia)",
        "40201104": "Ecoendoscopia Alta Sem Punção",
        "40201120": "Endoscopia Digestiva Alta",
        "40201171": "Retossigmoidoscopia Flexível",
        "40201376": "Aplicação De Plasma De Argônio Por Endoscopia Digestiva Alta",
        "40202038": "Endoscopia Digestiva Alta Com Biópsia E/Ou Citologia",
        "40202135": "Colonoscopia Com Magnificação E Tatuagem",
        "40202143": "Descompressão Colônica Por Colonoscopia",
        "40202186": "Dilatação De Esôfago Com Balão Pneumático",
        "40202240": "Ecoendoscopia Alta Com Punção",
        "40202283": "Gastrostomia Endoscópica",
        "40202291": "Hemostasia Mecânica Do Esôfago, Estômago Ou Duodeno",
        "40202313": "Hemostasias De Cólon",
        "40202453": "Ligadura Elástica Do Esôfago, Estômago Ou Duodeno",
        "40202470": "Mucosectomia Do Esôfago, Estômago Ou Duodeno",
        "40202534": "Passagem De Sonda Naso-Enteral",
        "40202542": "Polipectomia De Cólon (Independente Do Número De Pólipos)",
        "40202550": "Polipectomia Do Esôfago, Estômago Ou Duodeno (Independente Do Número De Pólipos)",
        "40202569": "Retirada De Corpo Estranho Do Cólon",
        "40202577": "Retirada De Corpo Estranho Do Esôfago, Estômago Ou Duodeno",
        "40202615": "Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)",
        "40202666": "Colonoscopia Com Biópsia E/Ou Citologia",
        "40202682": "Retossigmoidoscopia Flexível Com Polipectomia",
        "40202690": "Retossigmoidoscopia Flexível Com Biópsia E/Ou Citologia",
        "40202712": "Colonoscopia Com Mucosectomia",
        "40813320": "Colocação De Stent Biliar",
    }

    def _desc(cod, desc_real_map):
        return desc_real_map.get(cod) or _DESC_OFICIAL.get(cod) or f"Código TUSS {cod}"

    def _confianca(sub: pd.Series) -> str:
        qtd = len(sub)
        if qtd == 0:
            return "Sem dados"
        if qtd == 1:
            return "Baixa"
        std = float(sub.std())
        media = float(sub.mean())
        desvio_pct = (std / media * 100) if media > 0 else 100
        if qtd >= 5 and desvio_pct <= 5:
            return "Alta"
        if qtd >= 5 and desvio_pct <= 20:
            return "Boa"
        if qtd >= 2:
            return "Moderada"
        return "Baixa"

    # ── Inferir ano dominante ──────────────────────────────────────────────
    if ano is None:
        col_data = next((c for c in ["Data_REPASSE", "Data_PRODUCAO"] if c in df_correlacao.columns), None)
        if col_data:
            try:
                # Usa datetime parsed para ser robusto a qualquer formato de data
                _datas_parsed = pd.to_datetime(
                    df_correlacao[col_data].dropna(), dayfirst=True, errors="coerce"
                ).dropna()
                ano = int(_datas_parsed.dt.year.mode().iloc[0]) if len(_datas_parsed) > 0 else datetime.now().year
            except Exception:
                ano = datetime.now().year
        else:
            ano = datetime.now().year

    # ── Preparar dados válidos do REPASSE ──────────────────────────────────
    df_rep = df_correlacao.copy()
    df_rep["_cod"]  = (df_rep.get("CodigoTUSS_REPASSE", pd.Series(dtype=str))
                       .astype(str).str.replace(".0", "", regex=False).str.strip())
    df_rep["_val"]  = pd.to_numeric(
        df_rep.get("ValorLiberado_REPASSE", pd.Series(dtype=float)), errors="coerce")
    df_rep["_data"] = pd.to_datetime(
        df_rep.get("Data_REPASSE", pd.Series(dtype=str)), dayfirst=True, errors="coerce")
    df_rep["_conv"] = (df_rep.get("Convenio_PRODUCAO", pd.Series(dtype=str))
                       .astype(str).apply(_normalizar_convenio))

    # Filtrar: ano corrente, valor > 0, código numérico
    mask_ano = df_rep["_data"].dt.year == ano
    mask_val = df_rep["_val"] > 0
    mask_cod = df_rep["_cod"].str.match(r"^\d+$", na=False)
    df_ano   = df_rep[mask_ano & mask_val & mask_cod].copy()

    # Mapa de descrição real (mais frequente por código)
    desc_real_map: dict[str, str] = {}
    if "Procedimento_REPASSE" in df_correlacao.columns and not df_ano.empty:
        desc_real = (
            df_ano.groupby(["_cod", "Procedimento_REPASSE"])
            .size().reset_index(name="freq")
            .sort_values(["_cod", "freq"], ascending=[True, False])
            .groupby("_cod").first().reset_index()[["_cod", "Procedimento_REPASSE"]]
        )
        desc_real_map = dict(zip(desc_real["_cod"], desc_real["Procedimento_REPASSE"]))

    novas_linhas: list[dict] = []

    # ── 1. Linhas por (Convênio normalizado, Código) ───────────────────────
    if not df_ano.empty:
        df_sorted = df_ano.sort_values("_data")
        for (conv_norm, cod), grp in df_sorted.groupby(["_conv", "_cod"]):
            if conv_norm in ("NAN", ""):
                continue
            sub   = grp["_val"]
            ult   = grp.iloc[-1]
            novas_linhas.append({
                "Ano":          ano,
                "Convenio":     conv_norm,
                "CodigoTUSS":   cod,
                "Descricao":    _desc(cod, desc_real_map),
                "Qtd":          len(sub),
                "Media":        round(float(sub.mean()), 2),
                "UltimoValor":  round(float(ult["_val"]), 2),
                "DataUltimo":   ult["_data"].strftime("%d/%m/%Y") if pd.notna(ult["_data"]) else "",
                "Confianca":    _confianca(sub),
            })

    # ── 2. Linhas GERAL (fallback agregado por código, todos convênios) ────
    todos_codigos: set[str] = set(df_ano["_cod"].unique()) if not df_ano.empty else set()
    # incluir códigos da tabela TUSS mesmo sem valor histórico
    tabela = _carregar_tabela_tuss()
    for entry in tabela.values():
        for c in [x.strip() for x in str(entry.get("CodigosTUSS", "")).split(",") if x.strip()]:
            if c.isdigit():
                todos_codigos.add(c)

    df_sorted_geral = df_ano.sort_values("_data") if not df_ano.empty else df_ano
    for cod in sorted(todos_codigos):
        sub_geral = df_ano[df_ano["_cod"] == cod]["_val"] if not df_ano.empty else pd.Series(dtype=float)
        qtd       = len(sub_geral)
        if qtd > 0:
            ult_geral = df_sorted_geral[df_sorted_geral["_cod"] == cod].iloc[-1]
            ultimo_val  = round(float(ult_geral["_val"]), 2)
            ultimo_data = ult_geral["_data"].strftime("%d/%m/%Y") if pd.notna(ult_geral["_data"]) else ""
        else:
            ultimo_val  = None
            ultimo_data = ""
        novas_linhas.append({
            "Ano":          ano,
            "Convenio":     "GERAL",
            "CodigoTUSS":   cod,
            "Descricao":    _desc(cod, desc_real_map),
            "Qtd":          qtd,
            "Media":        round(float(sub_geral.mean()), 2) if qtd > 0 else None,
            "UltimoValor":  ultimo_val,
            "DataUltimo":   ultimo_data,
            "Confianca":    _confianca(sub_geral) if qtd > 0 else "Sem dados",
        })

    if not novas_linhas:
        logger.warning("_gerar_valores_tuss: nenhum dado de valor encontrado no DataFrame")
        return pd.DataFrame()

    df_novo = pd.DataFrame(novas_linhas)

    # ── Acumular: preservar anos/convênios anteriores, sobrescrever (Convenio, Ano) atual ──
    if caminho_saida.exists():
        df_existente = pd.read_csv(caminho_saida, dtype={"CodigoTUSS": str})
        # Retrocompatibilidade: CSV antigo sem coluna Convenio → descartar e reger
        if "Convenio" not in df_existente.columns:
            logger.info("tuss_valores.csv antigo (sem Convenio) — regerando com nova estrutura")
            df_existente = pd.DataFrame()
        else:
            df_existente = df_existente[df_existente["Ano"] != ano]
        df_final_val = pd.concat([df_existente, df_novo], ignore_index=True).sort_values(
            ["Ano", "Convenio", "CodigoTUSS"])
    else:
        df_final_val = df_novo

    df_final_val.to_csv(caminho_saida, index=False, encoding="utf-8-sig")
    n_conv = df_novo[df_novo["Convenio"] != "GERAL"]["Convenio"].nunique()
    logger.info(f"tuss_valores.csv atualizado: ano={ano}, {n_conv} convênios, {len(df_novo)} linhas")
    return df_final_val


@st.cache_data(show_spinner=False)
def _carregar_tabela_tuss() -> dict:
    """
    Carrega tuss_lookup_table.csv como dict {chave_norm → entry}.
    Prioridade: arquivo local → XLSX → constante embutida (_TABELA_TUSS_EMBUTIDA).
    A constante embutida garante funcionamento em Docker sem volume montado.
    """
    if _TUSS_LOOKUP_PATH.exists():
        df = pd.read_csv(_TUSS_LOOKUP_PATH, dtype={"CodigosTUSS": str, "codigo_base_proc_principal": str})
        logger.info(f"tabela_tuss carregada do arquivo local ({len(df)} linhas)")
        return {str(r["chave_norm"]): dict(r) for _, r in df.iterrows()}
    elif _TUSS_XLSX_PATH.exists():
        logger.info("tuss_lookup_table.csv não encontrado — gerando a partir do XLSX")
        df = _gerar_tabela_tuss()
        return {str(r["chave_norm"]): dict(r) for _, r in df.iterrows()}
    else:
        logger.info("tuss_lookup_table.csv não encontrado — usando constante embutida _TABELA_TUSS_EMBUTIDA")
        return dict(_TABELA_TUSS_EMBUTIDA)


@st.cache_data(show_spinner=False)
def _carregar_valores_tuss() -> dict:
    """
    Carrega tuss_valores.csv como dict de dois níveis (ano mais recente):
      - chave str  "CodigoTUSS"            → stats GERAL (fallback)
      - chave tuple (conv_norm, CodigoTUSS) → stats por convênio

    Retorna {} se arquivo não existir ou sem coluna Convenio (formato antigo).
    """
    if not _TUSS_VALORES_PATH.exists():
        logger.warning("tuss_valores.csv não encontrado — será gerado automaticamente")
        return {}
    df = pd.read_csv(_TUSS_VALORES_PATH, dtype={"CodigoTUSS": str})
    if df.empty or "Convenio" not in df.columns:
        return {}
    ano_mais_recente = int(df["Ano"].max())
    df_atual = df[df["Ano"] == ano_mais_recente]
    resultado: dict = {}
    for _, r in df_atual.iterrows():
        cod  = str(r["CodigoTUSS"])
        conv = str(r.get("Convenio", "GERAL"))
        stats = dict(r)
        if conv == "GERAL":
            resultado[cod] = stats            # fallback por código
        else:
            resultado[(conv, cod)] = stats    # específico por plano
    return resultado


def _construir_indice_tuss_repasse(df_rep: pd.DataFrame) -> set:
    """Retorna set de (paciente_norm, data, codigo_tuss) para lookup O(1)."""
    idx: set = set()
    for _, r in df_rep.iterrows():
        pac  = _normalizar_nome_paciente(str(r.get("Paciente", "")))
        data = str(r.get("Data", "")).strip()
        cod  = str(r.get("CodigoTUSS", "")).replace(".0", "").strip()
        if pac and data and cod and cod != "nan":
            idx.add((pac, data, cod))
    return idx


def _construir_desc_por_tuss_code(tabela_tuss: dict) -> dict:
    """
    Reverse-map {cod_tuss: descricao} a partir do tuss_lookup_table.
    Prioridade: código que é codigo_base_proc_principal (descrição canônica)
    sobre códigos adicionais. Ignora entradas 'Sem correspondência exata TUSS'.
    """
    desc_map: dict[str, str] = {}
    # Primeira passagem: apenas entradas onde o código é o código-base (mais canônico)
    for entry in tabela_tuss.values():
        desc = str(entry.get("Descricao_REPASSE", "")).strip()
        if not desc or "sem correspondência" in desc.lower():
            continue
        base = str(entry.get("codigo_base_proc_principal", "")).replace(".0", "").strip()
        if base and base not in desc_map:
            desc_map[base] = desc
    # Segunda passagem: demais códigos (adicionais) da combinação
    for entry in tabela_tuss.values():
        desc = str(entry.get("Descricao_REPASSE", "")).strip()
        if not desc or "sem correspondência" in desc.lower():
            continue
        for cod in str(entry.get("CodigosTUSS", "")).split(","):
            cod = cod.strip().replace(".0", "")
            if cod and cod != "nan" and cod not in desc_map:
                desc_map[cod] = desc
    return desc_map


def _build_upgrade_map(tabela_tuss: dict) -> dict[str, set]:
    """
    Extrai da tabela_tuss um índice {base_code → {upgrade_codes}}.

    Baseado exclusivamente nas entradas com
      TipoCobranca = 'unico_cod_tuss_inclui_proc_adicional_e_principal':
    essas entradas representam o mesmo procedimento principal + adicional
    cobrados num único código combinado.  O campo codigo_base_proc_principal
    aponta o código simples do qual CodigosTUSS é o upgrade.

    Exemplo:
      codigo_base = '40201082'  (colonoscopia simples)
      CodigosTUSS = '40202666'  (colonoscopia com biópsia)
      → _upgrade_map['40201082'] = {'40202666', '40202542', ...}

    Usado em verificar_tuss_adicionais() para classificar divergências entre
    cod_repasse e esperado como UPGRADE (OK) ou DOWNGRADE (COBRAR).
    """
    _map: dict[str, set] = {}
    for entry in tabela_tuss.values():
        if str(entry.get("TipoCobranca", "")) != "unico_cod_tuss_inclui_proc_adicional_e_principal":
            continue
        base = str(entry.get("codigo_base_proc_principal", "")).replace(".0", "").strip()
        codes = [
            c.strip() for c in str(entry.get("CodigosTUSS", "")).split(",")
            if c.strip() and c.strip() != "nan"
        ]
        if base and codes:
            _map.setdefault(base, set()).update(codes)
    return _map


def _construir_desc_por_codigo(df_rep: pd.DataFrame) -> dict:
    """
    Retorna dict {codigo_tuss: descricao_oficial} extraído das linhas reais do REPASSE.
    Usa a descrição mais frequente para cada código como fonte canônica.
    """
    if "CodigoTUSS" not in df_rep.columns or "Procedimento" not in df_rep.columns:
        return {}
    df_rep = df_rep.copy()
    df_rep["_cod"] = df_rep["CodigoTUSS"].astype(str).str.replace(".0", "", regex=False).str.strip()
    grp = (
        df_rep[df_rep["_cod"].str.match(r"^\d+$", na=False)]
        .groupby(["_cod", "Procedimento"]).size().reset_index(name="freq")
        .sort_values(["_cod", "freq"], ascending=[True, False])
        .groupby("_cod").first().reset_index()
    )
    return dict(zip(grp["_cod"], grp["Procedimento"]))


def _enriquecer_com_valores_tuss(
    df: pd.DataFrame,
    valores_tuss: dict,
    desc_lookup: dict,
) -> pd.DataFrame:
    """
    Enriquece o DataFrame de correlação com:
      - ValorEstimado_TUSS: UltimoValor (ou Media) do código esperado, por convênio
      - DescricaoTUSS: preserva o que verificar_tuss_adicionais já preencheu via
        Descricao_REPASSE; completa/melhora via desc_lookup quando disponível.

    ValorEstimado_TUSS é preenchido para status que indicam divergência ou ausência:
      COBRAR_TUSS_PROC_ADICIONAL_COBRADO_COMO_SIMPLES
      COBRAR_TUSS_CODIGO_ADICIONAL_AUSENTE_NO_REPASSE
      COBRAR_TUSS_CODIGO_PRINCIPAL_DIVERGENTE
      COBRAR_TUSS_CODIGO_PRINCIPAL_DOWNGRADE
    Também preenche NAO_FATURADO_NO_REPASSE quando CodigosTUSS_Esperados estiver preenchido.
    DescricaoTUSS é preenchida/completada para qualquer linha com CodigosTUSS_Esperados.
    """
    if df.empty:
        return df

    # Garante que as colunas existam sem apagar o que verificar_tuss_adicionais já populou
    if "ValorEstimado_TUSS" not in df.columns:
        df = df.copy()
        df["ValorEstimado_TUSS"] = ""
    if "DescricaoTUSS" not in df.columns:
        df = df.copy()
        df["DescricaoTUSS"] = ""

    _STATUS_COM_VALOR = {
        "COBRAR_TUSS_PROC_ADICIONAL_COBRADO_COMO_SIMPLES",
        "COBRAR_TUSS_CODIGO_ADICIONAL_AUSENTE_NO_REPASSE",
        "COBRAR_TUSS_CODIGO_PRINCIPAL_DIVERGENTE",
        "COBRAR_TUSS_CODIGO_PRINCIPAL_DOWNGRADE",
    }

    vals:  list = [""] * len(df)
    # Inicializa preservando DescricaoTUSS já preenchida por verificar_tuss_adicionais
    descs: list = list(df["DescricaoTUSS"].fillna("").astype(str))

    for i, (_, row) in enumerate(df.iterrows()):
        st_tuss = str(row.get("StatusTUSS", "")).strip()
        st_corr = str(row.get("StatusCorrelacao", "")).strip().upper()
        cod = str(row.get("CodigosTUSS_Esperados", "")).split(",")[0].strip().replace(".0", "")
        if not cod:
            continue

        quer_valor = (
            st_tuss in _STATUS_COM_VALOR
            or st_corr == "NAO_FATURADO_NO_REPASSE"
        )

        # Sempre tenta completar DescricaoTUSS se ainda vazia
        if not descs[i]:
            d = desc_lookup.get(cod, "")
            if not d and valores_tuss:
                d = str(valores_tuss.get(cod, {}).get("Descricao", ""))
            if d:
                descs[i] = d

        if not quer_valor or not valores_tuss:
            continue

        conv  = _normalizar_convenio(str(row.get("Convenio_PRODUCAO", "")))
        entry = (valores_tuss.get((conv, cod), {}) if conv else {}) or valores_tuss.get(cod, {})
        v = entry.get("UltimoValor") or entry.get("Media")
        if v is not None:
            vals[i] = round(float(v), 2)

        # Melhora descrição se desc_lookup tiver algo mais preciso
        desc_novo = desc_lookup.get(cod) or str(entry.get("Descricao", ""))
        if desc_novo and (not descs[i] or descs[i].startswith("Código TUSS ")):
            descs[i] = desc_novo

    df = df.copy()
    df["ValorEstimado_TUSS"] = vals
    df["DescricaoTUSS"]      = descs
    return df


def verificar_tuss_adicionais(
    linhas_resultado: list,
    df_rep: pd.DataFrame,
    tabela_tuss: dict,
    tuss_idx: set,
) -> list:
    """
    Pós-processamento: para TODA linha CORRELACIONADO (com ou sem ProcedimentosAdicionais),
    verifica os códigos TUSS esperados e preenche StatusTUSS, CodigosTUSS_Esperados,
    CodigosTUSS_Ausentes e DescricaoTUSS.

    Casos cobertos:
    - Proc simples (sem PA): OK_TUSS_PROC_PRINCIPAL_OK / OK_TUSS_CODIGO_PRINCIPAL_UPGRADE /
        COBRAR_TUSS_CODIGO_PRINCIPAL_DOWNGRADE / COBRAR_TUSS_CODIGO_PRINCIPAL_DIVERGENTE
    - PA incorporado no principal: OK_TUSS_ADICIONAL_INCORPORADO_NO_PRINCIPAL
    - PA com código único: OK_TUSS_PROC_ADICIONAL_RECONHECIDO / COBRAR_TUSS_PROC_ADICIONAL_COBRADO_COMO_SIMPLES
    - PA com múltiplos códigos: OK_TUSS_TODOS_CODIGOS_ADICIONAIS_FATURADOS / COBRAR_TUSS_CODIGO_ADICIONAL_AUSENTE_NO_REPASSE
    - Companion rows (MetodoMatch=5_FALLBACK_COMPANION sem proc_princ): OK_TUSS_PROC_ADICIONAL_RECONHECIDO
    - Proc simples com código divergente:
        OK_TUSS_CODIGO_PRINCIPAL_UPGRADE    (repasse pagou upgrade do esperado)
        COBRAR_TUSS_CODIGO_PRINCIPAL_DOWNGRADE (repasse pagou versão mais simples)
        COBRAR_TUSS_CODIGO_PRINCIPAL_DIVERGENTE (relação não determinável)
    """
    from datetime import timedelta

    # Índice de upgrades derivado da tabela_tuss (build único antes do loop)
    _upgrade_map = _build_upgrade_map(tabela_tuss)

    for linha in linhas_resultado:
        status = str(linha.get("StatusCorrelacao", ""))

        # Branch A: procedimento produção que não foi faturado no repasse
        if status == "NAO_FATURADO_NO_REPASSE":
            proc_princ_a = str(linha.get("Procedimento_PRODUCAO", "")).strip()
            proc_adic_a  = str(linha.get("ProcedimentosAdicionais_PRODUCAO", "")).strip()
            sem_pa_a     = not proc_adic_a or proc_adic_a == "nan"
            chave_a      = _normalizar_chave_tuss(
                f"{proc_princ_a}_{proc_adic_a}" if not sem_pa_a else f"{proc_princ_a}_"
            )
            entry_a = tabela_tuss.get(chave_a)
            if entry_a:
                codigos_raw_a = str(entry_a.get("CodigosTUSS", ""))
                codigos_a     = [c.strip() for c in codigos_raw_a.split(",") if c.strip() and c.strip() != "nan"]
                desc_a        = str(entry_a.get("Descricao_REPASSE", "")).strip()
                linha["StatusTUSS"]          = "COBRAR_TUSS_NAO_FATURADO_MAPEADO"
                linha["CodigosTUSS_Esperados"] = ", ".join(codigos_a)
                if desc_a:
                    linha["DescricaoTUSS"] = desc_a
            else:
                linha["StatusTUSS"] = "CORRELACIONAR_MANUAL_TUSS_COMBINACAO_SEM_MAPEAMENTO"
            continue

        # Branch B: repasse sem produção correspondente — registrar o código do repasse
        if status in ("REPASSE_NAO_IDENTIFICADO_NA_PRODUCAO", "REPASSE_DATA_FORA_DO_PERIODO_PRODUCAO"):
            cod_rep_b = str(linha.get("CodigoTUSS_REPASSE", "")).replace(".0", "").strip()
            linha["StatusTUSS"] = "CORRELACIONAR_MANUAL_TUSS_REPASSE_SEM_PRODUCAO"
            if cod_rep_b and cod_rep_b != "nan":
                linha["CodigosTUSS_Esperados"] = cod_rep_b
            continue

        # Branch C: demais status não-CORRELACIONADO — pular
        if not status.startswith("CORRELACIONADO"):
            continue

        proc_adic  = str(linha.get("ProcedimentosAdicionais_PRODUCAO", "")).strip()
        sem_pa     = not proc_adic or proc_adic == "nan"
        proc_princ = str(linha.get("Procedimento_PRODUCAO", "")).strip()
        sem_princ  = not proc_princ or proc_princ == "nan"

        # Causa 2: companion row — correlacionado via FALLBACK_COMPANION sem proc_princ nem PA
        # O código já está presente no repasse; apenas registrar como reconhecido.
        if sem_pa and sem_princ:
            if linha.get("MetodoMatch") == "5_FALLBACK_COMPANION_PROCEDIMENTO_ADICIONAL":
                linha["StatusTUSS"] = "OK_TUSS_PROC_ADICIONAL_RECONHECIDO"
            continue

        # Monta chave: com PA → "PROC_PA", sem PA → "PROC_"
        chave = _normalizar_chave_tuss(
            f"{proc_princ}_{proc_adic}" if not sem_pa else f"{proc_princ}_"
        )
        entry = tabela_tuss.get(chave)

        if not entry:
            linha["StatusTUSS"] = "CORRELACIONAR_MANUAL_TUSS_COMBINACAO_SEM_MAPEAMENTO"
            continue

        tipo       = str(entry.get("TipoCobranca", ""))
        codigos_raw = str(entry.get("CodigosTUSS", ""))
        codigos    = [c.strip() for c in codigos_raw.split(",") if c.strip() and c.strip() != "nan"]
        base       = str(entry.get("codigo_base_proc_principal", "")).replace(".0", "").strip()
        desc       = str(entry.get("Descricao_REPASSE", "")).strip()

        if tipo == "sem_mapeamento_tuss":
            linha["StatusTUSS"] = "CORRELACIONAR_MANUAL_TUSS_COMBINACAO_SEM_MAPEAMENTO"
            if desc:
                linha["DescricaoTUSS"] = desc
            continue

        # Proc simples sem PA: verificar se o código no REPASSE bate com o esperado
        if sem_pa and tipo == "unico_cod_tuss_somente_proc_principal":
            cod_repasse = str(linha.get("CodigoTUSS_REPASSE", "")).replace(".0", "").strip()
            esperado    = codigos[0] if codigos else ""
            if esperado and cod_repasse == esperado:
                # Código exato — OK
                linha["StatusTUSS"] = "OK_TUSS_PROC_PRINCIPAL_OK"
            elif cod_repasse and cod_repasse in _upgrade_map.get(esperado, set()):
                # Repasse pagou um código que é upgrade do esperado (cobre mais) — OK
                linha["StatusTUSS"] = "OK_TUSS_CODIGO_PRINCIPAL_UPGRADE"
            elif esperado and esperado in _upgrade_map.get(cod_repasse, set()):
                # Repasse pagou o código base mais simples; esperado é o upgrade — cobrar diferença
                linha["StatusTUSS"] = "COBRAR_TUSS_CODIGO_PRINCIPAL_DOWNGRADE"
            else:
                # Código divergente sem relação determinável na tabela
                linha["StatusTUSS"] = "COBRAR_TUSS_CODIGO_PRINCIPAL_DIVERGENTE"
            if esperado:
                linha["CodigosTUSS_Esperados"] = esperado
            if desc:
                linha["DescricaoTUSS"] = desc
            continue

        # PA com código igual ao do proc principal (adicional incorporado)
        if tipo == "unico_cod_tuss_somente_proc_principal":
            linha["StatusTUSS"] = "OK_TUSS_ADICIONAL_INCORPORADO_NO_PRINCIPAL"
            if codigos:
                linha["CodigosTUSS_Esperados"] = codigos[0]
            if desc:
                linha["DescricaoTUSS"] = desc
            continue

        cod_repasse = str(linha.get("CodigoTUSS_REPASSE", "")).replace(".0", "").strip()

        if tipo == "unico_cod_tuss_inclui_proc_adicional_e_principal":
            if codigos and cod_repasse == codigos[0]:
                linha["StatusTUSS"] = "OK_TUSS_PROC_ADICIONAL_RECONHECIDO"
            else:
                linha["StatusTUSS"] = "COBRAR_TUSS_PROC_ADICIONAL_COBRADO_COMO_SIMPLES"
            linha["CodigosTUSS_Esperados"] = codigos[0] if codigos else ""
            if desc:
                linha["DescricaoTUSS"] = desc
            continue

        # tipo == "multiplos_cod_tuss_proced_adicional"
        # Verificar cada código adicional (diferente do base) no índice do REPASSE
        pac  = _normalizar_nome_paciente(str(linha.get("Paciente_PRODUCAO", "")))
        data = str(linha.get("Data_PRODUCAO", "")).strip()
        try:
            data_dt = datetime.strptime(data, "%d/%m/%Y")
            datas_check = [(data_dt + timedelta(days=d)).strftime("%d/%m/%Y") for d in (0, -1, 1)]
        except ValueError:
            datas_check = [data]

        codigos_adicionais = [c for c in codigos if c != base]
        if not codigos_adicionais:
            codigos_adicionais = codigos

        encontrados, ausentes = [], []
        for cod in codigos_adicionais:
            achou = any((pac, d, cod) in tuss_idx for d in datas_check)
            (encontrados if achou else ausentes).append(cod)

        if not ausentes:
            linha["StatusTUSS"] = "OK_TUSS_TODOS_CODIGOS_ADICIONAIS_FATURADOS"
        else:
            linha["StatusTUSS"] = "COBRAR_TUSS_CODIGO_ADICIONAL_AUSENTE_NO_REPASSE"

        linha["CodigosTUSS_Esperados"] = ", ".join(codigos_adicionais)
        if ausentes:
            linha["CodigosTUSS_Ausentes"] = ", ".join(ausentes)
        if desc:
            linha["DescricaoTUSS"] = desc

    return linhas_resultado


def _refinar_status_por_valor(df: pd.DataFrame) -> pd.DataFrame:
    """
    Pós-processamento: refina StatusCorrelacao para linhas CORRELACIONADO onde
    ValorLiberado_REPASSE é menor que ValorEstimado_TUSS (tolerância 5%).

    Casos cobertos:
    - valor_rep == 0 e val_est > 0 → CORRELACIONADO (glosa total — ver ValorLiberado_REPASSE)
    - 0 < valor_rep < val_est*0.95 → CORRELACIONADO (glosa parcial — ver ValorLiberado_REPASSE)

    Requer que _enriquecer_com_valores_tuss já tenha rodado (coluna ValorEstimado_TUSS).
    """
    if df.empty:
        return df
    if "ValorEstimado_TUSS" not in df.columns or "ValorLiberado_REPASSE" not in df.columns:
        return df

    df = df.copy()
    mask_corr     = df["StatusCorrelacao"].str.startswith("CORRELACIONADO", na=False)
    mask_sem_glosa = ~df["StatusCorrelacao"].str.contains("GLOSA", na=False)

    for i in df.index[mask_corr & mask_sem_glosa]:
        val_rep_raw = df.at[i, "ValorLiberado_REPASSE"]
        val_est_raw = df.at[i, "ValorEstimado_TUSS"]
        if val_est_raw == "" or val_est_raw is None or str(val_est_raw).strip() == "":
            continue
        try:
            val_est = float(val_est_raw)
        except (ValueError, TypeError):
            continue
        if val_est <= 0:
            continue
        val_rep = _extrair_valor_numerico(val_rep_raw)
        if val_rep == 0:
            df.at[i, "StatusCorrelacao"] = "CORRELACIONADO"
        elif val_rep < val_est * 0.95:
            df.at[i, "StatusCorrelacao"] = "CORRELACIONADO"

    return df


def _verificar_erros_producao(df: pd.DataFrame) -> tuple:
    """
    Verifica erros de qualidade nos dados de PRODUCAO.

    Regras verificadas:
      • Data_PRODUCAO deve conter uma data válida no formato dd/mm/yyyy
      • Paciente_PRODUCAO não pode ser vazio
      • NrAtendimento_PRODUCAO não pode ser vazio
      • Procedimento_PRODUCAO não pode ser vazio

    Returns:
        (resumo: dict, df_erros: DataFrame com as linhas problemáticas e colunas de diagnóstico)
    """
    import re as _re

    _DATE_RE = _re.compile(r"^\d{2}/\d{2}/\d{4}$")

    def _data_invalida(v) -> bool:
        if not isinstance(v, str) or not v.strip():
            return True  # vazio também é inválido
        s = v.strip()
        if not _DATE_RE.match(s):
            return True
        try:
            datetime.strptime(s, "%d/%m/%Y")
            return False
        except ValueError:
            return True

    def _vazio(v) -> bool:
        return not isinstance(v, str) or not v.strip()

    # Suporta tanto o CSV processado (colunas sem sufixo: "Data", "Paciente"...)
    # quanto o CSV correlacionado (colunas com sufixo: "Data_PRODUCAO", ...).
    def _col(df: pd.DataFrame, *nomes: str) -> pd.Series:
        for n in nomes:
            if n in df.columns:
                return df[n]
        return pd.Series("", index=df.index)

    mask_data    = _col(df, "Data_PRODUCAO",         "Data").map(_data_invalida)
    mask_pac     = _col(df, "Paciente_PRODUCAO",      "Paciente").map(_vazio)
    mask_nratend = _col(df, "NrAtendimento_PRODUCAO", "NrAtendimento").map(_vazio)
    mask_proc    = _col(df, "Procedimento_PRODUCAO",  "Procedimento").map(_vazio)

    mask_qualquer = mask_data | mask_pac | mask_nratend | mask_proc

    df_erros = df[mask_qualquer].copy()
    df_erros["❌ Data inválida"]       = mask_data[mask_qualquer].map(lambda x: "Sim" if x else "")
    df_erros["❌ Paciente vazio"]      = mask_pac[mask_qualquer].map(lambda x: "Sim" if x else "")
    df_erros["❌ NrAtendimento vazio"] = mask_nratend[mask_qualquer].map(lambda x: "Sim" if x else "")
    df_erros["❌ Procedimento vazio"]  = mask_proc[mask_qualquer].map(lambda x: "Sim" if x else "")

    resumo = {
        "total":              int(mask_qualquer.sum()),
        "data_invalida":      int(mask_data.sum()),
        "paciente_vazio":     int(mask_pac.sum()),
        "nratendimento_vazio":int(mask_nratend.sum()),
        "procedimento_vazio": int(mask_proc.sum()),
    }

    return resumo, df_erros


def correlacionar_csv_arquivos(
    csv_producao: str,
    csv_repasse: str,
    limiar_similaridade: float = 0.65,
    tabela_tuss_preloaded: dict | None = None,
    valores_tuss_preloaded: dict | None = None,
) -> str:
    """
    Correlaciona dois CSVs (PRODUCAO e REPASSE) usando chave composta otimizada.

    Garantias de schema:
    - TODAS as colunas canonicas de _MAP_PRODUCAO aparecem com sufixo _PRODUCAO
    - TODAS as colunas canonicas de _MAP_REPASSE aparecem com sufixo _REPASSE
    - Colunas ausentes no arquivo fisico ficam vazias — nunca omitidas

    Otimizacoes implementadas:
    - Indice hash para busca O(1) por Data+Paciente
    - Fallback por NrAtendimento quando nao encontra por nome
    - Normalizacao de nomes sem acentos e caracteres especiais
    - Cache de similaridade de procedimentos
    - Busca com tolerancia de +-1 dia

    Args:
        csv_producao: CSV padronizado da PRODUCAO
        csv_repasse: CSV padronizado do REPASSE
        limiar_similaridade: Threshold para match de procedimento (0.0-1.0)

    Returns:
        CSV correlacionado com sufixos _PRODUCAO e _REPASSE em todas as colunas
    """
    try:
        # ── Carrega DataFrames ────────────────────────────────────────────────
        df_prod = pd.read_csv(io.StringIO(csv_producao), dtype=str)
        df_rep  = pd.read_csv(io.StringIO(csv_repasse),  dtype=str)
        df_prod.fillna("", inplace=True)
        df_rep.fillna("",  inplace=True)

        logger.info(f"Correlacao: {len(df_prod)} linhas PRODUCAO, {len(df_rep)} linhas REPASSE")

        # ── Deriva schemas canonicos a partir dos mapas ───────────────────────
        # Todas as colunas canonicas de cada mapa (valores unicos, ordem estavel)
        _COLS_PROD = list(dict.fromkeys(
            list(_MAP_PRODUCAO.values()) +
            list(_MAP_PRODUCAO_2025_NOVO.values()) +
            list(_MAP_PRODUCAO_2026.values())
        ))
        # Remove Procedimento2 — coluna interna descartada antes de chegar aqui
        _COLS_PROD = [c for c in _COLS_PROD if c != "Procedimento2"]
        # AbaOrigemDados e adicionada pelo processador, nao consta nos mapas — inclui explicitamente
        if "AbaOrigemDados" not in _COLS_PROD:
            _COLS_PROD.append("AbaOrigemDados")

        _COLS_REP = list(dict.fromkeys(_MAP_REPASSE.values()))
        # AbaOrigemDados e adicionada pelo processador, nao consta no _MAP_REPASSE — inclui explicitamente
        if "AbaOrigemDados" not in _COLS_REP:
            _COLS_REP.append("AbaOrigemDados")

        # Colunas internas/de controle que nao devem aparecer no CSV final
        _EXCLUIR_PROD = {"TipoArquivo"}
        _EXCLUIR_REP  = {"TipoArquivo", "_matched"}

        # Garante que todas as colunas canonicas existam nos DataFrames
        # (arquivo fisico pode nao ter todas)
        for col in _COLS_PROD:
            if col not in df_prod.columns:
                df_prod[col] = ""
        for col in _COLS_REP:
            if col not in df_rep.columns:
                df_rep[col] = ""

        # ── Normaliza datas ───────────────────────────────────────────────────
        for df in [df_prod, df_rep]:
            if "Data" in df.columns:
                df["Data"] = df["Data"].apply(_padronizar_data)

        # ── Cria indices de busca rapida ──────────────────────────────────────
        indice_repasse      = _criar_indice_repasse(df_rep)
        indice_atendimento  = _criar_indice_repasse_atendimento(df_rep)
        logger.info(f"Indice por nome: {len(indice_repasse)} chaves | por atendimento: {len(indice_atendimento)} chaves")

        df_rep["_matched"] = False
        cache_similaridade: dict = {}

        linhas_resultado       = []
        matches_encontrados    = 0
        matches_por_atendimento = 0

        # ── Itera PRODUCAO e busca match no REPASSE ───────────────────────────
        for _, row_prod in df_prod.iterrows():
            data_prod     = row_prod.get("Data", "")
            paciente_norm = _normalizar_nome_paciente(row_prod.get("Paciente", ""))
            nr_atend_prod = str(row_prod.get("NrAtendimento", "")).strip()
            proc_prod     = row_prod.get("Procedimento", "")

            candidatos   = []
            metodo_busca = "1_NOME_COMPLETO_DATA_PROCEDIMENTO"

            # Busca exata por nome
            key_exata = (data_prod, paciente_norm)
            if key_exata in indice_repasse:
                candidatos.extend(indice_repasse[key_exata])

            # Busca com tolerancia de +-1 dia por nome
            if not candidatos:
                try:
                    data_dt = datetime.strptime(data_prod, "%d/%m/%Y")
                    for delta in [-1, 1]:
                        key_tol = ((data_dt + timedelta(days=delta)).strftime("%d/%m/%Y"), paciente_norm)
                        if key_tol in indice_repasse:
                            candidatos.extend(indice_repasse[key_tol])
                except Exception:
                    pass

            # Fallback: busca por NrAtendimento
            if not candidatos and nr_atend_prod and nr_atend_prod not in ("", "nan", "NaN"):
                metodo_busca = "2_FALLBACK_NR-ATENDIMENTO_DATA_PROCEDIMENTO"
                key_atend = (data_prod, nr_atend_prod)
                if key_atend in indice_atendimento:
                    candidatos.extend(indice_atendimento[key_atend])
                if not candidatos:
                    try:
                        data_dt = datetime.strptime(data_prod, "%d/%m/%Y")
                        for delta in [-1, 1]:
                            key_tol = ((data_dt + timedelta(days=delta)).strftime("%d/%m/%Y"), nr_atend_prod)
                            if key_tol in indice_atendimento:
                                candidatos.extend(indice_atendimento[key_tol])
                    except Exception:
                        pass

            # Seleciona melhor match por similaridade de procedimento
            melhor_match = None
            melhor_score = 0.0
            melhor_idx   = None

            for idx_rep in candidatos:
                if df_rep.at[idx_rep, "_matched"]:
                    continue
                sim = _similaridade_procedimento(proc_prod, df_rep.iloc[idx_rep].get("Procedimento", ""), cache_similaridade)
                if sim >= limiar_similaridade and sim > melhor_score:
                    melhor_score = sim
                    melhor_match = df_rep.iloc[idx_rep]
                    melhor_idx   = idx_rep

            # ── Monta linha correlacionada ────────────────────────────────────
            proc_norm = _normalizar_procedimento(proc_prod)
            linha_corr: dict = {
                "ChaveCorrelacao": f"{paciente_norm}_{nr_atend_prod}_{data_prod}_{proc_norm}".replace(" ", "-")
            }

            # Todas as colunas canonicas da PRODUCAO com sufixo _PRODUCAO
            for col in _COLS_PROD:
                if col not in _EXCLUIR_PROD:
                    linha_corr[f"{col}_PRODUCAO"] = row_prod.get(col, "")

            # ── Fallbacks quando chave principal não encontrou match ──────────
            metodo_match = metodo_busca  # 1_NOME_COMPLETO ou 2_FALLBACK_NR-ATENDIMENTO

            if melhor_match is None:
                # Fallback 1: combinações de tokens do nome + data ±1 dia
                idx_fb1, score_fb1 = _buscar_fallback1_combinacoes_nome(
                    data_prod, row_prod.get("Paciente", ""),
                    df_rep, indice_repasse,
                    proc_prod, cache_similaridade,
                    limiar_similaridade,
                )
                if idx_fb1 is not None:
                    melhor_match = df_rep.iloc[idx_fb1]
                    melhor_idx   = idx_fb1
                    melhor_score = score_fb1
                    metodo_match = "3_FALLBACK_NOME_PARCIAL_FUZZY_DATA_FIXA"

            if melhor_match is None:
                # Fallback 2: nome exato + procedimento + data ±7 dias
                idx_fb2, score_fb2 = _buscar_fallback2_paciente_proc_data_ampla(
                    data_prod, paciente_norm, proc_prod,
                    df_rep, indice_repasse, cache_similaridade,
                    limiar_similaridade,
                )
                if idx_fb2 is not None:
                    melhor_match = df_rep.iloc[idx_fb2]
                    melhor_idx   = idx_fb2
                    melhor_score = score_fb2
                    metodo_match = "4_FALLBACK_NOME_COMPLETO_DATA-FLEXIVEL"

            # ── StatusCorrelacao e SimilaridadeProcedimento ───────────────────
            if melhor_match is not None:
                df_rep.at[melhor_idx, "_matched"] = True
                matches_encontrados += 1
                if metodo_busca == "2_FALLBACK_NR-ATENDIMENTO_DATA_PROCEDIMENTO":
                    matches_por_atendimento += 1

                valor_rep   = _extrair_valor_numerico(melhor_match.get("ValorLiberado", "0"))
                status_base = _determinar_status_correlacao(valor_rep, True)

                status = status_base

                # Ajuste A: sinaliza procedimentos anatomicamente divergentes para revisão humana
                # (ex: ENDOSCOPIA da PRODUCAO casou com "Colonoscopia" do REPASSE via
                # similaridade acidental de string — precisam ser conferidos manualmente)
                if _sao_anatomicamente_divergentes(proc_prod, melhor_match.get("Procedimento", "")):
                    metodo_match = f"{metodo_match}_PROCEDIMENTO_DIVERGENTE"

                linha_corr["SimilaridadeProcedimento"] = f"{melhor_score:.2f}"
                linha_corr["MetodoMatch"]              = metodo_match
            else:
                status = "NAO_FATURADO_NO_REPASSE"
                linha_corr["SimilaridadeProcedimento"] = "0.00"
                linha_corr["MetodoMatch"]              = "SEM_MATCH"

            linha_corr["StatusCorrelacao"] = status

            # Todas as colunas canonicas do REPASSE com sufixo _REPASSE
            if melhor_match is not None:
                for col in _COLS_REP:
                    if col not in _EXCLUIR_REP:
                        linha_corr[f"{col}_REPASSE"] = melhor_match.get(col, "")
            else:
                for col in _COLS_REP:
                    if col not in _EXCLUIR_REP:
                        linha_corr[f"{col}_REPASSE"] = ""

            linhas_resultado.append(linha_corr)

        logger.info(f"Matches encontrados: {matches_encontrados}/{len(df_prod)} ({matches_encontrados/len(df_prod)*100:.1f}%)")
        logger.info(f"Matches por atendimento: {matches_por_atendimento}")

        # ── Pré-computação para Fallbacks 5 e 6 ──────────────────────────────

        # Fallback 5: índice (paciente_norm, data) das linhas PRODUCAO já correlacionadas.
        # Permite detectar se um procedimento companion do REPASSE tem um principal
        # já matcheado no mesmo episódio (mesmo paciente + data ±1 dia).
        _corr_idx: Dict[Tuple[str, str], bool] = {}
        for lr in linhas_resultado:
            if str(lr.get("StatusCorrelacao", "")).startswith("CORRELACIONADO"):
                pac_lr  = _normalizar_nome_paciente(lr.get("Paciente_PRODUCAO", ""))
                data_lr = lr.get("Data_PRODUCAO", "")
                if pac_lr and data_lr:
                    _corr_idx[(pac_lr, data_lr)] = True

        # Fallback 6: data mínima da PRODUCAO com buffer de 30 dias.
        # Entradas do REPASSE anteriores a essa data são faturamentos tardios (late billing)
        # sem correspondência possível na PRODUCAO — recebem status informativo próprio.
        try:
            _datas_prod_parsed = [
                datetime.strptime(d.strip(), "%d/%m/%Y")
                for d in df_prod["Data"].dropna()
                if re.match(r"\d{2}/\d{2}/\d{4}", str(d).strip())
            ]
            _data_min_producao = min(_datas_prod_parsed) - timedelta(days=30) if _datas_prod_parsed else None
        except Exception:
            _data_min_producao = None

        # ── Linhas do REPASSE sem match → inseridas no final ──────────────────
        nao_matcheados = 0
        for idx_rep, row_rep in df_rep.iterrows():
            if row_rep["_matched"]:
                continue
            nao_matcheados += 1

            paciente_norm  = _normalizar_nome_paciente(row_rep.get("Paciente", ""))
            nr_atend_rep   = str(row_rep.get("NrAtendimento", "")).strip()
            data_rep       = row_rep.get("Data", "")
            proc_norm      = _normalizar_procedimento(row_rep.get("Procedimento", ""))

            linha_corr = {
                "ChaveCorrelacao": f"{paciente_norm}_{nr_atend_rep}_{data_rep}_{proc_norm}".replace(" ", "-")
            }

            # PRODUCAO vazia (schema completo, tudo vazio)
            for col in _COLS_PROD:
                if col not in _EXCLUIR_PROD:
                    linha_corr[f"{col}_PRODUCAO"] = ""

            # Determina status do REPASSE não matcheado (com fallbacks 5 e 6)
            status_repasse = "REPASSE_NAO_IDENTIFICADO_NA_PRODUCAO"
            linha_corr["MetodoMatch"] = ""

            # Fallback 5: procedimento companion (Urease/Helicobacter standalone) cujo
            # procedimento principal já foi correlacionado no mesmo episódio.
            # Condições: (a) proc é companion puro, (b) mesmo paciente tem CORRELACIONADO
            # na mesma data ±1 dia. Risco de falso positivo: muito baixo — requer
            # keyword específico + ausência de proc principal + match de episódio.
            if _e_procedimento_companion(row_rep.get("Procedimento", "")):
                try:
                    _data_rep_dt = datetime.strptime(data_rep, "%d/%m/%Y")
                    _datas_f5 = [data_rep] + [
                        (_data_rep_dt + timedelta(days=d)).strftime("%d/%m/%Y")
                        for d in [-1, 1]
                    ]
                except ValueError:
                    _datas_f5 = [data_rep]
                for _dc in _datas_f5:
                    if (paciente_norm, _dc) in _corr_idx:
                        status_repasse = "CORRELACIONADO"
                        linha_corr["MetodoMatch"] = "5_FALLBACK_COMPANION_PROCEDIMENTO_ADICIONAL"
                        break

            # Fallback 6: data do REPASSE anterior ao período coberto pela PRODUCAO
            # (faturamento tardio de exames de períodos anteriores ao arquivo enviado).
            if status_repasse == "REPASSE_NAO_IDENTIFICADO_NA_PRODUCAO" and _data_min_producao:
                try:
                    if datetime.strptime(data_rep, "%d/%m/%Y") < _data_min_producao:
                        status_repasse = "REPASSE_DATA_FORA_DO_PERIODO_PRODUCAO"
                except ValueError:
                    pass

            linha_corr["SimilaridadeProcedimento"] = "0.00"
            linha_corr["StatusCorrelacao"]         = status_repasse

            # REPASSE preenchido (schema completo)
            for col in _COLS_REP:
                if col not in _EXCLUIR_REP:
                    linha_corr[f"{col}_REPASSE"] = row_rep.get(col, "")

            linhas_resultado.append(linha_corr)

        logger.info(f"Linhas REPASSE nao matcheadas: {nao_matcheados}")
        logger.info(f"Cache de similaridade: {len(cache_similaridade)} entradas")

        # ── Verificação TUSS pós-correlação ───────────────────────────────────
        tabela_tuss: dict = {}
        _tuss_debug_log = _TUSS_DIR / "tuss_debug.log"
        try:
            def _dbg(msg):
                """Grava diagnóstico em arquivo para inspeção mesmo sem acesso ao terminal."""
                import threading
                linha = f"[{datetime.now().isoformat()} T={threading.current_thread().name}] {msg}\n"
                logger.info(msg)
                try:
                    with open(_tuss_debug_log, "a", encoding="utf-8") as _f:
                        _f.write(linha)
                except Exception:
                    pass

            _dbg(f"TUSS-INIT preloaded={bool(tabela_tuss_preloaded)} len={len(tabela_tuss_preloaded) if tabela_tuss_preloaded else 0}")
            _dbg(f"TUSS-PATH exists={_TUSS_LOOKUP_PATH.exists()} path={_TUSS_LOOKUP_PATH}")

            if tabela_tuss_preloaded:
                tabela_tuss = tabela_tuss_preloaded
                _dbg(f"TUSS-LOAD via preloaded ({len(tabela_tuss)} entradas)")
            elif _TUSS_LOOKUP_PATH.exists():
                _df_lookup = pd.read_csv(
                    _TUSS_LOOKUP_PATH,
                    dtype={"CodigosTUSS": str, "codigo_base_proc_principal": str},
                )
                tabela_tuss = {str(r["chave_norm"]): dict(r) for _, r in _df_lookup.iterrows()}
                _dbg(f"TUSS-LOAD via read_csv ({len(tabela_tuss)} entradas)")
            else:
                tabela_tuss = dict(_TABELA_TUSS_EMBUTIDA)
                _dbg(f"TUSS-LOAD via embutida ({len(tabela_tuss)} entradas)")

            if tabela_tuss:
                tuss_idx = _construir_indice_tuss_repasse(df_rep)
                _dbg(f"TUSS-IDX construído ({len(tuss_idx)} entradas)")
                linhas_antes = len(linhas_resultado)
                linhas_resultado = verificar_tuss_adicionais(linhas_resultado, df_rep, tabela_tuss, tuss_idx)
                n_com_status = sum(1 for l in linhas_resultado if "StatusTUSS" in l)
                _dbg(f"TUSS-VERIFY concluído: {linhas_antes} linhas, {n_com_status} com StatusTUSS")
            else:
                _dbg("TUSS-SKIP tabela_tuss vazia")
        except Exception as _e_tuss:
            logger.warning(f"Verificação TUSS ignorada: {_e_tuss}", exc_info=True)
            try:
                with open(_tuss_debug_log, "a", encoding="utf-8") as _f:
                    import traceback
                    _f.write(f"TUSS-EXCEPTION: {_e_tuss}\n{traceback.format_exc()}\n")
            except Exception:
                pass

        # ── Monta DataFrame final e ordena por data ───────────────────────────
        df_final = pd.DataFrame(linhas_resultado)
        df_final.fillna("", inplace=True)

        # ── Gera/atualiza tuss_valores.csv ANTES do enriquecimento ────────────
        try:
            if not valores_tuss_preloaded or not valores_tuss_preloaded:
                _gerar_valores_tuss(df_final)
                _carregar_valores_tuss.clear()
                logger.info("tuss_valores.csv gerado/atualizado durante correlação")
        except Exception as _e_gen:
            logger.warning(f"Geração tuss_valores ignorada: {_e_gen}", exc_info=True)

        # ── Enriquecimento com valores e descrições TUSS ──────────────────────
        try:
            # Recarrega valores_tuss após geração
            _valores_tuss_corr = (
                valores_tuss_preloaded if valores_tuss_preloaded
                else _carregar_valores_tuss()
            )
            _desc_lookup_corr  = _construir_desc_por_tuss_code(tabela_tuss)
            df_final = _enriquecer_com_valores_tuss(df_final, _valores_tuss_corr, _desc_lookup_corr)
            logger.info("Enriquecimento TUSS concluído")
            df_final = _refinar_status_por_valor(df_final)
            logger.info("Refinamento de status por valor concluído")
        except Exception as _e_enr:
            logger.warning(f"Enriquecimento TUSS ignorado: {_e_enr}", exc_info=True)

        if "Data_PRODUCAO" in df_final.columns:
            df_final["_sort_date"] = df_final.apply(
                lambda r: r["Data_PRODUCAO"] if r["Data_PRODUCAO"] else r.get("Data_REPASSE", ""),
                axis=1,
            )
            df_final.sort_values("_sort_date", inplace=True)
            df_final.drop(columns=["_sort_date"], inplace=True)

        return df_final.to_csv(index=False, encoding="utf-8")

    except Exception as e:
        logger.error(f"Erro na correlacao local: {e}", exc_info=True)
        return ""


# =============================================================================
# CONFIGURAÇÃO DO LLM
# =============================================================================

@st.cache_resource
def get_llm(provider: str, model_name: str, temperature: float, api_key: str, base_url: str = None) -> LLM:
    if not api_key:
        st.error("⚠️ API Key não configurada!")
        st.stop()

    model_name = model_name.strip()
    
    # Mapeamento de prefixos por provider
    provider_prefixes = {
        "Google Gemini": "google/",
        "OpenAI": "openai/",
        "Anthropic": "anthropic/",
        "Groq": "groq/",
        "Ollama": "ollama/",
        "Azure OpenAI": "azure/",
        "AWS Bedrock": "bedrock/",
        "Cohere": "cohere/",
        "HuggingFace": "huggingface/",
        "OpenRouter": "openrouter/",
        "Mistral AI": "mistral/",
        "Grok (X.AI)": "xai/",
    }
    
    prefix = provider_prefixes.get(provider, "")
    
    # Adiciona prefixo se não existir
    if prefix and not model_name.startswith(prefix):
        model_name = f"{prefix}{model_name}"
    
    try:
        llm_params = {
            "model": model_name,
            "api_key": api_key,
            "temperature": temperature
        }
        
        # Adiciona base_url se fornecido (para Ollama, Azure, etc)
        if base_url and base_url.strip():
            llm_params["base_url"] = base_url.strip()
        
        return LLM(**llm_params)
    except Exception as e:
        st.error(f"Erro ao inicializar o modelo '{model_name}': {e}")
        st.info(f"Verifique se o nome do modelo e a API Key estão corretos para {provider}")
        st.stop()


# =============================================================================
# CRIAÇÃO DOS AGENTES  ← usam session_state quando disponível
# =============================================================================

def create_agents(llm, verbose_mode: bool = False) -> list:
    cfg = _get_analista_cfg()
    multiendoscopia_analista = Agent(
        role=cfg["role"],
        goal=cfg["goal"],
        backstory=cfg["backstory"],
        llm=llm,
        verbose=verbose_mode,
        allow_delegation=False,
    )
    return [multiendoscopia_analista]


def create_correlator_agent(llm, verbose_mode: bool = False) -> Agent:
    cfg = _get_correlacionador_cfg()
    return Agent(
        role=cfg["role"],
        goal=cfg["goal"],
        backstory=cfg["backstory"],
        llm=llm,
        verbose=verbose_mode,
        allow_delegation=False,
    )


def create_correlation_task(correlator_agent: Agent, csvs_por_arquivo: dict) -> Task:
    cfg = _get_correlacionador_cfg()
    blocos = "\n\n".join(
        f"=== ARQUIVO: {fname} ===\n{csv_text}"
        for fname, csv_text in csvs_por_arquivo.items()
    )
    description = cfg["task_description_template"].replace("{blocos}", blocos)
    return Task(
        description=description,
        expected_output=cfg["task_expected_output"],
        agent=correlator_agent,
    )


# =============================================================================
# UTILITÁRIOS DE CSV
# =============================================================================

COLUNAS_ESPERADAS = [
    "Data", "NrAtendProducao", "NrAtendRepasse", "Paciente", "Convenio",
    "Procedimento", "CodigoTUSS", "MedicoExecutor",
    "QtProcedimento", "ValorLiberado", "StatusCorrelacao", "Observacao",
]

LABEL_STATUS_CORR: dict[str, str] = {
    "CORRELACIONADO":                        "Pago e Conferido",
    "NAO_FATURADO_NO_REPASSE":               "Procedimento não Repassado pelo Hospital",
    "REPASSE_NAO_IDENTIFICADO_NA_PRODUCAO":  "Pago pelo Hospital sem Registro na Produção",
    "REPASSE_DATA_FORA_DO_PERIODO_PRODUCAO": "Cobrança Fora do Período Analisado",
}

LABEL_STATUS_TUSS: dict[str, str] = {
    "OK_TUSS_PROC_PRINCIPAL_OK":                           "Código TUSS Correto",
    "OK_TUSS_PROC_ADICIONAL_RECONHECIDO":                  "Adicional Faturado com Código Correto",
    "OK_TUSS_TODOS_CODIGOS_ADICIONAIS_FATURADOS":          "Todos os Adicionais com Código Correto",
    "OK_TUSS_ADICIONAL_INCORPORADO_NO_PRINCIPAL":          "Adicional Incorporado ao Código Principal",
    "COBRAR_TUSS_PROC_ADICIONAL_COBRADO_COMO_SIMPLES":     "Subcobrança: Adicional Faturado como Simples",
    "OK_TUSS_CODIGO_PRINCIPAL_UPGRADE":                    "Repasse com Código Superior ao Esperado (sem cobrança)",
    "COBRAR_TUSS_CODIGO_PRINCIPAL_DOWNGRADE":              "Repasse com Código Mais Simples – Cobrar Diferença",
    "COBRAR_TUSS_CODIGO_PRINCIPAL_DIVERGENTE":             "Código do Proc. Principal Diverge do Esperado",
    "COBRAR_TUSS_CODIGO_ADICIONAL_AUSENTE_NO_REPASSE":     "Código Adicional Não Encontrado no Repasse",
    "COBRAR_TUSS_NAO_FATURADO_MAPEADO":                   "Não Cobrado – Código TUSS Identificado (Recuperável)",
    "CORRELACIONAR_MANUAL_TUSS_REPASSE_SEM_PRODUCAO":      "Cobrança no Repasse sem Registro de Produção",
    "CORRELACIONAR_MANUAL_TUSS_COMBINACAO_SEM_MAPEAMENTO": "Combinação de Procedimentos sem Mapeamento TUSS",
}

# Mapeamento inverso: label amigável → valor bruto (usado nos filtros multiselect)
_CORR_LABEL_TO_RAW: dict[str, str] = {v: k for k, v in LABEL_STATUS_CORR.items()}
_TUSS_LABEL_TO_RAW: dict[str, str] = {v: k for k, v in LABEL_STATUS_TUSS.items()}


def extrair_csv_do_texto(texto: str) -> str:
    match = re.search(r"```(?:csv)?\s*\n(.*?)```", texto, re.DOTALL | re.IGNORECASE)
    if match:
        return match.group(1).strip()
    linhas_csv = [
        linha for linha in texto.splitlines()
        if linha.count(",") >= 4 or linha.startswith("Data,")
    ]
    return "\n".join(linhas_csv).strip()


def separar_resumo_do_csv(texto_csv: str) -> tuple[str, str]:
    marcador = "# RESUMO_DIVERGENCIAS_POR_CONVENIO"
    if marcador in texto_csv:
        partes = texto_csv.split(marcador, 1)
        return partes[0].strip(), partes[1].strip()
    return texto_csv.strip(), ""


def texto_para_dataframe(csv_texto: str) -> pd.DataFrame | None:
    try:
        csv_limpo = extrair_csv_do_texto(csv_texto)
        csv_dados, _ = separar_resumo_do_csv(csv_limpo)
        df = pd.read_csv(io.StringIO(csv_dados), sep=",", dtype=str)
        df.columns = [c.strip() for c in df.columns]
        return df
    except Exception as e:
        logger.warning(f"Falha ao converter CSV para DataFrame: {e}")
        return None


# =============================================================================
# CRIAÇÃO DAS TASKS  ← usa session_state quando disponível
# =============================================================================

def create_tasks(agents: list, conteudo_arquivo: str) -> list:
    cfg = _get_analista_cfg()
    description = cfg["task_description_template"].replace("{conteudo_arquivo}", conteudo_arquivo)
    task_analise = Task(
        description=description,
        expected_output=cfg["task_expected_output"],
        agent=agents[0],
    )
    return [task_analise]


# =============================================================================
# INTERFACE STREAMLIT
# =============================================================================

def render_agent_form(agent_key: str, defaults: dict, title: str, icon: str):
    """
    Renderiza o formulário de edição de um agente.
    Salva os valores em st.session_state[agent_key].
    """
    cfg = st.session_state.get(agent_key, dict(defaults))

    st.markdown(f"### {icon} {title}")

    with st.container(border=True):
        col_info, col_reset = st.columns([5, 1])
        with col_info:
            st.caption("Edite os campos abaixo e clique em **Salvar Alterações** para aplicar.")
        with col_reset:
            if st.button("↩️ Restaurar", key=f"reset_{agent_key}", help="Volta aos valores padrão originais"):
                st.session_state[agent_key] = dict(defaults)
                st.rerun()

        new_role = st.text_input(
            "🏷️ Role (papel do agente)",
            value=cfg.get("role", ""),
            key=f"{agent_key}_role",
            help="Define o papel/função do agente dentro do crew.",
        )

        new_goal = st.text_area(
            "🎯 Goal (objetivo)",
            value=cfg.get("goal", ""),
            height=320,
            key=f"{agent_key}_goal",
            help="Descreve o objetivo principal do agente — instruções detalhadas de comportamento.",
        )

        new_backstory = st.text_area(
            "📖 Backstory (contexto/experiência)",
            value=cfg.get("backstory", ""),
            height=180,
            key=f"{agent_key}_backstory",
            help="Contexto e experiência do agente — usado para orientar o tom e o raciocínio.",
        )

        st.divider()
        st.markdown("#### 📋 Task associada")

        new_task_desc = st.text_area(
            "📝 Task Description (template)",
            value=cfg.get("task_description_template", ""),
            height=280,
            key=f"{agent_key}_task_desc",
            help=(
                "Template da instrução enviada ao agente em cada execução. "
                "Use {conteudo_arquivo} (Analista) ou {blocos} (Correlacionador) "
                "como placeholder do conteúdo dinâmico."
            ),
        )

        new_task_output = st.text_area(
            "✅ Task Expected Output (saída esperada)",
            value=cfg.get("task_expected_output", ""),
            height=100,
            key=f"{agent_key}_task_output",
            help="Descreve o formato/conteúdo esperado na saída da task.",
        )

        col_save, col_status = st.columns([2, 5])
        with col_save:
            if st.button(f"💾 Salvar Alterações — {title}", key=f"save_{agent_key}", type="primary"):
                st.session_state[agent_key] = {
                    "role": new_role,
                    "goal": new_goal,
                    "backstory": new_backstory,
                    "task_description_template": new_task_desc,
                    "task_expected_output": new_task_output,
                }
                with col_status:
                    st.success(f"✅ Configuração do **{title}** salva com sucesso!")


def main():
    st.set_page_config(
        page_title="Endoscopia | Controle de Procedimentos",
        page_icon="🔬",
        layout="wide",
    )

    # ── Auth guard ────────────────────────────────────────────────────────────
    if "auth_user" not in st.session_state:
        _show_login()
        st.stop()

    # Inicializa os defaults dos agentes no session_state (executado uma vez por sessão)
    _init_agent_session_state()

    st.markdown(
        '<h1 style="text-align: center;">🔬 Sistema de Controle de Procedimentos de Endoscopia</h1>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<p style="text-align: center;">'
        'Confronto automático entre Produtividade da Equipe e Repasse do Hospital<br>'
        'Identificação de Glosas, Divergências de Valor e Procedimentos Não Faturados<br>'
        'Apoio à cobrança estruturada junto ao hospital e convênios médicos'
        '</p>',
        unsafe_allow_html=True,
    )

    # ── Sidebar ──────────────────────────────────────────────────────────────
    with st.sidebar:
        # ── Usuário logado + logout ───────────────────────────────────────
        _nome  = st.session_state.get("auth_nome", "")
        _email = st.session_state.get("auth_email", "")
        _role  = st.session_state.get("auth_role", "")
        _rlabel = _ROLE_LABELS.get(_role, _role)
        _rcolor = _ROLE_COLORS.get(_role, "#374151")
        st.markdown(
            f"<div style='display:flex;align-items:center;gap:8px;padding:6px 0;'>"
            f"<div style='width:32px;height:32px;border-radius:50%;background:#2563eb;"
            f"display:flex;align-items:center;justify-content:center;"
            f"color:white;font-weight:700;font-size:13px;flex-shrink:0;'>"
            f"{(_nome or _email or '?')[0].upper()}</div>"
            f"<div style='min-width:0;'>"
            f"<div style='font-size:12px;font-weight:600;color:#111827;white-space:nowrap;"
            f"overflow:hidden;text-overflow:ellipsis;'>{_nome or _email}</div>"
            f"<span style='font-size:10px;font-weight:600;padding:1px 6px;border-radius:4px;"
            f"background:{_rcolor}22;color:{_rcolor};border:1px solid {_rcolor}44;'>"
            f"{_rlabel}</span>"
            f"</div></div>",
            unsafe_allow_html=True,
        )
        if st.button("↩️ Sair", use_container_width=True):
            _logout()
        st.divider()

        st.info("💡 Sistema exclusivo para controle de procedimentos endoscópicos e faturamento junto a convênios.")
        st.divider()
        st.header("⚙️ Configurações")

        # Seleção do Provider
        provider = st.selectbox(
            "🤖 Provider LLM",
            options=[
                "Google Gemini",
                "OpenAI",
                "Anthropic",
                "Groq",
                "OpenRouter",
                "HuggingFace",
                "Mistral AI",
                "Grok (X.AI)",
                "Ollama",
                "Azure OpenAI",
                "AWS Bedrock",
                "Cohere",
            ],
            index=0,
            help="Selecione o provedor do modelo de linguagem"
        )

        # API Key
        env_key_map = {
            "Google Gemini": "GOOGLE_API_KEY",
            "OpenAI": "OPENAI_API_KEY",
            "Anthropic": "ANTHROPIC_API_KEY",
            "Groq": "GROQ_API_KEY",
            "OpenRouter": "OPENROUTER_API_KEY",
            "HuggingFace": "HUGGINGFACE_API_KEY",
            "Mistral AI": "MISTRAL_API_KEY",
            "Grok (X.AI)": "XAI_API_KEY",
            "Azure OpenAI": "AZURE_API_KEY",
            "AWS Bedrock": "AWS_ACCESS_KEY_ID",
            "Cohere": "COHERE_API_KEY",
            "Ollama": "",
        }

        env_key = env_key_map.get(provider, "")
        env_api_key = os.getenv(env_key, "") if env_key else ""
        
        api_key = st.text_input(
            f"🔑 API Key ({provider})",
            value=env_api_key,
            type="password",
            help=f"Sua chave de API para {provider}" + (f" (variável: {env_key})" if env_key else ""),
        )

        if api_key or provider == "Ollama":
            st.success(f"✅ {provider} configurado")
        else:
            st.warning(f"⚠️ API Key não preenchida — modo Local disponível")

        st.divider()

        # Modelos principais por provider (baseado em docs.crewai.com)
        provider_models = {
            "Google Gemini": [
                "gemini-2.5-flash-lite",
                "gemini-2.5-flash",
                "gemini-2.5-pro",
                "gemini-1.5-pro",
                "gemini-1.5-flash",
            ],
            "OpenAI": [
                "gpt-4o",
                "gpt-4o-mini",
                "gpt-4-turbo",
                "gpt-4",
                "gpt-3.5-turbo",
                "o1-preview",
                "o1-mini",
            ],
            "Anthropic": [
                "claude-3-5-sonnet-20241022",
                "claude-3-5-haiku-20241022",
                "claude-3-opus-20240229",
                "claude-3-sonnet-20240229",
                "claude-3-haiku-20240307",
            ],
            "Groq": [
                "llama-3.3-70b-versatile",
                "llama-3.1-70b-versatile",
                "llama-3.1-8b-instant",
                "mixtral-8x7b-32768",
                "gemma2-9b-it",
            ],
            "OpenRouter": [
                "anthropic/claude-3.5-sonnet",
                "openai/gpt-4o",
                "google/gemini-2.0-flash-exp",
                "meta-llama/llama-3.3-70b-instruct",
                "deepseek/deepseek-chat",
                "qwen/qwen-2.5-72b-instruct",
                "x-ai/grok-2-1212",
            ],
            "HuggingFace": [
                "meta-llama/Meta-Llama-3.1-70B-Instruct",
                "meta-llama/Meta-Llama-3.1-8B-Instruct",
                "mistralai/Mixtral-8x7B-Instruct-v0.1",
                "microsoft/Phi-3-medium-4k-instruct",
                "Qwen/Qwen2.5-72B-Instruct",
            ],
            "Mistral AI": [
                "mistral-large-latest",
                "mistral-medium-latest",
                "mistral-small-latest",
                "open-mistral-7b",
                "open-mixtral-8x7b",
            ],
            "Grok (X.AI)": [
                "grok-2-1212",
                "grok-2-vision-1212",
                "grok-beta",
            ],
            "Ollama": [
                "llama3.2",
                "llama3.1",
                "mistral",
                "qwen2.5",
                "phi3",
                "gemma2",
            ],
            "Azure OpenAI": [
                "gpt-4o",
                "gpt-4-turbo",
                "gpt-4",
                "gpt-35-turbo",
            ],
            "AWS Bedrock": [
                "anthropic.claude-3-5-sonnet-20241022-v2:0",
                "anthropic.claude-3-sonnet-20240229-v1:0",
                "anthropic.claude-3-haiku-20240307-v1:0",
                "meta.llama3-70b-instruct-v1:0",
                "meta.llama3-8b-instruct-v1:0",
            ],
            "Cohere": [
                "command-r-plus",
                "command-r",
                "command",
                "command-light",
            ],
        }

        # Nome do modelo - selectbox com opções do provider
        available_models = provider_models.get(provider, [])
        
        if available_models:
            custom_model = st.selectbox(
                "📝 Modelo",
                options=available_models,
                index=0,
                help=f"Selecione o modelo {provider}"
            )
        else:
            custom_model = st.text_input(
                "📝 Nome do Modelo",
                value="",
                help=f"Digite o nome do modelo {provider}"
            )

        # Base URL (para Ollama, Azure, etc)
        show_base_url = provider in ["Ollama", "Azure OpenAI"]
        base_url = ""
        if show_base_url:
            default_url = "http://localhost:11434" if provider == "Ollama" else ""
            base_url = st.text_input(
                "🌐 Base URL",
                value=default_url,
                help=f"URL base da API {provider}"
            )

        temperature = st.slider(
            "🌡️ Temperatura",
            min_value=0.0,
            max_value=1.0,
            value=0.1,
            step=0.1,
            help="0 = determinístico | 1 = criativo",
        )

        verbose_mode = st.toggle(
            "🔊 Modo Verbose",
            value=False,
            help="Ativado: exibe o raciocínio detalhado dos agentes (mais tokens). "
                 "Desativado: apenas o resultado final (mais econômico).",
        )

        st.divider()
        st.header("👥 Agentes Disponíveis")

        agents_info = [
            ("🔍", "Analista de Endoscopia", "Lê e estrutura os arquivos PRODUCAO e REPASSE"),
            ("🔀", "Correlacionador", "Confronta registros e identifica divergências"),
        ]
        for icon, name, desc in agents_info:
            st.markdown(f"{icon} **{name}**")
            st.caption(desc)

        st.divider()
        st.header("📌 Tipos de Arquivo")
        st.markdown("""
- **PRODUCAO**: planilha da equipe de enfermagem com os procedimentos realizados (código TUSS)
- **REPASSE**: planilha emitida pelo hospital com valores faturados e pagos pelos convênios
        """)

    # ── Tabs ──────────────────────────────────────────────────────────────────
    tab1, tab2, tab3, tab4, tab_cobranca, tab_agents = st.tabs([
        "📄 Input",
        "🚀 Execução",
        "📊 Resultados",
        "🔀 Correlação",
        "📋 Gerar Cobrança",
        "🤖 Agentes",
    ])

    # ── TAB 1: INPUT ──────────────────────────────────────────────────────────
    with tab1:
        st.header("📂 Upload de Arquivos")
        st.markdown(
            "Envie os arquivos **PRODUCAO** (equipe de enfermagem) e **REPASSE** (hospital). "
            "Formatos aceitos: **Excel (.xlsx, .xls)**, PDF, Word (.docx), TXT e CSV."
        )

        col1, col2 = st.columns([2, 1])
        extratos_texto = []

        with col1:
            uploaded_files = st.file_uploader(
                "Upload de Arquivos de Endoscopia",
                type=["xlsx", "xls", "pdf", "docx", "txt", "csv"],
                accept_multiple_files=True,
                help="Envie os arquivos PRODUCAO e REPASSE",
            )

        with col2:
            if uploaded_files:
                st.subheader("Arquivos")
                for file in uploaded_files:
                    ext = file.name.rsplit(".", 1)[-1].lower()
                    icon = "📊" if ext in ("xlsx", "xls") else "📄"
                    st.success(f"{icon} {file.name}")
                    try:
                        file.seek(0, os.SEEK_END)
                        size = file.tell() / 1024
                        file.seek(0)
                        st.caption(f"{size:.1f} KB")
                    except Exception:
                        pass

        if uploaded_files:
            st.divider()
            _novos = [f for f in uploaded_files
                      if f"extract_{f.name}_{f.size}" not in st.session_state]
            if _novos:
                with st.spinner(f"📖 Extraindo dados de {len(_novos)} arquivo(s)..."):
                    for file in _novos:
                        _ck = f"extract_{file.name}_{file.size}"
                        st.session_state[_ck] = extract_text_from_file(file)
            for file in uploaded_files:
                _txt = st.session_state.get(f"extract_{file.name}_{file.size}")
                if _txt:
                    extratos_texto.append({"filename": file.name, "content": _txt})

            st.success(f"✅ {len(extratos_texto)} arquivo(s) processado(s)")
            st.session_state["extratos"] = extratos_texto

        if extratos_texto:
            st.divider()
            st.subheader("👁️ Preview dos Dados")
            for doc in extratos_texto:
                with st.expander(doc["filename"]):
                    text = doc["content"]
                    st.text_area("Dados RAW", value=text, height=400)
                    c1, c2, c3 = st.columns(3)
                    c1.metric("Caracteres", len(text))
                    c2.metric("Palavras", len(text.split()))
                    c3.metric("Linhas", len(text.splitlines()))
                    
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    nome_sem_ext = os.path.splitext(doc["filename"])[0]
                    st.download_button(
                        label="⬇️ Download do CSV RAW",
                        data=text.encode("utf-8-sig"),
                        file_name=f"raw_{nome_sem_ext}_{ts}.csv",
                        mime="text/csv",
                        type="primary",
                    )

    # ── TAB 2: EXECUÇÃO ───────────────────────────────────────────────────────
    with tab2:
        st.header("🚀 Executar Análise dos Arquivos")

        extratos = st.session_state.get("extratos", [])
        if not extratos:
            st.warning("Envie os arquivos na aba Input antes de executar.")

        # ── Seleção do modo de execução ───────────────────────────────────────
        st.subheader("⚙️ Modo de Processamento")

        modo_execucao = st.radio(
            "Escolha como os arquivos serão processados:",
            options=[
                "🔄 Transformação Local (sem API Key)",
                f"🤖 Agente IA — {provider} (requer API Key)",
            ],
            index=0,
            horizontal=True,
            help=(
                "**Transformação Local**: usa a função `transformar_csv_arquivo` diretamente, "
                "sem consumir tokens nem precisar de API Key. Recomendado para arquivos "
                "Excel/CSV nos formatos padrão PRODUCAO e REPASSE.\n\n"
                f"**Agente IA ({provider})**: usa o agente CrewAI com o modelo configurado "
                "na sidebar. Mais flexível para arquivos fora do padrão, mas requer API Key."
            ),
        )

        usa_llm = modo_execucao.startswith("🤖")

        # Descrição contextual do modo selecionado
        if usa_llm:
            st.info(
                f"🤖 **Modo LLM ativo** — O Agente Analista de Endoscopia ({provider}) irá interpretar "
                "cada arquivo e estruturar os dados em CSV. Requer **API Key** configurada na sidebar."
            )
            if not api_key and provider != "Ollama":
                st.warning("⚠️ Preencha a **API Key** na sidebar para usar este modo.")
        else:
            st.info(
                "🔄 **Modo Local ativo** — A função `transformar_csv_arquivo` será chamada diretamente, "
                "sem uso de IA. Mais rápido e sem custo de tokens. "
                "Ideal para arquivos Excel nos formatos padrão PRODUCAO e REPASSE."
            )

        st.divider()

        btn_label = f"🚀 Iniciar Análise com {provider}" if usa_llm else "🔄 Iniciar Transformação Local"
        btn_disabled = not extratos or (usa_llm and not api_key and provider != "Ollama")

        if st.button(btn_label, type="primary", disabled=btn_disabled):

            resultados = {}
            progress_bar = st.progress(0, text="Aguardando início...")

            # ── MODO LOCAL: transformar_csv_arquivo ───────────────────────────
            if not usa_llm:
                for i, doc in enumerate(extratos):
                    filename = doc["filename"]
                    text = doc["content"]
                    total = len(extratos)

                    progress_bar.progress(
                        i / total,
                        text=f"📂 Processando arquivo {i + 1} de {total}: **{filename}**",
                    )

                    with st.status(f"🔄 Transformando: {filename}", expanded=True) as status_box:
                        try:
                            csv_resultado = transformar_csv_arquivo(text, filename)
                            if csv_resultado:
                                resultados[filename] = csv_resultado
                                status_box.update(
                                    label=f"✅ Concluído: {filename}",
                                    state="complete",
                                    expanded=False,
                                )
                            else:
                                status_box.update(
                                    label=f"⚠️ Sem dados: {filename}",
                                    state="error",
                                    expanded=True,
                                )
                                st.warning(
                                    f"Nenhum dado extraído de **{filename}**. "
                                    "Verifique se o arquivo está no formato padrão PRODUCAO ou REPASSE."
                                )
                        except Exception as exc:
                            status_box.update(
                                label=f"❌ Erro em {filename}", state="error", expanded=True
                            )
                            st.error(f"Erro ao transformar **{filename}**: {exc}")
                            logger.error(
                                f"Erro em transformar_csv_arquivo({filename}): {exc}", exc_info=True
                            )

                    progress_bar.progress(
                        (i + 1) / total,
                        text=f"✅ {i + 1} de {total} arquivos concluídos",
                    )

                st.session_state["results"] = resultados
                # No modo local o CSV já vem puro — não precisa de extração adicional
                st.session_state["csvs_brutos"] = dict(resultados)
                st.success(
                    "🎉 Transformação concluída! Veja os resultados na aba **📊 Resultados** "
                    "ou inicie a correlação na aba **🔀 Correlação**."
                )

            # ── MODO LLM: CrewAI + LLM selecionado ───────────────────────────────
            else:
                llm = get_llm(provider, custom_model, temperature, api_key, base_url)
                agents = create_agents(llm, verbose_mode)

                for i, doc in enumerate(extratos):
                    filename = doc["filename"]
                    text = doc["content"]
                    total = len(extratos)

                    progress_bar.progress(
                        i / total,
                        text=f"📂 Processando arquivo {i + 1} de {total}: **{filename}**",
                    )

                    with st.status(
                        f"🤖 Agente trabalhando em: {filename}", expanded=True
                    ) as status_box:

                        st.markdown("##### 📡 Log de Execução em Tempo Real")
                        log_container = st.container(height=380, border=False)

                        log_queue: queue.Queue = queue.Queue(maxsize=500)
                        handler = StreamlitLogHandler(log_queue)
                        handler.setFormatter(
                            logging.Formatter("%(asctime)s | %(name)s | %(message)s",
                                              datefmt="%H:%M:%S")
                        )
                        root_logger = logging.getLogger()
                        root_logger.addHandler(handler)

                        thread_result: dict = {"value": None, "error": None}

                        def _run_crew(result_holder: dict, _text=text):
                            try:
                                tasks = create_tasks(agents, _text)
                                crew = Crew(
                                    agents=agents,
                                    tasks=tasks,
                                    process=Process.sequential,
                                    verbose=verbose_mode,
                                )
                                result_holder["value"] = str(crew.kickoff())
                            except Exception as exc:
                                result_holder["error"] = exc
                                logger.error(f"Erro durante kickoff: {exc}", exc_info=True)

                        crew_thread = threading.Thread(
                            target=_run_crew,
                            args=(thread_result,),
                            daemon=True,
                        )
                        crew_thread.start()

                        while crew_thread.is_alive():
                            drained = False
                            while not log_queue.empty():
                                line = log_queue.get_nowait()
                                render_log_line(line, log_container)
                                drained = True
                            if not drained:
                                time.sleep(0.15)

                        while not log_queue.empty():
                            render_log_line(log_queue.get_nowait(), log_container)

                        crew_thread.join()
                        root_logger.removeHandler(handler)

                        if thread_result["error"]:
                            status_box.update(
                                label=f"❌ Erro em {filename}", state="error", expanded=True
                            )
                            st.error(f"Erro: {thread_result['error']}")
                        else:
                            resultados[filename] = thread_result["value"]
                            status_box.update(
                                label=f"✅ Concluído: {filename}", state="complete", expanded=False
                            )

                    progress_bar.progress(
                        (i + 1) / total,
                        text=f"✅ {i + 1} de {total} arquivos concluídos",
                    )

                st.session_state["results"] = resultados
                st.session_state["csvs_brutos"] = {
                    fname: extrair_csv_do_texto(txt)
                    for fname, txt in resultados.items()
                }
                st.success(
                    "🎉 Análise concluída! Veja os resultados na aba **📊 Resultados** "
                    "ou inicie a correlação na aba **🔀 Correlação**."
                )

    # ── TAB 3: RESULTADOS ─────────────────────────────────────────────────────
    with tab3:
        st.header("📊 Resultados por Arquivo")
        results = st.session_state.get("results", None)

        if not results:
            st.info("Execute a análise na aba Execução")
        else:
            # ── Erros encontrados na PRODUCAO ─────────────────────────────────
            _csvs_brutos = st.session_state.get("csvs_brutos", {})
            _csv_prod_r3 = next(
                (txt for txt in _csvs_brutos.values()
                 if "TipoArquivo" in txt and len(txt.split("\n")) > 1
                 and "PRODUCAO" in txt.split("\n")[1]),
                None,
            )
            if _csv_prod_r3:
                _err_key = f"erros_prod_{hash(_csv_prod_r3[:300])}"
                if _err_key not in st.session_state:
                    _df_prod_r3 = texto_para_dataframe(_csv_prod_r3)
                    if _df_prod_r3 is not None and not _df_prod_r3.empty:
                        st.session_state[_err_key] = _verificar_erros_producao(_df_prod_r3)
                    else:
                        st.session_state[_err_key] = ({"total": 0}, pd.DataFrame())

                _resumo_err, _df_erros = st.session_state[_err_key]

                if _resumo_err["total"] > 0:
                    with st.expander(
                        f"🚨 Erros encontrados na PRODUÇÃO — **{_resumo_err['total']} linha(s)** com problemas",
                        expanded=True,
                    ):
                        _ce1, _ce2, _ce3, _ce4 = st.columns(4)
                        _ce1.metric("📅 Data inválida",        _resumo_err["data_invalida"])
                        _ce2.metric("👤 Paciente vazio",        _resumo_err["paciente_vazio"])
                        _ce3.metric("🔢 NrAtendimento vazio",   _resumo_err["nratendimento_vazio"])
                        _ce4.metric("🏥 Procedimento vazio",    _resumo_err["procedimento_vazio"])

                        # Colunas de dados: aceita tanto "Data" quanto "Data_PRODUCAO"
                        _cols_dados = [
                            next((c for c in [cn, cn + "_PRODUCAO"] if c in _df_erros.columns), None)
                            for cn in ["Data", "Paciente", "NrAtendimento", "Procedimento", "AbaOrigemDados"]
                        ]
                        _cols_exib = [
                            c for c in
                            [c for c in _cols_dados if c] +
                            ["❌ Data inválida", "❌ Paciente vazio",
                             "❌ NrAtendimento vazio", "❌ Procedimento vazio"]
                            if c in _df_erros.columns
                        ]
                        st.dataframe(_df_erros[_cols_exib], use_container_width=True)

                        _ts_err = datetime.now().strftime("%Y%m%d_%H%M%S")
                        st.download_button(
                            label="⬇️ Download erros de PRODUÇÃO (CSV)",
                            data=_df_erros.rename_axis("Linha").to_csv(index=True, encoding="utf-8-sig").encode("utf-8-sig"),
                            file_name=f"erros_producao_{_ts_err}.csv",
                            mime="text/csv",
                        )
                else:
                    st.success("✅ Nenhum erro encontrado nos dados de PRODUÇÃO.")

            # ── Resultados por arquivo ─────────────────────────────────────────
            for file, result in results.items():
                st.subheader(file)

                st.divider()
                st.subheader("📄 Resultado Consolidado")
                st.text_area("Relatório completo", value=result, height=400)
                c1, c2, c3 = st.columns(3)
                c1.metric("Caracteres", len(result))
                c2.metric("Palavras", len(result.split()))
                c3.metric("Linhas", len(result.splitlines()))
                
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                nome_sem_ext = os.path.splitext(file)[0]
                st.download_button(
                    label="⬇️ Download do CSV Processado",
                    data=result.encode("utf-8-sig"),
                    file_name=f"processado_{nome_sem_ext}_{ts}.csv",
                    mime="text/csv",
                    type="primary",
                )

    # ── TAB 4: CORRELAÇÃO ─────────────────────────────────────────────────────
    with tab4:
        st.header("🔀 Correlação — PRODUCAO × REPASSE")
        st.markdown(
            "Gera um **único CSV correlacionado** a partir de todos os arquivos processados, "
            "confrontando a produtividade da equipe com os repasses do hospital. "
            "Identifica glosas, divergências de valor e procedimentos não faturados."
        )

        csvs_brutos = st.session_state.get("csvs_brutos", {})

        if not csvs_brutos:
            st.info("⬅️ Primeiro processe os arquivos na aba **🚀 Execução**.")
        else:
            # Identifica arquivos PRODUCAO e REPASSE
            csv_producao = None
            csv_repasse = None
            nome_producao = ""
            nome_repasse = ""
            
            for fname, csv_txt in csvs_brutos.items():
                # Detecta tipo pelo conteúdo
                if "TipoArquivo" in csv_txt:
                    linhas = csv_txt.split("\n")
                    if len(linhas) > 1:
                        # Verifica primeira linha de dados
                        if "PRODUCAO" in linhas[1]:
                            csv_producao = csv_txt
                            nome_producao = fname
                        elif "REPASSE" in linhas[1]:
                            csv_repasse = csv_txt
                            nome_repasse = fname
                            st.session_state["csv_repasse_raw"] = csv_txt.encode("utf-8")
            
            # Mostra preview dos arquivos identificados
            col_prev1, col_prev2 = st.columns(2)
            with col_prev1:
                if csv_producao:
                    st.success(f"✅ PRODUCAO: {nome_producao}")
                else:
                    st.warning("⚠️ Arquivo PRODUCAO não identificado")
            with col_prev2:
                if csv_repasse:
                    st.success(f"✅ REPASSE: {nome_repasse}")
                else:
                    st.warning("⚠️ Arquivo REPASSE não identificado")
            
            if not csv_producao or not csv_repasse:
                st.error(
                    "❌ É necessário ter exatamente 1 arquivo PRODUCAO e 1 arquivo REPASSE processados. "
                    "Verifique se os arquivos foram processados corretamente na aba Execução."
                )
            else:
                # Informações dos arquivos testados — cacheado para não re-parsear a cada render
                _prev_key = f"corr_prev_{hash((csv_producao + csv_repasse)[:300])}"
                if _prev_key not in st.session_state:
                    _dp = texto_para_dataframe(csv_producao)
                    _dr = texto_para_dataframe(csv_repasse)
                    st.session_state[_prev_key] = (
                        _dp if _dp is not None else pd.DataFrame(),
                        _dr if _dr is not None else pd.DataFrame(),
                    )
                df_prod_prev, df_rep_prev = st.session_state[_prev_key]
                linhas_prod = len(df_prod_prev)
                linhas_rep  = len(df_rep_prev)

                st.markdown("### 📁 Arquivos para serem correlacionados")
                col_info1, col_info2 = st.columns(2)
                with col_info1:
                    st.metric("**PRODUCAO:**", f"{linhas_prod} linhas")
                with col_info2:
                    st.metric("**REPASSE:**", f"{linhas_rep} linhas")

                with st.expander(f"📋 Preview dos CSVs ({len(csvs_brutos)} arquivo(s))", expanded=False):
                    st.markdown("**PRODUCAO:**")
                    if not df_prod_prev.empty:
                        st.dataframe(df_prod_prev.head(5), use_container_width=True)

                    st.markdown("**REPASSE:**")
                    if not df_rep_prev.empty:
                        st.dataframe(df_rep_prev.head(5), use_container_width=True)

                st.divider()
                
                # Seleção do modo de correlação
                st.subheader("⚙️ Modo de Correlação")
                
                modo_correlacao = st.radio(
                    "Escolha como realizar a correlação:",
                    options=[
                        "🔄 Correlação Local (sem API Key)",
                        f"🤖 Correlação com Agente IA — {provider} (requer API Key)",
                    ],
                    index=0,
                    horizontal=True,
                    help=(
                        "**Correlação Local**: usa a função `correlacionar_csv_arquivos` com "
                        "correspondência semântica de procedimentos. Rápido e sem custo de tokens.\n\n"
                        f"**Agente IA ({provider})**: usa o agente Correlacionador com LLM para "
                        "interpretação mais flexível. Requer API Key."
                    ),
                )
                
                usa_llm_corr = modo_correlacao.startswith("🤖")
                
                if usa_llm_corr:
                    st.info(
                        f"🤖 **Modo LLM ativo** — O Agente Correlacionador ({provider}) irá realizar "
                        "o batimento dos dados. Requer **API Key** configurada na sidebar."
                    )
                    if not api_key and provider != "Ollama":
                        st.warning("⚠️ Preencha a **API Key** na sidebar para usar este modo.")
                else:
                    st.info(
                        "🔄 **Modo Local ativo** — A função `correlacionar_csv_arquivos` será chamada "
                        "diretamente com correspondência semântica de procedimentos. Rápido e sem custo."
                    )
                
                st.divider()

                # ── Checkbox: excluir linhas com erro (só modo local) ─────────
                _err_key_t4 = f"erros_prod_{hash((csv_producao or '')[:300])}"
                if _err_key_t4 not in st.session_state:
                    if not df_prod_prev.empty:
                        st.session_state[_err_key_t4] = _verificar_erros_producao(df_prod_prev)
                    else:
                        st.session_state[_err_key_t4] = ({"total": 0}, pd.DataFrame())
                _resumo_t4, _ = st.session_state[_err_key_t4]
                _n_erros_t4 = _resumo_t4["total"]

                _col_btn, _col_chk = st.columns([2, 3])
                with _col_btn:
                    btn_label_corr    = f"🤖 Gerar Correlação com {provider}" if usa_llm_corr else "🔄 Gerar Correlação Local"
                    btn_disabled_corr = usa_llm_corr and not api_key and provider != "Ollama"
                    _clicked_corr = st.button(btn_label_corr, type="primary", disabled=btn_disabled_corr)
                with _col_chk:
                    if not usa_llm_corr:
                        _lbl_chk = (
                            f"Excluir dados de PRODUÇÃO com erros ({_n_erros_t4} linha(s))"
                            if _n_erros_t4 > 0
                            else "Excluir dados de PRODUÇÃO com erros (nenhum encontrado)"
                        )
                        _excluir_erros = st.checkbox(
                            _lbl_chk,
                            value=True,
                            disabled=_n_erros_t4 == 0,
                            help=(
                                "Quando marcado, as linhas com Data inválida, Paciente, "
                                "NrAtendimento ou Procedimento vazios são removidas antes "
                                "de iniciar a correlação."
                            ),
                        )
                    else:
                        _excluir_erros = False

                if _clicked_corr:

                    # ── Filtra PRODUCAO removendo linhas com erro (se checkbox ativo) ──
                    _csv_producao_corr = csv_producao
                    if not usa_llm_corr and _excluir_erros and _n_erros_t4 > 0:
                        _, _df_erros_t4 = st.session_state[_err_key_t4]
                        _df_prod_limpo = df_prod_prev.drop(index=_df_erros_t4.index, errors="ignore")
                        _csv_producao_corr = _df_prod_limpo.to_csv(index=False)
                        st.info(
                            f"🧹 {_n_erros_t4} linha(s) com erro excluída(s) da PRODUCAO antes da correlação. "
                            f"Correlacionando com {len(_df_prod_limpo)} linha(s)."
                        )

                    # ── MODO LOCAL: correlacionar_csv_arquivos (com thread para não travar UI) ──
                    if not usa_llm_corr:
                        with st.status("🔄 Correlacionando localmente...", expanded=True) as status_local:
                            st.markdown("##### ⏳ Aguarde — correlacionando PRODUCAO × REPASSE...")
                            _prog_local = st.progress(0, "Iniciando correlação...")

                            # Pré-carrega tabelas TUSS na thread principal.
                            # Limpa o cache antes de carregar para evitar retornar {}
                            # stale cacheado de uma execução anterior com arquivo ausente.
                            _tabela_tuss_pre: dict = {}
                            _valores_tuss_pre: dict = {}
                            try:
                                _carregar_tabela_tuss.clear()
                                _carregar_valores_tuss.clear()
                                _tabela_tuss_pre = _carregar_tabela_tuss()
                                _valores_tuss_pre = _carregar_valores_tuss()
                                logger.info(f"Tabelas TUSS pré-carregadas: {len(_tabela_tuss_pre)} entradas lookup, {len(_valores_tuss_pre)} entradas valores")
                            except Exception as _e_pre:
                                logger.warning(f"Pré-carga TUSS ignorada: {_e_pre}")

                            _res_local: dict = {"value": None, "error": None}

                            def _run_local_corr(
                                holder,
                                _prod=_csv_producao_corr,
                                _rep=csv_repasse,
                                _tab=_tabela_tuss_pre,
                                _vals=_valores_tuss_pre,
                            ):
                                try:
                                    holder["value"] = correlacionar_csv_arquivos(
                                        _prod, _rep,
                                        tabela_tuss_preloaded=_tab,
                                        valores_tuss_preloaded=_vals,
                                    )
                                except Exception as _exc:
                                    holder["error"] = _exc
                                    logger.error(f"Erro em correlacionar_csv_arquivos: {_exc}", exc_info=True)

                            _t_local = threading.Thread(
                                target=_run_local_corr, args=(_res_local,), daemon=True
                            )
                            _t_local.start()

                            _step = 0
                            _msgs = [
                                "Indexando REPASSE...", "Correlacionando por nome e data...",
                                "Aplicando fallbacks...", "Verificando códigos TUSS...",
                                "Finalizando...",
                            ]
                            while _t_local.is_alive():
                                _prog_local.progress(
                                    min(0.9, _step * 0.18),
                                    _msgs[min(_step, len(_msgs) - 1)],
                                )
                                _step += 1
                                time.sleep(0.8)

                            _t_local.join()
                            _prog_local.progress(1.0, "Concluído!")

                            if _res_local["error"]:
                                status_local.update(
                                    label="❌ Erro na correlação local",
                                    state="error", expanded=True
                                )
                                st.error(f"Erro: {_res_local['error']}")
                            elif _res_local["value"]:
                                st.session_state["csv_correlacionado"] = _res_local["value"]
                                # Invalida caches dependentes do resultado anterior
                                for _k in list(st.session_state.keys()):
                                    if (
                                        _k.startswith("corr_df_")
                                        or _k.startswith("corr_style_")
                                        or _k.startswith("corr_csv_dl_")
                                        or _k.startswith("cob_")
                                        or _k == "corr_show_table"
                                    ):
                                        del st.session_state[_k]
                                status_local.update(
                                    label="✅ Correlação local concluída!",
                                    state="complete", expanded=False
                                )
                            else:
                                status_local.update(
                                    label="❌ Erro na correlação local",
                                    state="error", expanded=True
                                )
                                st.error("Não foi possível gerar a correlação. Verifique os logs.")
                    
                    # ── MODO LLM: Agente Correlacionador ──────────────────────────
                    else:
                        with st.status("🤖 Agente Correlacionador trabalhando...", expanded=True) as status_corr:
                            st.markdown("##### 📡 Log de Correlação em Tempo Real")
                            log_container_corr = st.container(height=320, border=False)

                            log_queue_corr: queue.Queue = queue.Queue(maxsize=500)
                            handler_corr = StreamlitLogHandler(log_queue_corr)
                            handler_corr.setFormatter(
                                logging.Formatter("%(asctime)s | %(name)s | %(message)s", datefmt="%H:%M:%S")
                            )
                            root_logger = logging.getLogger()
                            root_logger.addHandler(handler_corr)

                            thread_result_corr: dict = {"value": None, "error": None}

                            def _run_correlation(result_holder: dict):
                                try:
                                    llm_corr = get_llm(
                                        provider, custom_model, temperature, api_key, base_url
                                    )
                                    correlator = create_correlator_agent(llm_corr, verbose_mode)
                                    # Passa apenas PRODUCAO e REPASSE
                                    csvs_para_correlacao = {
                                        nome_producao: csv_producao,
                                        nome_repasse: csv_repasse
                                    }
                                    task_corr = create_correlation_task(correlator, csvs_para_correlacao)
                                    crew_corr = Crew(
                                        agents=[correlator],
                                        tasks=[task_corr],
                                        process=Process.sequential,
                                        verbose=verbose_mode,
                                    )
                                    result_holder["value"] = str(crew_corr.kickoff())
                                except Exception as exc:
                                    result_holder["error"] = exc
                                    logger.error(f"Erro na correlação: {exc}", exc_info=True)

                            thread_corr = threading.Thread(
                                target=_run_correlation,
                                args=(thread_result_corr,),
                                daemon=True,
                            )
                            thread_corr.start()

                            while thread_corr.is_alive():
                                drained = False
                                while not log_queue_corr.empty():
                                    render_log_line(log_queue_corr.get_nowait(), log_container_corr)
                                    drained = True
                                if not drained:
                                    time.sleep(0.15)

                            while not log_queue_corr.empty():
                                render_log_line(log_queue_corr.get_nowait(), log_container_corr)

                            thread_corr.join()
                            root_logger.removeHandler(handler_corr)

                            if thread_result_corr["error"]:
                                status_corr.update(
                                    label="❌ Erro na correlação", state="error", expanded=True
                                )
                                st.error(f"Erro: {thread_result_corr['error']}")
                            else:
                                status_corr.update(
                                    label="✅ Correlação concluída!", state="complete", expanded=False
                                )
                                st.session_state["csv_correlacionado"] = thread_result_corr["value"]
                                # Invalida caches dependentes
                                for _k in list(st.session_state.keys()):
                                    if (
                                        _k.startswith("corr_df_")
                                        or _k.startswith("corr_style_")
                                        or _k.startswith("corr_csv_dl_")
                                        or _k.startswith("cob_")
                                        or _k == "corr_show_table"
                                    ):
                                        del st.session_state[_k]

            # ── Exibe resultado correlacionado ────────────────────────────────
            csv_correlacionado_raw = st.session_state.get("csv_correlacionado")

            if csv_correlacionado_raw:
                st.divider()

                # Cache do parse — evita re-parsear 5000+ linhas a cada interação de filtro
                _df_cache_key = f"corr_df_{hash(csv_correlacionado_raw[:300])}"
                if _df_cache_key not in st.session_state:
                    _csv_limpo = extrair_csv_do_texto(csv_correlacionado_raw)
                    _csv_dados_str, _resumo_str = separar_resumo_do_csv(_csv_limpo)
                    st.session_state[_df_cache_key] = (
                        texto_para_dataframe(_csv_dados_str),
                        _csv_dados_str,
                        _resumo_str,
                    )
                df_final, csv_dados_str, resumo_str = st.session_state[_df_cache_key]

                st.subheader("📊 Tabela Correlacionada")

                if df_final is not None and not df_final.empty:
                    # ── Deduplicação intra-lote ───────────────────────────────
                    # Usa as mesmas 8 colunas de input bruto do trigger Supabase.
                    # Remove linhas 100 % idênticas que o hospital enviou em duplicata
                    # (ex: mesmo registro REPASSE aparecendo duas vezes no arquivo).
                    # df_metrics é usado para métricas, filtros e upload.
                    # df_final (raw) permanece intacto para o download CSV.
                    import hashlib as _hl
                    def _hash8(row):
                        _f = ["ChaveCorrelacao","ProcedimentosAdicionais_PRODUCAO",
                              "MedicoExecutor_REPASSE","NrRepasse_REPASSE",
                              "AbaOrigemDados_REPASSE","ValorLiberado_REPASSE",
                              "NrInternoConta_REPASSE","Via_REPASSE"]
                        return _hl.md5("|".join(str(row.get(c,"") or "") for c in _f).encode()).hexdigest()
                    _h8        = df_final.apply(_hash8, axis=1)
                    _n_dupes   = int(_h8.duplicated().sum())
                    df_metrics = df_final[~_h8.duplicated(keep="first")].copy()
                    if _n_dupes > 0:
                        _mask_dups    = _h8.duplicated(keep=False)
                        _df_dups      = df_final[_mask_dups].copy()
                        _df_dups["_h8"] = _h8[_mask_dups]
                        _n_dup_groups = int(_df_dups["_h8"].nunique())
                    else:
                        _df_dups, _n_dup_groups = pd.DataFrame(), 0

                    col_m1, col_m2, col_m3, col_m4, col_m5 = st.columns(5)
                    total_linhas = len(df_metrics)

                    def _perc(n, base):
                        return f"{n / base * 100:.1f}%" if base > 0 else "0.0%"

                    status_col = df_metrics.get("StatusCorrelacao", pd.Series(dtype=str))
                    mm = df_metrics.get("MetodoMatch", pd.Series(dtype=str)).fillna("")
                    # Correlacionados = qualquer status que comece com CORRELACIONADO
                    n_correlacionado           = status_col.str.upper().str.startswith("CORRELACIONADO").sum()
                    n_repasse_nao_identificado = (status_col.str.upper() == "REPASSE_NAO_IDENTIFICADO_NA_PRODUCAO").sum()
                    n_fora_periodo             = (status_col.str.upper() == "REPASSE_DATA_FORA_DO_PERIODO_PRODUCAO").sum()
                    n_nao_faturado             = (status_col.str.upper() == "NAO_FATURADO_NO_REPASSE").sum()
                    n_proc_divergente          = mm.str.contains("PROCEDIMENTO_DIVERGENTE").sum()

                    col_m1.metric("📄 Total de Linhas",                   total_linhas)
                    col_m2.metric("✅ Correlacionados",                    n_correlacionado)
                    col_m3.metric("⚠️ Repasse não identificado na produção", int(n_repasse_nao_identificado))
                    col_m4.metric("📅 Repasse fora do período de produção",  int(n_fora_periodo))
                    col_m5.metric("❌ Não Faturados",                       int(n_nao_faturado))

                    if n_proc_divergente > 0:
                        st.warning(
                            f"🔬 **{int(n_proc_divergente)} correlações com procedimento divergente** "
                            f"({_perc(n_proc_divergente, n_correlacionado)} dos correlacionados) — "
                            "PRODUCAO e REPASSE têm procedimentos anatomicamente distintos. Será Revisado manualmente.",
                            icon=None,
                        )

                    if _n_dupes > 0:
                        st.markdown(
                            f"⚠️ **{_n_dupes} linha(s) duplicada(s)** detectada(s) no arquivo fonte "
                            f"(mesmo registro REPASSE enviado múltiplas vezes pelo hospital) — "
                            "excluídas das métricas e da carga."
                        )
                        with st.expander(
                            f"🔍 Ver linhas duplicadas — {_n_dup_groups} grupo(s), {_n_dupes} linha(s)",
                            expanded=False,
                        ):
                            _dup_cols = [c for c in [
                                "Paciente_PRODUCAO", "Data_PRODUCAO", "Procedimento_PRODUCAO",
                                "StatusCorrelacao", "StatusTUSS",
                                "NrRepasse_REPASSE", "NrInternoConta_REPASSE",
                                "ValorLiberado_REPASSE", "AbaOrigemDados_REPASSE",
                            ] if c in _df_dups.columns]
                            st.dataframe(
                                _df_dups[_dup_cols].reset_index(drop=True),
                                use_container_width=True,
                            )
                            _ts_dup = datetime.now().strftime("%Y%m%d_%H%M%S")
                            st.download_button(
                                label="⬇️ Download duplicatas (CSV)",
                                data=_df_dups[_dup_cols].to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig"),
                                file_name=f"duplicatas_repasse_{_ts_dup}.csv",
                                mime="text/csv",
                            )

                    # Métricas detalhadas de correlação
                    st.markdown("---")
                    st.markdown("### 📊 Resultados da Correlação")

                    n_companion = (mm == "5_FALLBACK_COMPANION_PROCEDIMENTO_ADICIONAL").sum()

                    # Contagem por MetodoMatch (inclui sufixo _PROCEDIMENTO_DIVERGENTE quando aplicável)
                    n_m1  = mm.str.startswith("1_NOME_COMPLETO_DATA_PROCEDIMENTO").sum()
                    n_m2  = mm.str.startswith("2_FALLBACK_NR-ATENDIMENTO").sum()
                    n_m3  = mm.str.startswith("3_FALLBACK_NOME_PARCIAL_FUZZY_DATA_FIXA").sum()
                    n_m4  = mm.str.startswith("4_FALLBACK_NOME_COMPLETO_DATA-FLEXIVEL").sum()
                    n_m5  = (mm == "5_FALLBACK_COMPANION_PROCEDIMENTO_ADICIONAL").sum()
                    n_sem = (mm == "SEM_MATCH").sum()

                    # ── Linha 1: métricas de matches por método ───────────────
                    st.markdown("#### STATUS CORRELAÇÃO: Matches por método de correlação")
                    mc1, mc2, mc3, mc4, mc5 = st.columns(5)
                    mc1.metric(
                        "1 Nome + Data + Proc",
                        f"{n_m1}",
                        delta=_perc(n_m1, total_linhas),
                        delta_color="off",
                        help="Nome completo normalizado + Data ±1 dia + Procedimento",
                    )
                    mc2.metric(
                        "2 NrAtendimento",
                        f"{n_m2}",
                        delta=_perc(n_m2, total_linhas),
                        delta_color="off",
                        help="NrAtendimento + Data ±1 dia + Procedimento (nome não encontrado)",
                    )
                    mc3.metric(
                        "3 Nome Muito Parecido",
                        f"{n_m3}",
                        delta=_perc(n_m3, total_linhas),
                        delta_color="off",
                        help="≥3 tokens do nome com similaridade fuzzy + Data ±1 dia + Procedimento",
                    )
                    mc4.metric(
                        "4 Data Flexível",
                        f"{n_m4}",
                        delta=_perc(n_m4, total_linhas),
                        delta_color="off",
                        help="Nome completo normalizado + Data ±7 dias + Procedimento",
                    )
                    mc5.metric(
                        "5 Proc. Adicional",
                        f"{n_m5}",
                        delta=_perc(n_m5, total_linhas),
                        delta_color="off",
                        help="Urease/Helicobacter companion: principal já correlacionado no mesmo episódio",
                    )

                    # ── STATUS TUSS (se disponível) ──────────────────────────
                    tuss_col = df_metrics.get("StatusTUSS", pd.Series(dtype=str)).fillna("")

                    # Grupos semânticos por prefixo
                    n_tuss_ok_total     = tuss_col.str.startswith("OK_").sum()
                    n_tuss_cobrar_total = tuss_col.str.startswith("COBRAR_").sum()
                    n_tuss_manual_total = tuss_col.str.startswith("CORRELACIONAR_MANUAL_").sum()
                    n_tuss_com_status   = int(n_tuss_ok_total + n_tuss_cobrar_total + n_tuss_manual_total)

                    # Desdobramento do COBRAR: correlacionados vs não faturados
                    _cobrar_mask_col = tuss_col.str.startswith("COBRAR_")
                    n_cobrar_correlac = int(
                        (_cobrar_mask_col & status_col.str.startswith("CORRELACIONADO")).sum()
                    )
                    n_cobrar_nao_fat  = int((tuss_col == "COBRAR_TUSS_NAO_FATURADO_MAPEADO").sum())

                    if n_tuss_com_status > 0:
                        st.markdown("#### STATUS TUSS: Verificação de Repasse com base no código TUSS dos procedimentos")
                        tc1, tc2a, tc2b, tc3 = st.columns(4)
                        tc1.metric(
                            "✅ OK — Sem ação necessária",
                            int(n_tuss_ok_total),
                            delta=_perc(n_tuss_ok_total, total_linhas),
                            delta_color="off",
                            help="Código TUSS do repasse correto ou superior ao esperado (inclui adicionais reconhecidos e incorporados)",
                        )
                        tc2a.metric(
                            "🔴 COBRAR — Correlacionados",
                            n_cobrar_correlac,
                            delta=_perc(n_cobrar_correlac, total_linhas),
                            delta_color="off",
                            help=(
                                "Procedimento correlacionado, mas o código TUSS do repasse está divergente, "
                                "mais simples (downgrade) ou ausente — cobrança sobre repasse já existente. "
                                "Inclui: Downgrade, Divergente, Adicional Cobrado como Simples, Adicional Ausente."
                            ),
                        )
                        tc2b.metric(
                            "🔴 COBRAR — Não Faturados",
                            n_cobrar_nao_fat,
                            delta=_perc(n_cobrar_nao_fat, total_linhas),
                            delta_color="off",
                            help=(
                                "Procedimento ausente no repasse (não faturado pelo hospital), "
                                "mas com código TUSS identificado no mapeamento — "
                                "cobrança sobre repasse inexistente."
                            ),
                        )
                        tc3.metric(
                            "🟡 Revisar Manualmente",
                            int(n_tuss_manual_total),
                            delta=_perc(n_tuss_manual_total, total_linhas),
                            delta_color="off",
                            help="Combinação sem mapeamento TUSS ou repasse sem produção correspondente — requer análise manual",
                        )

                        # ── Detalhamento por status individual ─────────────────
                        _TUSS_DETAIL: list[tuple[str, str, str]] = [
                            ("OK_TUSS_PROC_PRINCIPAL_OK",                           "Proc. Principal Correto",                   "✅ OK"),
                            ("OK_TUSS_ADICIONAL_INCORPORADO_NO_PRINCIPAL",          "Adicional Incorporado no Principal",        "✅ OK"),
                            ("OK_TUSS_PROC_ADICIONAL_RECONHECIDO",                  "Adicional Faturado com Código Correto",     "✅ OK"),
                            ("OK_TUSS_TODOS_CODIGOS_ADICIONAIS_FATURADOS",          "Todos os Adicionais com Código Correto",    "✅ OK"),
                            ("OK_TUSS_CODIGO_PRINCIPAL_UPGRADE",                    "Código Superior ao Esperado (Upgrade)",     "✅ OK"),
                            ("COBRAR_TUSS_PROC_ADICIONAL_COBRADO_COMO_SIMPLES",     "Adicional Cobrado como Simples",            "🔴 COBRAR"),
                            ("COBRAR_TUSS_CODIGO_ADICIONAL_AUSENTE_NO_REPASSE",     "Adicional Ausente no Repasse",              "🔴 COBRAR"),
                            ("COBRAR_TUSS_CODIGO_PRINCIPAL_DIVERGENTE",             "Código Principal Divergente",               "🔴 COBRAR"),
                            ("COBRAR_TUSS_CODIGO_PRINCIPAL_DOWNGRADE",              "Código Principal Mais Simples (Downgrade)", "🔴 COBRAR"),
                            ("COBRAR_TUSS_NAO_FATURADO_MAPEADO",                    "Não Cobrado – TUSS Identificado",           "🔴 COBRAR"),
                            ("CORRELACIONAR_MANUAL_TUSS_REPASSE_SEM_PRODUCAO",      "Repasse sem Registro de Produção",          "🟡 Manual"),
                            ("CORRELACIONAR_MANUAL_TUSS_COMBINACAO_SEM_MAPEAMENTO", "Combinação sem Mapeamento TUSS",            "🟡 Manual"),
                        ]
                        _detail_rows = [
                            {"Grupo": grupo, "Status": label, "Qtd": cnt, "%": _perc(cnt, total_linhas)}
                            for status, label, grupo in _TUSS_DETAIL
                            if (cnt := int((tuss_col == status).sum())) > 0
                        ]
                        if _detail_rows:
                            with st.expander("📋 Detalhamento por StatusTUSS", expanded=False):
                                st.dataframe(
                                    pd.DataFrame(_detail_rows),
                                    use_container_width=True,
                                    hide_index=True,
                                )

                        # ── Estimativa de impacto financeiro (financeiro/admin) ──
                        _val_col = df_metrics.get("ValorEstimado_TUSS", pd.Series(dtype=str))
                        _val_num = pd.to_numeric(_val_col, errors="coerce").dropna()
                        if not _val_num.empty and can_view_financial():
                            st.markdown("---")
                            st.markdown("#### 💰 Estimativa de Impacto Financeiro")
                            _total_est    = _val_num.sum()
                            _n_com_valor  = len(_val_num)
                            _cobrar_mask  = tuss_col.str.startswith("COBRAR_")
                            _n_cobrar_sem_valor = max(
                                0,
                                int(_cobrar_mask.sum()) -
                                int(pd.to_numeric(_val_col[_cobrar_mask], errors="coerce").notna().sum()),
                            )
                            iv1, iv2, iv3 = st.columns(3)
                            iv1.metric(
                                "💵 Valor Total Estimado",
                                f"R$ {_total_est:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."),
                                help="Soma dos UltimoValor por convênio para todos os itens elegíveis à cobrança",
                            )
                            iv2.metric(
                                "✅ Itens com estimativa",
                                int(_n_com_valor),
                                help="Itens cujo código TUSS consta em tuss_valores.csv",
                            )
                            iv3.metric(
                                "⚠️ Itens COBRAR sem estimativa",
                                _n_cobrar_sem_valor,
                                help="Itens com status COBRAR_ mas sem histórico de valor TUSS cadastrado",
                            )
                            # Breakdown por convênio
                            _conv_col = df_metrics.get("Convenio_PRODUCAO", df_metrics.get("Convenio", pd.Series(dtype=str)))
                            _mask_elig = _val_col.replace("", float("nan")).notna()
                            if _mask_elig.any():
                                _df_breakdown = (
                                    df_metrics[_mask_elig]
                                    .assign(_val_num=pd.to_numeric(_val_col[_mask_elig], errors="coerce"))
                                    .groupby(_conv_col[_mask_elig].fillna("(sem convênio)"))
                                    .agg(Itens=("_val_num", "count"), ValorTotal=("_val_num", "sum"))
                                    .sort_values("ValorTotal", ascending=False)
                                    .reset_index()
                                )
                                _df_breakdown.columns = ["Convênio", "Itens", "Valor Total (R$)"]
                                _df_breakdown["Valor Total (R$)"] = _df_breakdown["Valor Total (R$)"].map(
                                    lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                                )
                                with st.expander("📊 Breakdown por convênio", expanded=False):
                                    st.dataframe(_df_breakdown, use_container_width=True, hide_index=True)

                    st.markdown("---")

                    with st.expander("🔍 Filtros", expanded=False):
                        fcol1, fcol2, fcol3, fcol4, fcol5 = st.columns(5)

                        convenios_disp = sorted(df_metrics.get("Convenio_PRODUCAO", pd.Series()).dropna().unique().tolist())
                        filtro_convenio = fcol1.multiselect(
                            "Convênio", convenios_disp,
                            placeholder="Todos",
                        )

                        _status_corr_raw = sorted(df_metrics.get("StatusCorrelacao", pd.Series()).dropna().unique().tolist())
                        _status_corr_labels = sorted([LABEL_STATUS_CORR.get(s, s) for s in _status_corr_raw])
                        _filtro_status_labels = fcol2.multiselect(
                            "Status Correlação", _status_corr_labels,
                            placeholder="Todos",
                        )
                        filtro_status = [_CORR_LABEL_TO_RAW.get(l, l) for l in _filtro_status_labels]

                        tuss_disp = sorted(df_metrics.get("CodigoTUSS_REPASSE", pd.Series()).dropna().unique().tolist())
                        filtro_tuss = fcol3.multiselect(
                            "Código TUSS", tuss_disp,
                            placeholder="Todos",
                        )

                        _status_tuss_raw = sorted(df_metrics.get("StatusTUSS", pd.Series()).dropna().unique().tolist())
                        _status_tuss_labels = sorted([LABEL_STATUS_TUSS.get(s, s) for s in _status_tuss_raw])
                        _filtro_status_tuss_labels = fcol4.multiselect(
                            "Status TUSS", _status_tuss_labels,
                            placeholder="Todos",
                        )
                        filtro_status_tuss = [_TUSS_LABEL_TO_RAW.get(l, l) for l in _filtro_status_tuss_labels]

                        # Filtro de ValorEstimado_TUSS (com/sem valor)
                        filtro_valor_est = fcol5.selectbox(
                            "Valor Estimado",
                            ["Todos", "Com Valor", "Sem Valor"],
                            index=0,
                        )

                    # Filtro com máscara booleana — sem copiar o DataFrame inteiro
                    # Seleção vazia = sem filtro (exibe tudo); coerente com placeholder="Todos"
                    _mask = pd.Series(True, index=df_metrics.index)
                    if "Convenio_PRODUCAO" in df_metrics.columns and filtro_convenio:
                        _mask &= df_metrics["Convenio_PRODUCAO"].isin(filtro_convenio)
                    if "StatusCorrelacao" in df_metrics.columns and filtro_status:
                        _mask &= df_metrics["StatusCorrelacao"].isin(filtro_status)
                    if "CodigoTUSS_REPASSE" in df_metrics.columns and filtro_tuss:
                        _mask &= df_metrics["CodigoTUSS_REPASSE"].isin(filtro_tuss)
                    if "StatusTUSS" in df_metrics.columns and filtro_status_tuss:
                        _mask &= df_metrics["StatusTUSS"].isin(filtro_status_tuss)
                    if "ValorEstimado_TUSS" in df_metrics.columns and filtro_valor_est != "Todos":
                        if filtro_valor_est == "Com Valor":
                            _mask &= df_metrics["ValorEstimado_TUSS"].replace("", float("nan")).notna()
                        else:  # Sem Valor
                            _mask &= df_metrics["ValorEstimado_TUSS"].replace("", float("nan")).isna()
                    df_filtrado = df_metrics[_mask]
                    _n_filtrado = len(df_filtrado)

                    # Styling vetorizado (axis=None): ~60× mais rápido que apply(axis=1) por linha
                    def _build_style_df(df: pd.DataFrame) -> pd.DataFrame:
                        import numpy as _np
                        # Índice alinhado — necessário após filtro booleano criar índice esparso
                        _empty = pd.Series("", index=df.index, dtype=str)
                        _st  = (df["StatusCorrelacao"].fillna("").str.upper()
                                if "StatusCorrelacao" in df.columns else _empty)
                        _tss = (df["StatusTUSS"].fillna("")
                                if "StatusTUSS" in df.columns else _empty)
                        _mm  = (df["MetodoMatch"].fillna("").str.upper()
                                if "MetodoMatch" in df.columns else _empty)
                        cor  = pd.Series("", index=df.index, dtype=str)
                        cor[_st.str.contains("NAO_IDENTIFICADO",     na=False)] = "background-color: #e2e3e5"
                        cor[_st.str.contains("DATA_FORA_DO_PERIODO", na=False)] = "background-color: #e2e3e5; color: #888"
                        cor[_st.str.contains("NAO_FATURADO",         na=False)] = "background-color: #f8d7da"
                        cor[_st == "CORRELACIONADO"]                             = "background-color: #d4edda"
                        cor[_mm.str.contains("3_FALLBACK_NOME_PARCIAL",  na=False)] = "background-color: #fff3cd"
                        cor[_mm.str.contains("4_FALLBACK_NOME_COMPLETO", na=False)] = "background-color: #d1ecf1"
                        cor[_mm.str.contains("2_FALLBACK_NR-ATENDIMENTO",na=False)] = "background-color: #cce5ff"
                        cor[_mm == "5_FALLBACK_COMPANION_PROCEDIMENTO_ADICIONAL"]    = "background-color: #d4edda"
                        cor[_mm.str.contains("PROCEDIMENTO_DIVERGENTE",  na=False)] = "background-color: #ffc107; color: #000"
                        # TUSS sobrescreve — máxima prioridade (colorização por prefixo)
                        cor[_tss.str.startswith("OK_")]                   = "background-color: #d4edda"
                        cor[_tss.str.startswith("COBRAR_")]               = "background-color: #c0392b; color: #fff"
                        cor[_tss.str.startswith("CORRELACIONAR_MANUAL_")] = "background-color: #fff3cd"
                        return pd.DataFrame(
                            _np.repeat(cor.values[:, None], len(df.columns), axis=1),
                            columns=df.columns, index=df.index,
                        )

                    # ── Camada 1: botão gate ──────────────────────────────────
                    # A tabela só é renderizada ao clicar — evita re-render em
                    # qualquer outra interação da aba (filtros, sidebar, etc.)
                    _btn_col, _ocultar_col, _ = st.columns([2, 2, 6])
                    if not st.session_state.get("corr_show_table", False):
                        if _btn_col.button(
                            "👁️ Exibir tabela de resultados",
                            key="btn_show_corr_table",
                            type="primary",
                            help=f"{_n_filtrado:,} linhas disponíveis",
                        ):
                            st.session_state["corr_show_table"] = True
                            st.rerun()
                    else:
                        if _btn_col.button("🔒 Ocultar tabela", key="btn_hide_corr_table"):
                            st.session_state["corr_show_table"] = False
                            st.rerun()

                        # ── Camada 3: cap de linhas na exibição interativa ────
                        # Reduz custo de serialização do st.dataframe pelo WebSocket
                        _MAX_ROWS_DISPLAY = 200
                        if _n_filtrado > _MAX_ROWS_DISPLAY:
                            st.info(
                                f"ℹ️ Exibindo as primeiras **{_MAX_ROWS_DISPLAY:,}** de "
                                f"**{_n_filtrado:,}** linhas filtradas. "
                                "Use os filtros para refinar ou baixe o CSV completo abaixo."
                            )
                            df_display = df_filtrado.head(_MAX_ROWS_DISPLAY)
                        else:
                            df_display = df_filtrado

                        # ── Ocultar colunas financeiras para editor ───────────
                        df_display = _strip_financial(df_display)

                        # ── Reordenar colunas: pinadas primeiro ───────────────
                        _colunas_pinadas = ["StatusCorrelacao", "CodigoTUSS_REPASSE", "StatusTUSS", "CodigosTUSS_Esperados"]
                        _colunas_existentes = [c for c in _colunas_pinadas if c in df_display.columns]
                        _colunas_restantes = [c for c in df_display.columns if c not in _colunas_pinadas]
                        df_display = df_display[_colunas_existentes + _colunas_restantes]

                        # ── Camada 2: cache do style por combinação de filtro ─
                        # _build_style_df só reexecuta quando os filtros mudam;
                        # interações não relacionadas reutilizam o resultado cacheado
                        _filter_hash = hash((
                            frozenset(filtro_convenio),
                            frozenset(filtro_status),
                            frozenset(filtro_tuss),
                            frozenset(filtro_status_tuss),
                            filtro_valor_est,
                        ))
                        _style_cache_key = f"corr_style_{_df_cache_key}_{_filter_hash}"
                        if _style_cache_key not in st.session_state:
                            # Remove caches de style anteriores para não acumular memória
                            for _sk in [k for k in st.session_state if k.startswith("corr_style_")]:
                                del st.session_state[_sk]
                            st.session_state[_style_cache_key] = _build_style_df(df_display)
                        _cached_style = st.session_state[_style_cache_key]

                        # Aplica labels amigáveis apenas na camada de exibição
                        df_display_labeled = df_display.copy()
                        if "StatusCorrelacao" in df_display_labeled.columns:
                            df_display_labeled["StatusCorrelacao"] = (
                                df_display_labeled["StatusCorrelacao"]
                                .map(lambda x: LABEL_STATUS_CORR.get(x, x) if pd.notna(x) else x)
                            )
                        if "StatusTUSS" in df_display_labeled.columns:
                            df_display_labeled["StatusTUSS"] = (
                                df_display_labeled["StatusTUSS"]
                                .map(lambda x: LABEL_STATUS_TUSS.get(x, x) if pd.notna(x) else x)
                            )

                        st.dataframe(
                            df_display_labeled.style.apply(lambda _: _cached_style, axis=None),
                            use_container_width=True,
                            height=460,
                            column_config={
                                "StatusCorrelacao": st.column_config.TextColumn("Status Correlação", pinned=True),
                                "CodigoTUSS_REPASSE": st.column_config.TextColumn("Código TUSS Repasse", pinned=True),
                                "StatusTUSS": st.column_config.TextColumn("Status TUSS", pinned=True),
                                "CodigosTUSS_Esperados": st.column_config.TextColumn("Códigos TUSS Esperados", pinned=True),
                            }
                        )

                    # ── Download CSV — sem cap, sempre disponível ─────────────
                    # Cacheado em session_state para evitar to_csv() a cada render
                    _csv_dl_key = f"corr_csv_dl_{_df_cache_key}_{hash((frozenset(filtro_convenio), frozenset(filtro_status), frozenset(filtro_tuss), frozenset(filtro_status_tuss), filtro_valor_est))}"
                    if _csv_dl_key not in st.session_state:
                        for _ck in [k for k in st.session_state if k.startswith("corr_csv_dl_")]:
                            del st.session_state[_ck]
                        st.session_state[_csv_dl_key] = _strip_financial(df_filtrado).to_csv(
                            index=False, sep=",", encoding="utf-8-sig"
                        )
                    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                    _dl_col, _log_col = st.columns([3, 1])
                    _dl_col.download_button(
                        label=f"⬇️ Download do CSV Correlacionado ({_n_filtrado:,} linhas)",
                        data=st.session_state[_csv_dl_key].encode("utf-8-sig"),
                        file_name=f"correlacao_endoscopia_{ts}.csv",
                        mime="text/csv",
                        type="primary",
                    )
                    # ── Log de diagnóstico TUSS (útil em Docker) ─────────────
                    _tuss_log_path = _TUSS_DIR / "tuss_debug.log"
                    if _tuss_log_path.exists():
                        _log_bytes = _tuss_log_path.read_bytes()
                        _log_col.download_button(
                            label="🪲 Log TUSS",
                            data=_log_bytes,
                            file_name="tuss_debug.log",
                            mime="text/plain",
                            help="Log de diagnóstico da verificação TUSS — gerado durante a correlação",
                        )
                    
                    # ── Botão de Recarregar no Supabase ──────────────────────
                    st.divider()
                    if not is_admin():
                        st.caption("🔒 Recarregar dados no Supabase — apenas Admin")
                    if is_admin() and st.button("🔄 Recarregar Dados no Supabase", type="secondary", help="Carregar para a tabela correlacao_endoscopia"):
                        try:
                            from supabase import create_client

                            # Configuração do Supabase
                            supabase_url = os.getenv("SUPABASE_URL")
                            # Prefere secret key (bypassa RLS); cai para publishable com aviso
                            supabase_key = (
                                os.getenv("SUPABASE_SECRET_KEY")
                                or os.getenv("SUPABASE_PUBLISHABLE_KEY")
                            )
                            _usando_secret_key = bool(os.getenv("SUPABASE_SECRET_KEY"))

                            if not supabase_url or not supabase_key:
                                st.error(
                                    "❌ Variáveis de ambiente não configuradas. "
                                    "Defina **SUPABASE_URL** e **SUPABASE_SECRET_KEY** "
                                    "(secret key do projeto Supabase)."
                                )
                            elif not _usando_secret_key:
                                st.error(
                                    "❌ **SUPABASE_SECRET_KEY** não configurada. "
                                    "A chave publishable não tem permissão para inserir dados (RLS). "
                                    "Adicione a variável `SUPABASE_SECRET_KEY` com a **secret key** "
                                    "encontrada em: Supabase → Project Settings → API Keys → Secret keys."
                                )
                            else:
                                with st.spinner("🔄 Conectando ao Supabase..."):
                                    supabase = create_client(supabase_url, supabase_key)
                                
                                # Valida se a tabela existe
                                with st.spinner("🔍 Validando tabela..."):
                                    try:
                                        supabase.table("correlacao_endoscopia").select("*").limit(1).execute()
                                    except Exception as e:
                                        st.error(f"❌ Tabela 'correlacao_endoscopia' não existe ou não está acessível: {e}")
                                        st.stop()
                                
                                # ── Define identificador do novo lote ────────────────
                                lote_novo = datetime.now().strftime("%Y%m%d_%H%M%S")

                                # ── Prepara registros ────────────────────────────────
                                # Usa df_metrics (df_final já deduplicado) para garantir
                                # que Streamlit e frontend exibam os mesmos números.
                                # Duplicatas intra-lote do arquivo fonte já foram removidas.
                                # O trigger BEFORE INSERT do banco cuida de:
                                #   • hash_conteudo, is_duplicata, id_original
                                #   • carry-over de decisao_humana/revisado_em/notas_revisor
                                #   • versionamento: desativa versão anterior do mesmo
                                #     ChaveCorrelacao (se conteúdo mudou entre lotes)
                                # Campos de soft-delete e controle de lote também são
                                # exclusividade do banco — não enviar pelo app.py.
                                _COLUNAS_DB = {
                                    "id", "hash_conteudo", "is_duplicata", "id_original",
                                    "criado_em", "decisao_humana", "revisado_em", "notas_revisor",
                                    # soft-delete (rollback-migration.sql)
                                    "ativo", "desativado_em", "desativado_por",
                                    "motivo_desativacao", "rollback_operacao_id",
                                }
                                _records_raw = df_metrics \
                                    .where(df_metrics.notna(), other=None) \
                                    .to_dict("records")
                                _DATE_COLS = {"Data_PRODUCAO", "Data_REPASSE"}

                                def _fmt_date_iso(v):
                                    if not v or not isinstance(v, str):
                                        return v
                                    v = v.strip()
                                    if not v:
                                        return None
                                    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
                                        try:
                                            return datetime.strptime(v, fmt).strftime("%Y-%m-%d")
                                        except ValueError:
                                            pass
                                    return None  # valor não reconhecido como data → NULL

                                def _fmt_pct(v):
                                    """Remove ' %' de valores como '100 %' → 100.0 (float) para colunas numeric."""
                                    if not v or not isinstance(v, str):
                                        return v
                                    v = v.strip()
                                    if not v:
                                        return None
                                    cleaned = v.replace('%', '').strip()
                                    try:
                                        return float(cleaned)
                                    except ValueError:
                                        return None

                                _PCT_COLS = {"Porcentagem_REPASSE"}

                                records = []
                                for _r in _records_raw:
                                    _row = {k: v for k, v in _r.items() if k not in _COLUNAS_DB}
                                    for _dc in _DATE_COLS:
                                        if _dc in _row:
                                            _row[_dc] = _fmt_date_iso(_row[_dc])
                                    for _pc in _PCT_COLS:
                                        if _pc in _row:
                                            _row[_pc] = _fmt_pct(_row[_pc])
                                    _row["lote_processamento"] = lote_novo
                                    records.append(_row)

                                # ── 1. Registra lote como "iniciado" em lotes_carga ──
                                # Checkpoint inicial: permite ao frontend identificar
                                # cargas incompletas mesmo que o INSERT de dados falhe.
                                supabase \
                                    .table("lotes_carga") \
                                    .insert({
                                        "lote_id":        lote_novo,
                                        "status":         "iniciado",
                                        "total_inserido": len(records),
                                    }) \
                                    .execute()

                                try:
                                    # ── 2. INSERT dados em correlacao_endoscopia ──────
                                    with st.spinner(f"📤 Inserindo {len(records):,} registros (lote {lote_novo})..."):
                                        _batch_size = 500
                                        _n_batches  = -(-len(records) // _batch_size)
                                        for _i in range(0, len(records), _batch_size):
                                            _batch = records[_i:_i + _batch_size]
                                            supabase \
                                                .table("correlacao_endoscopia") \
                                                .upsert(
                                                    _batch,
                                                    on_conflict="lote_processamento,hash_conteudo",
                                                    ignore_duplicates=True,
                                                ) \
                                                .execute()
                                            st.progress(
                                                (_i + len(_batch)) / len(records),
                                                f"Inserindo lote {_i // _batch_size + 1}/{_n_batches}"
                                            )

                                    # ── 3. Apura contagens reais no banco ─────────────
                                    _total_lote    = len(records)
                                    _total_dup     = 0
                                    _total_validos = len(records)
                                    try:
                                        _cnt_resp = supabase \
                                            .table("correlacao_endoscopia") \
                                            .select("is_duplicata", count="exact") \
                                            .eq("lote_processamento", lote_novo) \
                                            .execute()
                                        _total_lote = _cnt_resp.count or len(records)
                                        _cnt_dup_resp = supabase \
                                            .table("correlacao_endoscopia") \
                                            .select("id", count="exact") \
                                            .eq("lote_processamento", lote_novo) \
                                            .eq("is_duplicata", True) \
                                            .execute()
                                        _total_dup     = _cnt_dup_resp.count or 0
                                        _total_validos = _total_lote - _total_dup
                                    except Exception:
                                        pass  # contagens informativas — não bloquear

                                    # ── 4. Atualiza lotes_carga → "ativo" ─────────────
                                    supabase \
                                        .table("lotes_carga") \
                                        .update({
                                            "status":           "ativo",
                                            "total_inserido":   _total_lote,
                                            "total_validos":    _total_validos,
                                            "total_duplicatas": _total_dup,
                                        }) \
                                        .eq("lote_id", lote_novo) \
                                        .execute()

                                    st.success(
                                        f"✅ **{len(records):,} registros** carregados com sucesso!\n\n"
                                        f"🏷️ Lote: `{lote_novo}`"
                                    )

                                except Exception as _e_insert:
                                    # ── Erro durante INSERT: marca lote como "erro" ───
                                    # Os dados parcialmente inseridos ficam no banco com
                                    # lote_processamento = lote_novo e podem ser removidos
                                    # pelo frontend ao invalidar/excluir este lote.
                                    try:
                                        supabase \
                                            .table("lotes_carga") \
                                            .update({"status": "erro"}) \
                                            .eq("lote_id", lote_novo) \
                                            .execute()
                                    except Exception:
                                        pass
                                    st.error(f"❌ Erro ao recarregar dados: {_e_insert}")
                                    st.warning(
                                        f"⚠️ O lote `{lote_novo}` foi marcado como **erro** em `lotes_carga`. "
                                        "Use o frontend para remover a carga incompleta antes de tentar novamente."
                                    )
                                    logger.error(f"Erro INSERT Supabase lote {lote_novo}: {_e_insert}", exc_info=True)

                        except ImportError:
                            st.error("❌ Biblioteca supabase-py não instalada. Execute: pip install supabase")
                        except Exception as e:
                            st.error(f"❌ Erro ao recarregar dados: {e}")
                            logger.error(f"Erro ao recarregar Supabase: {e}", exc_info=True)

                else:
                    st.warning("⚠️ Não foi possível renderizar o DataFrame. Exibindo CSV bruto.")
                    st.text_area("CSV bruto", value=csv_dados_str, height=300)
                    st.download_button(
                        label="⬇️ Download do CSV Bruto",
                        data=csv_dados_str.encode("utf-8-sig"),
                        file_name=f"correlacao_bruta_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
                        mime="text/csv",
                    )

                if resumo_str:
                    st.divider()
                    st.subheader("📋 Resumo — Divergências por Convênio")
                    st.markdown(resumo_str)

    # ── TAB 5: GERAR COBRANÇA (financeiro / admin) ───────────────────────────
    with tab_cobranca:
        st.header("📋 Gerar Formulário de Cobrança")

        if not can_view_financial():
            st.warning(
                "🔒 **Acesso restrito** — esta aba está disponível apenas para os perfis "
                "**Financeiro** e **Admin**."
            )
            st.stop()

        st.markdown(
            "Gera o **Formulário para Solicitação de Revisão de Procedimentos Não Repassados** "
            "preenchido automaticamente com os casos identificados na correlação."
        )

        csv_corr_raw = st.session_state.get("csv_correlacionado")
        if not csv_corr_raw:
            st.info("⬅️ Execute a correlação na aba **🔀 Correlação** primeiro.")
        else:
            from io import BytesIO
            from openpyxl.styles import PatternFill, Font

            # Carregar dados de correlação — cacheado para não re-parsear em cada interação
            _cob_df_key = f"cob_df_{hash(csv_corr_raw[:300])}"
            if _cob_df_key not in st.session_state:
                _csv_limpo = extrair_csv_do_texto(csv_corr_raw)
                _csv_dados, _ = separar_resumo_do_csv(_csv_limpo)
                _df_tmp = texto_para_dataframe(_csv_dados)
                if _df_tmp is not None and not _df_tmp.empty:
                    # Pré-computar colunas de data vetorizadas uma única vez
                    _df_tmp["_dt_prod"] = pd.to_datetime(
                        _df_tmp.get("Data_PRODUCAO", pd.Series(dtype=str)),
                        dayfirst=True, errors="coerce",
                    )
                    _df_tmp["_ano_prod"] = _df_tmp["_dt_prod"].dt.year.astype("Int64")
                    _df_tmp["_mes_prod"] = _df_tmp["_dt_prod"].dt.month.astype("Int64")
                st.session_state[_cob_df_key] = _df_tmp
            df_corr = st.session_state[_cob_df_key]
            if df_corr is None or df_corr.empty:
                st.error("Não foi possível carregar os dados de correlação.")
                st.stop()

            valores_tuss = _carregar_valores_tuss()
            if not valores_tuss and not df_corr.empty:
                try:
                    with st.spinner("Calculando estimativas de valor a partir do histórico de repasse..."):
                        _gerar_valores_tuss(df_corr)
                    _carregar_valores_tuss.clear()
                    valores_tuss = _carregar_valores_tuss()
                    if valores_tuss:
                        logger.info(f"tuss_valores.csv gerado com {len(valores_tuss)} entradas")
                    else:
                        logger.warning("_gerar_valores_tuss executou mas retornou vazio — verifique os dados de repasse")
                except Exception as _e_gen_tuss:
                    st.error(f"Erro ao gerar estimativas de valor TUSS: {_e_gen_tuss}")
                    logger.exception("_gerar_valores_tuss falhou na aba GERAR COBRANÇA")

            desc_por_cod  = _construir_desc_por_codigo(
                pd.read_csv(BytesIO(st.session_state.get("csv_repasse_raw", b"")))
                if st.session_state.get("csv_repasse_raw") else pd.DataFrame()
            )

            # ── Descricao de código: REPASSE real > tabela DESC_OFICIAL ───────
            _DESC_OFICIAL_COB = {
                "40201074": "Colangiopancreatografia Retrógrada Endoscópica",
                "40201082": "Colonoscopia (Inclui A Retossigmoidoscopia)",
                "40201120": "Endoscopia Digestiva Alta",
                "40201171": "Retossigmoidoscopia Flexível",
                "40202038": "Endoscopia Digestiva Alta Com Biópsia E/Ou Citologia",
                "40202283": "Gastrostomia Endoscópica",
                "40202291": "Hemostasia Mecânica Do Esôfago, Estômago Ou Duodeno",
                "40202313": "Hemostasias De Cólon",
                "40202453": "Ligadura Elástica Do Esôfago, Estômago Ou Duodeno",
                "40202470": "Mucosectomia Do Esôfago, Estômago Ou Duodeno",
                "40202534": "Passagem De Sonda Naso-Enteral",
                "40202542": "Polipectomia De Cólon (Independente Do Número De Pólipos)",
                "40202550": "Polipectomia Do Esôfago, Estômago Ou Duodeno (Independente Do Número De Pólipos)",
                "40202577": "Retirada De Corpo Estranho Do Esôfago, Estômago Ou Duodeno",
                "40202615": "Endoscopia Digestiva Alta Com Biópsia E Teste De Urease (Pesquisa Helicobacter Pylori)",
                "40202666": "Colonoscopia Com Biópsia E/Ou Citologia",
                "40202682": "Retossigmoidoscopia Flexível Com Polipectomia",
                "40202690": "Retossigmoidoscopia Flexível Com Biópsia E/Ou Citologia",
                "40202712": "Colonoscopia Com Mucosectomia",
                "40813320": "Colocação De Stent Biliar",
            }

            # Reverse-map de descrições canônicas a partir do tuss_lookup_table
            _desc_tuss_lookup = _construir_desc_por_tuss_code(_carregar_tabela_tuss())

            def _desc_cod(cod: str) -> str:
                cod = str(cod).replace(".0", "").strip()
                return (desc_por_cod.get(cod)           # 1º: REPASSE real
                        or _desc_tuss_lookup.get(cod)   # 2º: tuss_lookup_table ← novo
                        or _DESC_OFICIAL_COB.get(cod)   # 3º: dict hardcoded
                        or f"Código TUSS {cod}")

            def _lookup_entry(cod: str, convenio: str = "") -> dict:
                """Busca stats do código TUSS: primeiro específico por convênio, depois GERAL."""
                c = str(cod).replace(".0", "").strip()
                conv_norm = _normalizar_convenio(convenio) if convenio else ""
                if conv_norm:
                    entry = valores_tuss.get((conv_norm, c), {})
                    if entry:
                        return entry
                return valores_tuss.get(c, {})   # fallback GERAL

            def _valor_estimado(
                cod_esperado: str,
                cod_pago: str | None,
                tipo_origem: str,
                convenio: str = "",
            ) -> tuple[float | None, str]:
                """Retorna (valor_estimado, confiança) usando o valor mais recente por plano."""
                ce = str(cod_esperado).replace(".0", "").strip()
                cp = str(cod_pago or "").replace(".0", "").strip()

                def _val_base(entry: dict) -> float | None:
                    """Retorna UltimoValor se disponível, senão Media."""
                    v = entry.get("UltimoValor") or entry.get("Media")
                    return float(v) if v is not None else None

                entry_esp  = _lookup_entry(ce, convenio)
                conf       = str(entry_esp.get("Confianca", "Sem dados"))

                if tipo_origem == "downgrade" and cp:
                    entry_pago = _lookup_entry(cp, convenio)
                    v_esp  = _val_base(entry_esp)
                    v_pago = _val_base(entry_pago)
                    if v_esp is not None and v_pago is not None:
                        return round(v_esp - v_pago, 2), conf
                elif tipo_origem in ("ausente", "nao_faturado"):
                    v = _val_base(entry_esp)
                    if v is not None:
                        return round(v, 2), conf
                return None, "Sem dados"

            # ── Seção 1: Filtros ──────────────────────────────────────────────
            st.subheader("1. Filtros")

            # Montar mapa {ano: [meses ordenados]} — usa coluna _dt_prod já pré-computada
            _NOMES_MESES = {1:"Jan",2:"Fev",3:"Mar",4:"Abr",5:"Mai",6:"Jun",
                            7:"Jul",8:"Ago",9:"Set",10:"Out",11:"Nov",12:"Dez"}
            _dt_col_cob = df_corr.get("_dt_prod", pd.Series(dtype="datetime64[ns]")).dropna()
            _periodos_disp: dict[int, list[int]] = {}
            for _dt in _dt_col_cob:
                _periodos_disp.setdefault(int(_dt.year), set()).add(int(_dt.month))
            _periodos_disp = {a: sorted(m) for a, m in sorted(_periodos_disp.items())}

            # Inicializar session_state para cada checkbox (default: marcado)
            for _ano, _meses in _periodos_disp.items():
                for _mes in _meses:
                    _k = f"cob_ref_{_ano}_{_mes}"
                    if _k not in st.session_state:
                        st.session_state[_k] = True

            # Callbacks para Todos/Nenhum por ano
            def _marcar_ano_cob(ano_cb: int, valor: bool) -> None:
                for _m in _periodos_disp.get(ano_cb, []):
                    st.session_state[f"cob_ref_{ano_cb}_{_m}"] = valor

            st.markdown("**Período de referência**")
            if not _periodos_disp:
                st.warning("Não foi possível identificar datas na coluna Data_PRODUCAO.")
            for _ano, _meses in _periodos_disp.items():
                _ca, _cb1, _cb2 = st.columns([5, 1, 1])
                _ca.markdown(f"**{_ano}**")
                _cb1.button("✅ Todos",  key=f"cob_btn_todos_{_ano}",
                            on_click=_marcar_ano_cob, args=(_ano, True))
                _cb2.button("☐ Nenhum", key=f"cob_btn_nenhum_{_ano}",
                            on_click=_marcar_ano_cob, args=(_ano, False))
                _cols_mes = st.columns(6)
                for _i, _mes in enumerate(_meses):
                    _cols_mes[_i % 6].checkbox(
                        _NOMES_MESES[_mes], key=f"cob_ref_{_ano}_{_mes}"
                    )

            # Conjunto de (ano, mês) selecionados
            filtro_periodos: set[tuple[int, int]] = {
                (_ano, _mes)
                for _ano, _meses in _periodos_disp.items()
                for _mes in _meses
                if st.session_state.get(f"cob_ref_{_ano}_{_mes}", True)
            }

            st.divider()
            fc3, fc4, fc5 = st.columns(3)
            inc_downgrade   = fc3.checkbox("🔴 Cobrado Como Simples",  value=True, key="cob_dg",
                help="COBRAR_TUSS_PROC_ADICIONAL_COBRADO_COMO_SIMPLES — convênio usou código menor, ignorando o adicional")
            inc_ausente     = fc4.checkbox("🟠 Código Adicional Ausente", value=True, key="cob_aus",
                help="COBRAR_TUSS_CODIGO_ADICIONAL_AUSENTE_NO_REPASSE — código separado (ex.: Urease) não faturado")
            inc_nao_faturado = fc5.checkbox("❌ Procedimento Não Faturado", value=True, key="cob_nf",
                help="NAO_FATURADO_NO_REPASSE — procedimento inteiro ausente do repasse")

            # ── Montar itens de cobrança ──────────────────────────────────────
            # Cache do loop pesado: chave = hash dos dados + flags de tipo + estado valores_tuss
            # O sufixo _vtuss (0/1) força rebuild quando valores_tuss transita de vazio → populado
            _itens_key = f"cob_itens_{hash(csv_corr_raw[:300])}_{inc_downgrade}_{inc_ausente}_{inc_nao_faturado}_{int(bool(valores_tuss))}"
            if _itens_key not in st.session_state:
                _todos_itens: list[dict] = []
                # Carregar tabela TUSS uma única vez (fora do loop)
                _tabela_tuss_loop = _carregar_tabela_tuss()
                for _idx_row, _row in df_corr.iterrows():
                    _conv = str(_row.get("Convenio_PRODUCAO", "")).strip()
                    _ano_r = _row.get("_ano_prod")
                    _mes_r = _row.get("_mes_prod")
                    _ano_i = int(_ano_r) if pd.notna(_ano_r) else None
                    _mes_i = int(_mes_r) if pd.notna(_mes_r) else None
                    _sc   = str(_row.get("StatusCorrelacao", "")).upper()
                    _st_t = str(_row.get("StatusTUSS", ""))

                    if inc_downgrade and _st_t == "COBRAR_TUSS_PROC_ADICIONAL_COBRADO_COMO_SIMPLES":
                        _cod_esp  = str(_row.get("CodigosTUSS_Esperados", "")).strip()
                        _cod_pago = str(_row.get("CodigoTUSS_REPASSE", "")).replace(".0", "").strip()
                        _val, _conf = _valor_estimado(_cod_esp, _cod_pago, "downgrade", convenio=_conv)
                        _obs = (f"Faturado como {_cod_pago} — {_desc_cod(_cod_pago)}. "
                                f"Correto conforme TUSS: {_cod_esp} — {_desc_cod(_cod_esp)}. "
                                f"Solicita revisão e reprocessamento.")
                        _todos_itens.append({"_origem": "downgrade", "_conf": _conf,
                            "_ano": _ano_i, "_mes": _mes_i,
                            "DATA": _row.get("Data_PRODUCAO", ""),
                            "NR_ATEND": _row.get("NrAtendimento_REPASSE", ""),
                            "PACIENTE": _row.get("Paciente_PRODUCAO", ""),
                            "CONVENIO": _conv,
                            "PRESTADOR": _row.get("MedicoExecutor_PRODUCAO", ""),
                            "CODIGO": _cod_esp,
                            "PROCEDIMENTO": _desc_cod(_cod_esp),
                            "FUNCAO": _row.get("Funcao_REPASSE", "Cirurgiao"),
                            "OBSERVACAO": _obs,
                            "VALOR": _val,
                        })
                    elif inc_ausente and _st_t == "COBRAR_TUSS_CODIGO_ADICIONAL_AUSENTE_NO_REPASSE":
                        _ausentes_str = str(_row.get("CodigosTUSS_Ausentes", _row.get("CodigosTUSS_Esperados", ""))).strip()
                        for _cod_aus in [c.strip() for c in _ausentes_str.split(",") if c.strip()]:
                            _val, _conf = _valor_estimado(_cod_aus, None, "ausente", convenio=_conv)
                            _proc_adic = str(_row.get("ProcedimentosAdicionais_PRODUCAO", "")).strip()
                            _nr_base   = str(_row.get("NrAtendimento_REPASSE", "")).strip()
                            _obs = (f"Procedimento adicional '{_proc_adic}' realizado junto ao "
                                    f"{_row.get('Procedimento_PRODUCAO','')} não consta no repasse. "
                                    f"Código TUSS: {_cod_aus}. "
                                    f"Nr atendimento base: {_nr_base}. Solicita inclusão.")
                            _todos_itens.append({"_origem": "ausente", "_conf": _conf,
                                "_ano": _ano_i, "_mes": _mes_i,
                                "DATA": _row.get("Data_PRODUCAO", ""),
                                "NR_ATEND": _nr_base,
                                "PACIENTE": _row.get("Paciente_PRODUCAO", ""),
                                "CONVENIO": _conv,
                                "PRESTADOR": _row.get("MedicoExecutor_PRODUCAO", ""),
                                "CODIGO": _cod_aus,
                                "PROCEDIMENTO": _desc_cod(_cod_aus),
                                "FUNCAO": _row.get("Funcao_REPASSE", "Cirurgiao"),
                                "OBSERVACAO": _obs,
                                "VALOR": _val,
                            })
                    elif inc_nao_faturado and _sc == "NAO_FATURADO_NO_REPASSE":
                        _proc_p = str(_row.get("Procedimento_PRODUCAO", "")).strip()
                        _chave_base = _normalizar_chave_tuss(f"{_proc_p}_")
                        _entry_base = _tabela_tuss_loop.get(_chave_base, {})
                        _cod_base = str(_entry_base.get("CodigosTUSS", "")).split(",")[0].strip() if _entry_base else ""
                        _val, _conf = _valor_estimado(_cod_base, None, "nao_faturado", convenio=_conv) if _cod_base else (None, "Sem dados")
                        _obs = (f"Procedimento {_proc_p} realizado em {_row.get('Data_PRODUCAO','')} "
                                f"não localizado no repasse. Solicita inclusão e faturamento.")
                        _todos_itens.append({"_origem": "nao_faturado", "_conf": _conf,
                            "_ano": _ano_i, "_mes": _mes_i,
                            "DATA": _row.get("Data_PRODUCAO", ""),
                            "NR_ATEND": _row.get("NrAtendimento_PRODUCAO", ""),
                            "PACIENTE": _row.get("Paciente_PRODUCAO", ""),
                            "CONVENIO": _conv,
                            "PRESTADOR": _row.get("MedicoExecutor_PRODUCAO", ""),
                            "CODIGO": _cod_base,
                            "PROCEDIMENTO": _desc_cod(_cod_base) if _cod_base else _proc_p,
                            "FUNCAO": "Cirurgiao",
                            "OBSERVACAO": _obs,
                            "VALOR": _val,
                        })
                st.session_state[_itens_key] = _todos_itens

            # Filtro de período aplicado como máscara barata sobre o cache
            _todos_itens = st.session_state[_itens_key]
            itens = [
                i for i in _todos_itens
                if (i["_ano"], i["_mes"]) in filtro_periodos
                or i["_ano"] is None
            ]

            # ── Seção 2: Prévia ───────────────────────────────────────────────
            st.subheader("2. Prévia dos itens selecionados")
            if not itens:
                st.info("Nenhum item encontrado com os filtros selecionados.")
            else:
                df_prev = pd.DataFrame(itens)
                n_baixa_conf = (df_prev["_conf"].isin(["Baixa", "Sem dados"])).sum()
                col_res1, col_res2, col_res3 = st.columns(3)
                col_res1.metric("Total de linhas", len(df_prev))
                col_res2.metric("🔴 Cobrado Como Simples", int((df_prev["_origem"] == "downgrade").sum()))
                col_res3.metric("🟠 Código Adicional Ausente + Não Faturado",
                    int((df_prev["_origem"].isin(["ausente", "nao_faturado"])).sum()))

                colunas_prev = ["DATA", "PACIENTE", "CONVENIO", "CODIGO", "PROCEDIMENTO", "VALOR", "OBSERVACAO"]
                _pb_col, _ = st.columns([2, 8])
                if not st.session_state.get("cob_show_previa", False):
                    if _pb_col.button("👁️ Exibir prévia", key="btn_show_cob_previa", type="primary",
                                      help=f"{len(df_prev):,} linhas disponíveis"):
                        st.session_state["cob_show_previa"] = True
                        st.rerun()
                else:
                    if _pb_col.button("🔒 Ocultar prévia", key="btn_hide_cob_previa"):
                        st.session_state["cob_show_previa"] = False
                        st.rerun()
                    st.dataframe(
                        df_prev[[c for c in colunas_prev if c in df_prev.columns]],
                        use_container_width=True, height=300,
                    )

                # ── Seção 3: Cabeçalho do formulário ─────────────────────────
                st.subheader("3. Cabeçalho do formulário")
                hc1, hc2, hc3 = st.columns(3)
                campo_empresa = hc1.text_input("Empresa / Prestador", value="ENDOPRIME SERVICOS MEDICOS", key="cob_empresa")
                campo_medico  = hc2.text_input("Médico Responsável", value="", key="cob_medico")
                campo_data    = hc3.date_input("Data do formulário", value=datetime.today(), key="cob_data")

                # ── Seção 4: Estimativa de valor ─────────────────────────────
                st.subheader("4. Estimativa de valor")
                if not valores_tuss:
                    st.info("ℹ️ Sem dados de valor disponíveis. Execute a correlação para gerar estimativas por convênio.")
                    estimar_valor = False
                else:
                    # Códigos presentes nos itens selecionados
                    _codigos_itens = {str(i.get("CODIGO", "")).strip() for i in itens if i.get("CODIGO")}
                    # Montar tabela de referência direto do dict valores_tuss (já em memória)
                    _ref_rows: list[dict] = []
                    for _vk, _vstats in valores_tuss.items():
                        if isinstance(_vk, tuple):
                            _vconv, _vcod = _vk
                        else:
                            _vconv, _vcod = "GERAL", _vk
                        if _codigos_itens and _vcod not in _codigos_itens:
                            continue
                        _ref_rows.append({
                            "CodigoTUSS":  _vcod,
                            "Descricao":   _vstats.get("Descricao") or _desc_cod(_vcod),
                            "Convenio":    _vconv,
                            "UltimoValor": _vstats.get("UltimoValor"),
                            "Media":       _vstats.get("Media"),
                            "Qtd":         _vstats.get("Qtd", 0),
                            "Confianca":   _vstats.get("Confianca", ""),
                        })
                    if _ref_rows:
                        _df_disp = (
                            pd.DataFrame(_ref_rows)
                            .rename(columns={
                                "CodigoTUSS":  "Código TUSS",
                                "Descricao":   "Procedimento",
                                "Convenio":    "Convênio",
                                "UltimoValor": "Último Valor (R$)",
                                "Media":       "Média (R$)",
                                "Qtd":         "Amostras",
                                "Confianca":   "Confiança",
                            })
                            .sort_values(
                                ["Código TUSS", "Convênio"],
                                key=lambda s: s.where(s != "GERAL", "ZZZZ"),
                            )
                        )
                        _eb_col, _dl_col, _ = st.columns([2, 2, 6])
                        # Botão de download sempre visível (independente de exibir/ocultar)
                        _csv_estimativa = _df_disp.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")
                        _dl_col.download_button(
                            label="⬇️ Download da Estimativa",
                            data=_csv_estimativa,
                            file_name="estimativa_tuss.csv",
                            mime="text/csv",
                            key="btn_dl_estimativa",
                            help=f"Baixar tabela de referência com {len(_df_disp):,} linhas (CSV)",
                        )
                        if not st.session_state.get("cob_show_estimativa", False):
                            if _eb_col.button("👁️ Exibir estimativa", key="btn_show_cob_est", type="primary",
                                              help=f"{len(_df_disp):,} linhas de referência"):
                                st.session_state["cob_show_estimativa"] = True
                                st.rerun()
                        else:
                            if _eb_col.button("🔒 Ocultar estimativa", key="btn_hide_cob_est"):
                                st.session_state["cob_show_estimativa"] = False
                                st.rerun()
                            st.dataframe(
                                _df_disp,
                                use_container_width=True,
                                hide_index=True,
                                height=min(35 * len(_df_disp) + 40, 320),
                                column_config={
                                    "Último Valor (R$)": st.column_config.NumberColumn(format="R$ %.2f"),
                                    "Média (R$)":        st.column_config.NumberColumn(format="R$ %.2f"),
                                },
                            )
                    else:
                        st.caption("Nenhum código dos itens selecionados encontrado nos dados de valor TUSS.")

                    if n_baixa_conf > 0:
                        st.caption(f"⚠️ {n_baixa_conf} item(ns) com confiança Baixa ou Sem dados — células destacadas em amarelo no XLSX.")
                    estimar_valor = st.checkbox(
                        "Incluir coluna VALOR no formulário XLSX",
                        value=True,
                        key="cob_estimar",
                        help=(
                            "Preenche a coluna VALOR do formulário com o Último Valor registrado por convênio "
                            "(ou Média GERAL como fallback). Downgrade: diferença entre código esperado e pago. "
                            "Ausente/Não Faturado: valor integral do código ausente."
                        ),
                    )

                # ── Seção 5: Gerar XLSX ───────────────────────────────────────
                st.subheader("5. Gerar formulário")
                if st.button("⚙️ Gerar Formulário XLSX", type="primary", key="cob_gerar"):
                    try:
                        wb = _criar_workbook_formulario_template()
                        ws = wb["Plan1"]

                        # Preencher campos de cabeçalho variáveis
                        ws["A7"] = campo_empresa
                        if campo_medico:
                            ws["D7"] = campo_medico
                        ws["G7"] = f"DATA: {campo_data.strftime('%d/%m/%Y')}"

                        # Ordenar itens por DATA antes de escrever no XLSX
                        _df_sort = pd.DataFrame(itens)
                        _df_sort["_dt"] = pd.to_datetime(
                            _df_sort["DATA"].astype(str), dayfirst=True, errors="coerce"
                        )
                        _df_sort = _df_sort.sort_values("_dt", na_position="last").drop(columns=["_dt"])
                        itens_xlsx = _df_sort.to_dict("records")

                        # Estilos
                        _fill_baixa_conf = PatternFill("solid", fgColor="FFF3CD")
                        _font_normal = Font(name="Arial", size=9)

                        for i, item in enumerate(itens_xlsx):
                            row_num = 11 + i
                            ws.cell(row_num, 1).value = item["DATA"]
                            ws.cell(row_num, 2).value = item["NR_ATEND"]
                            ws.cell(row_num, 3).value = str(item["PACIENTE"]).title()
                            ws.cell(row_num, 4).value = item["CONVENIO"]
                            ws.cell(row_num, 5).value = item["PRESTADOR"]
                            ws.cell(row_num, 6).value = item["CODIGO"]
                            ws.cell(row_num, 7).value = item["PROCEDIMENTO"]
                            ws.cell(row_num, 8).value = item["FUNCAO"]
                            ws.cell(row_num, 9).value = item["OBSERVACAO"]

                            if estimar_valor:
                                val = item.get("VALOR")
                                conf = item.get("_conf", "Sem dados")
                                ws.cell(row_num, 10).value = val
                                if conf in ("Baixa", "Sem dados"):
                                    ws.cell(row_num, 10).fill = _fill_baixa_conf
                            # Fonte uniforme
                            for col in range(1, 11):
                                ws.cell(row_num, col).font = _font_normal

                        # Salvar em buffer
                        buf = BytesIO()
                        wb.save(buf)
                        buf.seek(0)

                        _anos_sel = sorted({p[0] for p in filtro_periodos})
                        _periodo_slug = "_".join(str(a) for a in _anos_sel)[:20] if _anos_sel else "todos"
                        ts_now = datetime.now().strftime("%Y%m%d_%H%M%S")
                        fname = f"formulario_cobranca_{_periodo_slug}_{ts_now}.xlsx"

                        st.success(f"Formulário gerado: {len(itens_xlsx)} linhas")
                        st.download_button(
                            label="⬇️ Baixar Formulário XLSX",
                            data=buf.getvalue(),
                            file_name=fname,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary",
                        )
                    except Exception as _e_cob:
                        st.error(f"Erro ao gerar formulário: {_e_cob}")
                        logger.error(f"Erro gerar_cobranca: {_e_cob}", exc_info=True)

    # ── TAB 6: AGENTES ────────────────────────────────────────────────────────
    with tab_agents:
        st.header("🤖 Configuração dos Agentes")
        st.markdown(
            "Visualize e edite os prompts de cada agente e suas tasks. "
            "As alterações ficam ativas **durante esta sessão** e são aplicadas "
            "imediatamente nas próximas execuções.\n\n"
            "> 💡 Use o botão **↩️ Restaurar** para voltar ao texto original a qualquer momento."
        )

        # Indicador de estado atual
        analista_modificado = st.session_state.get("analista_cfg") != ANALISTA_DEFAULTS
        correlacionador_modificado = st.session_state.get("correlacionador_cfg") != CORRELACIONADOR_DEFAULTS

        col_ind1, col_ind2, col_ind3 = st.columns([2, 2, 3])
        with col_ind1:
            if analista_modificado:
                st.warning("✏️ Analista: configuração **personalizada**")
            else:
                st.success("✅ Analista: configuração **padrão**")
        with col_ind2:
            if correlacionador_modificado:
                st.warning("✏️ Correlacionador: configuração **personalizada**")
            else:
                st.success("✅ Correlacionador: configuração **padrão**")
        with col_ind3:
            if analista_modificado or correlacionador_modificado:
                if st.button("↩️ Restaurar TODOS os agentes ao padrão", type="secondary"):
                    st.session_state["analista_cfg"] = dict(ANALISTA_DEFAULTS)
                    st.session_state["correlacionador_cfg"] = dict(CORRELACIONADOR_DEFAULTS)
                    st.rerun()

        st.divider()

        agent_tab1, agent_tab2 = st.tabs([
            "🔍 Analista de Endoscopia",
            "🔀 Correlacionador",
        ])

        with agent_tab1:
            render_agent_form(
                agent_key="analista_cfg",
                defaults=ANALISTA_DEFAULTS,
                title="Analista de Endoscopia",
                icon="🔍",
            )
            st.divider()
            with st.expander("ℹ️ Sobre o placeholder `{conteudo_arquivo}`", expanded=False):
                st.markdown("""
O campo **Task Description** do Analista usa o placeholder `{conteudo_arquivo}`.

Durante a execução, ele é substituído automaticamente pelo conteúdo extraído de cada arquivo enviado na aba **📄 Input**.

**Não remova este placeholder** — sem ele, o agente não receberá os dados do arquivo.
                """)

        with agent_tab2:
            render_agent_form(
                agent_key="correlacionador_cfg",
                defaults=CORRELACIONADOR_DEFAULTS,
                title="Correlacionador",
                icon="🔀",
            )
            st.divider()
            with st.expander("ℹ️ Sobre o placeholder `{blocos}`", expanded=False):
                st.markdown("""
O campo **Task Description** do Correlacionador usa o placeholder `{blocos}`.

Durante a execução da correlação, ele é substituído automaticamente pelos CSVs gerados pelo Analista para cada arquivo processado.

**Não remova este placeholder** — sem ele, o agente não receberá os dados para o batimento.
                """)


if __name__ == "__main__":
    main()